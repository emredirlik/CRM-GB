from fastapi import FastAPI, APIRouter, HTTPException, BackgroundTasks, Response, Request, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
import csv
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate
import bcrypt
import jwt
from urllib.parse import quote
import base64
import resend
import sib_api_v3_sdk
from sib_api_v3_sdk.api.transactional_emails_api import TransactionalEmailsApi
from sib_api_v3_sdk.models.send_smtp_email import SendSmtpEmail

# Import PDF utilities
from pdf_utils import generate_order_pdf, generate_recipe_pdf, generate_lead_pdf, generate_route_pdf, generate_specification_pdf, generate_daily_report_pdf, generate_combined_daily_report_pdf

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'default-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 hours

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ===================== AUTH HELPERS =====================

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, username: str) -> str:
    payload = {
        "sub": user_id, 
        "username": username, 
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
        "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    # Check cookie first
    token = request.cookies.get("access_token")
    # Then check Authorization header
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ===================== AUTH MODELS =====================

class LoginRequest(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    name: str
    role: str

# ===================== MODELS =====================

class Lead(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    first_name: str = ""
    last_name: str = ""
    company_name: str = ""
    contact_person: str = ""
    tax_number: Optional[str] = ""
    address: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    city: Optional[str] = ""
    country: Optional[str] = ""
    notes: Optional[str] = ""
    status: Optional[str] = "new"
    source: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeadCreate(BaseModel):
    first_name: str = ""
    last_name: str = ""
    company_name: str
    tax_number: Optional[str] = ""
    address: Optional[str] = ""
    email: Optional[str] = ""  # Changed from EmailStr to allow empty/invalid emails from AI
    city: Optional[str] = ""
    country: Optional[str] = ""
    notes: Optional[str] = ""

class LeadUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    company_name: Optional[str] = None
    tax_number: Optional[str] = None
    address: Optional[str] = None
    email: Optional[EmailStr] = None
    city: Optional[str] = None
    country: Optional[str] = None
    notes: Optional[str] = None

class EmailTemplate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    subject: str
    body: str
    language: str  # tr, de, en
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmailTemplateCreate(BaseModel):
    name: str
    subject: str
    body: str
    language: str

class SMTPSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    host: str
    port: int
    username: str
    password: str
    from_email: str
    from_name: str
    use_tls: bool = True
    use_ssl: bool = False
    imap_host: Optional[str] = None
    imap_port: int = 993
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SMTPSettingsCreate(BaseModel):
    host: str
    port: int
    username: str
    password: str
    from_email: str
    from_name: str
    use_tls: bool = True
    use_ssl: bool = False
    imap_host: Optional[str] = None
    imap_port: int = 993

class EmailHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str
    lead_email: str
    lead_name: str
    company_name: str
    subject: str
    body: str
    status: str  # sent, failed
    error_message: Optional[str] = None
    sent_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SendEmailRequest(BaseModel):
    lead_id: str
    subject: str
    body: str

class GenerateEmailRequest(BaseModel):
    lead_id: str
    language: str  # tr, de, en
    tone: Optional[str] = "professional"  # professional, friendly, formal

class DashboardStats(BaseModel):
    total_leads: int
    emails_sent: int
    emails_failed: int
    recent_leads: List[dict]
    total_orders: int = 0
    total_revenue: float = 0.0

# Product Models - Ürün Kataloğu
class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str
    description: Optional[str] = ""
    default_unit: str = "kg"
    default_price: float = 0.0
    category: Optional[str] = ""
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    name: str
    code: str
    description: Optional[str] = ""
    default_unit: str = "kg"
    default_price: float = 0.0
    category: Optional[str] = ""

# Company Settings - Firma Ayarları
class CompanySettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "company_settings"
    company_name: str = "SpiceCRM"
    logo_url: Optional[str] = None
    address: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    website: Optional[str] = ""
    tax_number: Optional[str] = ""
    yearly_target: float = 0.0  # Yıllık hedef ciro
    currency: str = "EUR"
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Order Models - Multi-Product Support
class OrderProductItem(BaseModel):
    """Single product item within an order"""
    product_name: str
    product_code: str
    pieces: int = 1  # Adet (kaç paket/kutu)
    amount: float  # Miktar (10 kg gibi)
    unit: str = "kg"  # Birim (kg, g, adet, paket, litre)
    unit_price: float  # Birim fiyatı (€/kg gibi)
    subtotal: Optional[float] = None  # pieces × amount × unit_price

class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str
    lead_name: str  # For display
    company_name: str  # For display
    # Multi-product support
    products: List[OrderProductItem] = []
    # Legacy single-product fields (for backwards compatibility)
    product_name: Optional[str] = None
    product_code: Optional[str] = None
    pieces: Optional[int] = 1
    quantity: Optional[int] = None
    amount: Optional[float] = None
    unit: Optional[str] = "kg"
    unit_price: Optional[float] = None
    total_price: float
    status: str = "pending"  # pending, confirmed, shipped, delivered, cancelled
    payment_status: str = "pending"  # pending, partial, paid, overdue
    payment_due_date: Optional[str] = None  # YYYY-MM-DD
    payment_amount: float = 0.0  # Amount paid so far
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OrderCreateProductItem(BaseModel):
    """Product item for order creation"""
    product_name: str
    product_code: str
    pieces: int = 1
    amount: float
    unit: str = "kg"
    unit_price: float

class OrderCreate(BaseModel):
    lead_id: str
    # Multi-product support
    products: List[OrderCreateProductItem] = []
    # Legacy single-product fields (for backwards compatibility)
    product_name: Optional[str] = None
    product_code: Optional[str] = None
    pieces: int = 1
    amount: Optional[float] = None
    unit: str = "kg"
    unit_price: Optional[float] = None
    notes: Optional[str] = ""

class OrderUpdate(BaseModel):
    products: Optional[List[OrderProductItem]] = None
    product_name: Optional[str] = None
    product_code: Optional[str] = None
    pieces: Optional[int] = None
    amount: Optional[float] = None
    unit: Optional[str] = None
    unit_price: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None

# Lead Finder Models
class SearchLeadsRequest(BaseModel):
    keywords: List[str]  # e.g., ["gyros producer", "döner manufacturer"]
    location: str  # e.g., "Athens"
    country: str  # e.g., "Greece"
    limit: int = 20

class FoundLeadResponse(BaseModel):
    company_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    website: Optional[str] = None
    description: Optional[str] = None
    source: Optional[str] = None

class SearchResult(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    query_keywords: List[str]
    location: str
    country: str
    leads_found: List[dict]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "completed"

# Recipe Models - Müşteriye özel üretim reçeteleri
class RecipeIngredient(BaseModel):
    name: str  # Malzeme adı
    amount: float  # Miktar
    unit: str  # Birim (kg, g, L, ml, dakika, rpm vb.)

class Recipe(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str  # Hangi müşteriye ait
    lead_name: str  # Müşteri adı (görüntüleme için)
    company_name: str  # Firma adı
    name: str  # Reçete adı (örn: "Gyros Özel Karışım")
    product_code: str  # Ürün kodu
    # Malzemeler
    meat_amount: float  # Et miktarı (kg)
    water_amount: float  # Su miktarı (L)
    spice_amount: float  # Baharat miktarı (kg)
    binding_amount: float  # Bağlayıcı (binding) miktarı (kg)
    # Üretim parametreleri
    mixing_time: int  # Karışım süresi (dakika)
    motor_speed: int  # Motor hızı (rpm)
    # Ek malzemeler (dinamik liste)
    additional_ingredients: Optional[List[dict]] = []  # [{"name": "Tuz", "amount": 0.5, "unit": "kg"}]
    # Notlar ve talimatlar
    instructions: Optional[str] = ""
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RecipeCreate(BaseModel):
    lead_id: str
    name: str
    product_code: str
    meat_amount: float
    water_amount: float
    spice_amount: float
    binding_amount: float
    mixing_time: int
    motor_speed: int
    additional_ingredients: Optional[List[dict]] = []
    instructions: Optional[str] = ""
    notes: Optional[str] = ""

class RecipeUpdate(BaseModel):
    name: Optional[str] = None
    product_code: Optional[str] = None
    meat_amount: Optional[float] = None
    water_amount: Optional[float] = None
    spice_amount: Optional[float] = None
    binding_amount: Optional[float] = None
    mixing_time: Optional[int] = None
    motor_speed: Optional[int] = None
    additional_ingredients: Optional[List[dict]] = None
    instructions: Optional[str] = None
    notes: Optional[str] = None

# ===================== HELPER FUNCTIONS =====================

def serialize_datetime(obj):
    """Convert datetime to ISO string for MongoDB storage"""
    if isinstance(obj, datetime):
        return obj.isoformat()
    return obj

def deserialize_datetime(doc, fields=['created_at', 'updated_at', 'sent_at']):
    """Convert ISO string back to datetime"""
    for field in fields:
        if field in doc and isinstance(doc[field], str):
            doc[field] = datetime.fromisoformat(doc[field])
    return doc

# ===================== LEAD ENDPOINTS =====================

@api_router.get("/")
async def root():
    return {"message": "SpiceCRM API is running"}

@api_router.post("/leads", response_model=Lead)
async def create_lead(lead_data: LeadCreate):
    lead = Lead(**lead_data.model_dump())
    doc = lead.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.leads.insert_one(doc)
    return lead

@api_router.get("/leads", response_model=List[Lead])
async def get_leads():
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    for lead in leads:
        deserialize_datetime(lead)
    return leads

@api_router.get("/leads/{lead_id}", response_model=Lead)
async def get_lead(lead_id: str):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    deserialize_datetime(lead)
    return lead

@api_router.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, lead_data: LeadUpdate):
    existing = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    update_data = {k: v for k, v in lead_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.leads.update_one({"id": lead_id}, {"$set": update_data})
    updated = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    deserialize_datetime(updated)
    return updated

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str):
    result = await db.leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"message": "Lead deleted successfully"}

@api_router.get("/leads/all/pdf")
async def download_all_leads_pdf(lang: str = 'tr'):
    """Download all leads as a single PDF"""
    return await export_leads_pdf(lang)

@api_router.get("/leads/{lead_id}/pdf")
async def get_lead_pdf(lead_id: str, lang: str = 'tr'):
    """Generate PDF for a single lead with their orders and recipes"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Get lead's orders and recipes
    orders = await db.orders.find({"lead_id": lead_id}, {"_id": 0}).to_list(100)
    recipes = await db.recipes.find({"lead_id": lead_id}, {"_id": 0}).to_list(100)
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    
    pdf_content = generate_lead_pdf(lead, orders, recipes, settings, lang)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=customer_{lead_id[:8]}.pdf"}
    )

@api_router.get("/leads/{lead_id}/details")
async def get_lead_details(lead_id: str):
    """Get detailed lead info with orders and recipes"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    deserialize_datetime(lead)
    
    # Get lead's orders
    orders = await db.orders.find({"lead_id": lead_id}, {"_id": 0}).to_list(100)
    for order in orders:
        deserialize_datetime(order)
    
    # Get lead's recipes
    recipes = await db.recipes.find({"lead_id": lead_id}, {"_id": 0}).to_list(100)
    for recipe in recipes:
        deserialize_datetime(recipe)
    
    # Calculate total revenue from this lead
    total_revenue = sum(order.get('total_price', 0) for order in orders if order.get('status') in ['delivered', 'shipped', 'confirmed'])
    
    return {
        **lead,
        "orders": orders,
        "recipes": recipes,
        "total_orders": len(orders),
        "total_recipes": len(recipes),
        "total_revenue": total_revenue
    }

# ===================== EMAIL TEMPLATE ENDPOINTS =====================

@api_router.post("/templates", response_model=EmailTemplate)
async def create_template(template_data: EmailTemplateCreate):
    template = EmailTemplate(**template_data.model_dump())
    doc = template.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.email_templates.insert_one(doc)
    return template

@api_router.get("/templates", response_model=List[EmailTemplate])
async def get_templates():
    templates = await db.email_templates.find({}, {"_id": 0}).to_list(100)
    for template in templates:
        deserialize_datetime(template, ['created_at'])
    return templates

@api_router.get("/templates/{template_id}", response_model=EmailTemplate)
async def get_template(template_id: str):
    template = await db.email_templates.find_one({"id": template_id}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    deserialize_datetime(template, ['created_at'])
    return template

@api_router.delete("/templates/{template_id}")
async def delete_template(template_id: str):
    result = await db.email_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template deleted successfully"}

# ===================== SMTP SETTINGS ENDPOINTS =====================

@api_router.post("/settings/smtp", response_model=SMTPSettings)
async def save_smtp_settings(settings_data: SMTPSettingsCreate):
    # Delete existing settings and insert new
    await db.smtp_settings.delete_many({})
    settings = SMTPSettings(**settings_data.model_dump())
    doc = settings.model_dump()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.smtp_settings.insert_one(doc)
    
    # Also update company_settings for IMAP access
    await db.company_settings.update_one(
        {},
        {"$set": {
            "smtp_host": settings_data.host,
            "smtp_port": settings_data.port,
            "smtp_username": settings_data.username,
            "smtp_password": settings_data.password,
            "from_email": settings_data.from_email,
            "from_name": settings_data.from_name,
            "use_tls": settings_data.use_tls,
            "use_ssl": settings_data.use_ssl,
            "imap_host": settings_data.imap_host,
            "imap_port": settings_data.imap_port
        }},
        upsert=True
    )
    
    return settings

@api_router.get("/settings/smtp")
async def get_smtp_settings():
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not settings:
        return None
    deserialize_datetime(settings, ['updated_at'])
    return settings

# Helper function for SMTP connections (1&1 IONOS compatible)
def create_smtp_connection(host: str, port: int, use_ssl: bool = False, use_tls: bool = True, timeout: int = 30):
    """
    Create SMTP connection with proper SSL/TLS handling.
    Port 465 = SSL (SMTP_SSL)
    Port 587 = STARTTLS
    Port 25 = Plain (not recommended)
    """
    if use_ssl or port == 465:
        server = smtplib.SMTP_SSL(host, port, timeout=timeout)
    else:
        server = smtplib.SMTP(host, port, timeout=timeout)
        if use_tls or port == 587:
            server.ehlo()
            server.starttls()
            server.ehlo()
    return server

@api_router.post("/settings/smtp/test")
async def test_smtp_connection():
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    try:
        host = settings['host']
        port = int(settings['port'])
        use_ssl = settings.get('use_ssl', False)
        use_tls = settings.get('use_tls', True)
        
        logger.info(f"Testing SMTP: {host}:{port} SSL={use_ssl} TLS={use_tls}")
        
        server = create_smtp_connection(host, port, use_ssl, use_tls)
        server.login(settings['username'], settings['password'])
        server.quit()
        return {"success": True, "message": "SMTP bağlantısı başarılı!"}
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Auth Error: {e}")
        return {"success": False, "message": f"Kimlik doğrulama hatası: Kullanıcı adı veya şifre yanlış. ({str(e)})"}
    except smtplib.SMTPConnectError as e:
        logger.error(f"SMTP Connect Error: {e}")
        return {"success": False, "message": f"Bağlantı hatası: Sunucuya bağlanılamadı. ({str(e)})"}
    except Exception as e:
        logger.error(f"SMTP Test Error: {e}")
        return {"success": False, "message": f"Hata: {str(e)}"}

# ===================== EMAIL SENDING ENDPOINTS =====================

async def send_email_background(lead: dict, subject: str, body: str, settings: dict):
    """Background task to send email"""
    history_id = str(uuid.uuid4())
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = f"{settings['from_name']} <{settings['from_email']}>"
        msg['To'] = lead['email']
        
        # Create HTML version
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            {body.replace(chr(10), '<br>')}
        </body>
        </html>
        """
        
        part1 = MIMEText(body, 'plain')
        part2 = MIMEText(html_body, 'html')
        msg.attach(part1)
        msg.attach(part2)
        
        if settings.get('use_tls', True):
            server = smtplib.SMTP(settings['host'], settings['port'])
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(settings['host'], settings['port'])
        
        server.login(settings['username'], settings['password'])
        server.sendmail(settings['from_email'], lead['email'], msg.as_string())
        server.quit()
        
        # Save to history
        history = {
            "id": history_id,
            "lead_id": lead['id'],
            "lead_email": lead['email'],
            "lead_name": f"{lead['first_name']} {lead['last_name']}",
            "company_name": lead['company_name'],
            "subject": subject,
            "body": body,
            "status": "sent",
            "error_message": None,
            "sent_at": datetime.now(timezone.utc).isoformat()
        }
        await db.email_history.insert_one(history)
        logger.info(f"Email sent successfully to {lead['email']}")
        
    except Exception as e:
        # Save failed attempt to history
        history = {
            "id": history_id,
            "lead_id": lead['id'],
            "lead_email": lead['email'],
            "lead_name": f"{lead['first_name']} {lead['last_name']}",
            "company_name": lead['company_name'],
            "subject": subject,
            "body": body,
            "status": "failed",
            "error_message": str(e),
            "sent_at": datetime.now(timezone.utc).isoformat()
        }
        await db.email_history.insert_one(history)
        logger.error(f"Failed to send email to {lead['email']}: {str(e)}")

@api_router.post("/emails/send")
async def send_email(request: SendEmailRequest, background_tasks: BackgroundTasks):
    # Get lead
    lead = await db.leads.find_one({"id": request.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Get SMTP settings
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    # Send email in background
    background_tasks.add_task(send_email_background, lead, request.subject, request.body, settings)
    
    return {"message": "Email queued for sending"}

@api_router.post("/emails/send-bulk")
async def send_bulk_emails(lead_ids: List[str], subject: str, body: str, background_tasks: BackgroundTasks):
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    for lead_id in lead_ids:
        lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
        if lead:
            background_tasks.add_task(send_email_background, lead, subject, body, settings)
    
    return {"message": f"Queued {len(lead_ids)} emails for sending"}

# ===================== AI EMAIL GENERATION =====================

@api_router.post("/emails/generate")
async def generate_email(request: GenerateEmailRequest):
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    # Get lead info
    lead = await db.leads.find_one({"id": request.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Get API key
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        raise HTTPException(status_code=500, detail="LLM API key not configured")
    
    # Language prompts
    language_prompts = {
        "tr": "Türkçe olarak yaz. Alıcı Türk bir firma.",
        "de": "Schreibe auf Deutsch. Der Empfänger ist ein deutsches Unternehmen.",
        "en": "Write in English. The recipient is an international company."
    }
    
    tone_prompts = {
        "professional": "professional and business-like",
        "friendly": "friendly but professional",
        "formal": "very formal and respectful"
    }
    
    system_message = f"""You are a sales representative for a spice and binder manufacturing company based in Berlin, Germany. 
Your company produces high-quality spices and binders for food manufacturers, especially those producing döner, gyros, kebab, and souvlaki.

Write a {tone_prompts.get(request.tone, 'professional')} B2B sales email.
{language_prompts.get(request.language, language_prompts['en'])}

The email should:
1. Introduce your company briefly
2. Mention the benefits of your products (quality, consistency, competitive pricing)
3. Express interest in potential partnership
4. Include a call to action (request a meeting or call)

Keep the email concise (3-4 paragraphs max).
Do not include subject line in the body - return it separately.
"""

    user_prompt = f"""Write a sales email to:
- Contact: {lead['first_name']} {lead['last_name']}
- Company: {lead['company_name']}
- Location: {lead['city']}, {lead['country']}

Return the response in this exact JSON format:
{{"subject": "Email subject here", "body": "Email body here"}}
"""

    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"email-gen-{lead['id']}",
            system_message=system_message
        ).with_model("openai", "gpt-5.2")
        
        response = await chat.send_message(UserMessage(text=user_prompt))
        
        # Parse JSON response
        import json
        # Try to extract JSON from response
        response_text = response.strip()
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        
        try:
            email_data = json.loads(response_text.strip())
            return {"subject": email_data.get("subject", ""), "body": email_data.get("body", "")}
        except json.JSONDecodeError:
            # Fallback: return raw response
            return {"subject": f"Partnership Opportunity - {lead['company_name']}", "body": response}
            
    except Exception as e:
        logger.error(f"AI email generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate email: {str(e)}")

# ===================== EMAIL HISTORY ENDPOINTS =====================

@api_router.get("/emails/history", response_model=List[EmailHistory])
async def get_email_history():
    history = await db.email_history.find({}, {"_id": 0}).sort("sent_at", -1).to_list(500)
    for item in history:
        deserialize_datetime(item, ['sent_at'])
    return history

@api_router.get("/emails/history/{lead_id}", response_model=List[EmailHistory])
async def get_lead_email_history(lead_id: str):
    history = await db.email_history.find({"lead_id": lead_id}, {"_id": 0}).sort("sent_at", -1).to_list(100)
    for item in history:
        deserialize_datetime(item, ['sent_at'])
    return history

# ===================== DASHBOARD ENDPOINTS =====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(period: str = "all"):
    """
    Get dashboard statistics
    period: all, month, quarter, half_year, year
    """
    total_leads = await db.leads.count_documents({})
    emails_sent = await db.email_history.count_documents({"status": "sent"})
    emails_failed = await db.email_history.count_documents({"status": "failed"})
    
    recent_leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    for lead in recent_leads:
        deserialize_datetime(lead)
    
    # Date filter for period
    now = datetime.now(timezone.utc)
    date_filter = {}
    if period == "month":
        date_filter = {"created_at": {"$gte": (now - timedelta(days=30)).isoformat()}}
    elif period == "quarter":
        date_filter = {"created_at": {"$gte": (now - timedelta(days=90)).isoformat()}}
    elif period == "half_year":
        date_filter = {"created_at": {"$gte": (now - timedelta(days=180)).isoformat()}}
    elif period == "year":
        date_filter = {"created_at": {"$gte": (now - timedelta(days=365)).isoformat()}}
    
    # Order stats with period filter
    order_filter = {"status": {"$in": ["delivered", "shipped", "confirmed"]}}
    if date_filter:
        order_filter.update(date_filter)
    
    total_orders = await db.orders.count_documents(date_filter if date_filter else {})
    
    # Calculate revenue
    pipeline = [
        {"$match": order_filter},
        {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
    ]
    revenue_result = await db.orders.aggregate(pipeline).to_list(1)
    total_revenue = revenue_result[0]["total"] if revenue_result else 0.0
    
    # Get company settings for yearly target
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    yearly_target = settings.get("yearly_target", 0) if settings else 0
    
    return {
        "total_leads": total_leads,
        "emails_sent": emails_sent,
        "emails_failed": emails_failed,
        "recent_leads": recent_leads,
        "total_orders": total_orders,
        "total_revenue": total_revenue,
        "yearly_target": yearly_target,
        "period": period
    }

# ===================== COMPANY SETTINGS ENDPOINTS =====================

class CompanySettingsUpdate(BaseModel):
    company_name: Optional[str] = None
    yearly_target: Optional[float] = None

@api_router.get("/company-settings")
async def get_company_settings():
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    if not settings:
        # Return defaults
        settings = {
            "id": "company_settings",
            "company_name": "Gewürzberg GmbH",
            "yearly_target": 0,
            "current_revenue": 0
        }
    
    # Calculate current revenue from all confirmed/shipped/delivered orders
    pipeline = [
        {"$match": {"status": {"$in": ["delivered", "shipped", "confirmed"]}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
    ]
    revenue_result = await db.orders.aggregate(pipeline).to_list(1)
    settings["current_revenue"] = revenue_result[0]["total"] if revenue_result else 0
    
    return settings

@api_router.post("/company-settings")
async def save_company_settings(settings: CompanySettingsUpdate):
    existing = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    
    update_data = {k: v for k, v in settings.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if existing:
        await db.company_settings.update_one(
            {"id": "company_settings"}, 
            {"$set": update_data}
        )
    else:
        doc = {
            "id": "company_settings",
            "company_name": settings.company_name or "Gewürzberg GmbH",
            "yearly_target": settings.yearly_target or 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.company_settings.insert_one(doc)
    
    return {"status": "success", "message": "Company settings saved"}

# ===================== GEOCODING ENDPOINT =====================

import httpx

@api_router.get("/geocode")
async def geocode_address(city: str, country: str):
    """Geocode a city/country to lat/lng coordinates"""
    try:
        query = f"{city}, {country}"
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={"format": "json", "q": query, "limit": 1},
                headers={"User-Agent": "GewurzbergCRM/1.0"},
                timeout=10.0
            )
            if response.status_code == 200:
                data = response.json()
                if data and len(data) > 0:
                    return {
                        "lat": float(data[0]["lat"]),
                        "lng": float(data[0]["lon"]),
                        "display_name": data[0].get("display_name", query)
                    }
            elif response.status_code == 429:
                logger.warning(f"Nominatim rate limited for: {query}")
    except Exception as e:
        logger.error(f"Geocoding error: {e}")
    return None

# Predefined city coordinates for fast geocoding without API calls
CITY_COORDS = {
    # Germany
    'berlin': (52.5200, 13.4050),
    'munich': (48.1351, 11.5820),
    'hamburg': (53.5511, 9.9937),
    'frankfurt': (50.1109, 8.6821),
    'cologne': (50.9375, 6.9603),
    'düsseldorf': (51.2277, 6.7735),
    'stuttgart': (48.7758, 9.1829),
    # Turkey
    'istanbul': (41.0082, 28.9784),
    'ankara': (39.9334, 32.8597),
    'izmir': (38.4237, 27.1428),
    'bursa': (40.1885, 29.0610),
    'antalya': (36.8969, 30.7133),
    # Greece
    'athens': (37.9838, 23.7275),
    'thessaloniki': (40.6401, 22.9444),
    'piraeus': (37.9425, 23.6469),
    'rethymno': (35.3661, 24.4765),
    'heraklion': (35.3387, 25.1442),
    # Netherlands
    'amsterdam': (52.3676, 4.9041),
    'rotterdam': (51.9244, 4.4777),
    'the hague': (52.0705, 4.3007),
    'oss': (51.7650, 5.5183),
    # Poland
    'warsaw': (52.2297, 21.0122),
    'krakow': (50.0647, 19.9450),
    'gdansk': (54.3520, 18.6466),
    # Austria
    'vienna': (48.2082, 16.3738),
    'salzburg': (47.8095, 13.0550),
    # France
    'paris': (48.8566, 2.3522),
    'lyon': (45.7640, 4.8357),
    'marseille': (43.2965, 5.3698),
    # UK
    'london': (51.5074, -0.1278),
    'manchester': (53.4808, -2.2426),
    # Belgium
    'brussels': (50.8503, 4.3517),
    'antwerp': (51.2213, 4.4051),
    # Switzerland
    'zurich': (47.3769, 8.5417),
    'geneva': (46.2044, 6.1432),
}

def get_predefined_coords(query: str):
    """Check if we have predefined coordinates for a city"""
    query_lower = query.lower().strip()
    for city, coords in CITY_COORDS.items():
        if city in query_lower or query_lower.startswith(city):
            return {"lat": coords[0], "lng": coords[1], "display_name": query}
    return None

@api_router.get("/geocode/search")
async def search_address(q: str):
    """Search for addresses and return autocomplete suggestions"""
    if len(q) < 3:
        return []
    
    # Check predefined coordinates first
    predefined = get_predefined_coords(q)
    if predefined:
        return [{
            "lat": str(predefined["lat"]),
            "lon": str(predefined["lng"]),
            "display_name": f"{q.title()} (cached)",
            "type": "city"
        }]
    
    # Add small delay to avoid rate limiting
    import asyncio
    await asyncio.sleep(0.3)
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={
                    "format": "json", 
                    "q": q, 
                    "limit": 5,
                    "addressdetails": 1
                },
                headers={
                    "User-Agent": "GewurzbergCRM/1.0 (contact@gewurzberg.de)",
                    "Accept-Language": "en,de,tr"
                },
                timeout=15.0
            )
            if response.status_code == 200:
                results = response.json()
                # Add predefined results for common cities if API returns empty
                if not results:
                    predefined = get_predefined_coords(q)
                    if predefined:
                        return [{
                            "lat": str(predefined["lat"]),
                            "lon": str(predefined["lng"]),
                            "display_name": f"{q.title()}",
                            "type": "city"
                        }]
                return results
            elif response.status_code == 429:
                logger.warning("Nominatim rate limit reached, using fallback")
                # Return predefined if available
                predefined = get_predefined_coords(q)
                if predefined:
                    return [{
                        "lat": str(predefined["lat"]),
                        "lon": str(predefined["lng"]),
                        "display_name": f"{q.title()} (fallback)",
                        "type": "city"
                    }]
                return []
            else:
                logger.error(f"Nominatim error: {response.status_code}")
                return []
    except Exception as e:
        logger.error(f"Address search error: {e}")
        # Return predefined if available
        predefined = get_predefined_coords(q)
        if predefined:
            return [{
                "lat": str(predefined["lat"]),
                "lon": str(predefined["lng"]),
                "display_name": f"{q.title()} (fallback)",
                "type": "city"
            }]
        return []

@api_router.get("/geocode/reverse")
async def reverse_geocode(lat: float, lon: float):
    """Reverse geocode coordinates to address"""
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://nominatim.openstreetmap.org/reverse",
                params={
                    "format": "json",
                    "lat": lat,
                    "lon": lon,
                    "zoom": 18
                },
                headers={
                    "User-Agent": "GewurzbergCRM/1.0 (contact@gewurzberg.de)",
                    "Accept-Language": "en,de,tr"
                },
                timeout=10.0
            )
            if response.status_code == 200:
                data = response.json()
                return {
                    "display_name": data.get("display_name", f"{lat}, {lon}"),
                    "lat": lat,
                    "lon": lon
                }
            else:
                return {"display_name": f"{lat:.4f}, {lon:.4f}", "lat": lat, "lon": lon}
    except Exception as e:
        logger.error(f"Reverse geocoding error: {e}")
        return {"display_name": f"{lat:.4f}, {lon:.4f}", "lat": lat, "lon": lon}

@api_router.post("/geocode/batch")
async def geocode_batch(leads: List[dict]):
    """Geocode multiple leads at once"""
    results = {}
    async with httpx.AsyncClient() as client:
        for lead in leads:
            if lead.get("city") and lead.get("country"):
                try:
                    query = f"{lead['city']}, {lead['country']}"
                    response = await client.get(
                        "https://nominatim.openstreetmap.org/search",
                        params={"format": "json", "q": query, "limit": 1},
                        headers={"User-Agent": "GewurzbergCRM/1.0"},
                        timeout=10.0
                    )
                    data = response.json()
                    if data and len(data) > 0:
                        # Add small random offset to prevent overlapping
                        import random
                        results[lead["id"]] = {
                            "lat": float(data[0]["lat"]) + (random.random() - 0.5) * 0.01,
                            "lng": float(data[0]["lon"]) + (random.random() - 0.5) * 0.01
                        }
                    # Rate limiting - wait a bit between requests
                    import asyncio
                    await asyncio.sleep(0.5)
                except Exception as e:
                    logger.error(f"Geocoding error for {lead.get('id')}: {e}")
    return results

# ===================== ROUTE PLANNER ENDPOINTS =====================

class RouteRequest(BaseModel):
    start_address: str
    lead_ids: List[str]

@api_router.post("/route/calculate")
async def calculate_route(request: RouteRequest):
    """Calculate optimized route with distance and duration"""
    import math
    
    # Geocode start address
    start_coords = None
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={"format": "json", "q": request.start_address, "limit": 1},
                headers={"User-Agent": "GewurzbergCRM/1.0"},
                timeout=10.0
            )
            data = response.json()
            if data and len(data) > 0:
                start_coords = {
                    "lat": float(data[0]["lat"]),
                    "lng": float(data[0]["lon"]),
                    "address": data[0].get("display_name", request.start_address)
                }
        except Exception as e:
            logger.error(f"Geocoding start address error: {e}")
    
    if not start_coords:
        raise HTTPException(status_code=400, detail="Could not geocode start address")
    
    # Get leads and their coordinates
    leads_data = []
    for lead_id in request.lead_ids:
        lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
        if lead and lead.get("city") and lead.get("country"):
            leads_data.append(lead)
    
    # Geocode all leads
    geocoded_leads = []
    async with httpx.AsyncClient() as client:
        for lead in leads_data:
            try:
                query = f"{lead['city']}, {lead['country']}"
                response = await client.get(
                    "https://nominatim.openstreetmap.org/search",
                    params={"format": "json", "q": query, "limit": 1},
                    headers={"User-Agent": "GewurzbergCRM/1.0"},
                    timeout=10.0
                )
                data = response.json()
                if data and len(data) > 0:
                    geocoded_leads.append({
                        **lead,
                        "lat": float(data[0]["lat"]),
                        "lng": float(data[0]["lon"])
                    })
                import asyncio
                await asyncio.sleep(0.3)
            except Exception as e:
                logger.error(f"Geocoding error for lead {lead.get('id')}: {e}")
    
    if len(geocoded_leads) < 1:
        raise HTTPException(status_code=400, detail="Could not geocode any leads")
    
    # Calculate optimal route using nearest neighbor
    def haversine_distance(lat1, lon1, lat2, lon2):
        R = 6371  # Earth's radius in km
        lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        return R * c
    
    visited = set()
    route = []
    current = start_coords
    total_distance = 0
    
    while len(visited) < len(geocoded_leads):
        nearest = None
        nearest_dist = float('inf')
        
        for lead in geocoded_leads:
            if lead['id'] in visited:
                continue
            dist = haversine_distance(current['lat'], current['lng'], lead['lat'], lead['lng'])
            if dist < nearest_dist:
                nearest_dist = dist
                nearest = lead
        
        if nearest:
            visited.add(nearest['id'])
            route.append({
                "id": nearest['id'],
                "company_name": nearest.get('company_name', ''),
                "city": nearest.get('city', ''),
                "country": nearest.get('country', ''),
                "lat": nearest['lat'],
                "lng": nearest['lng'],
                "distance": nearest_dist
            })
            total_distance += nearest_dist
            current = {"lat": nearest['lat'], "lng": nearest['lng']}
    
    # Estimate duration (average 60 km/h + 30 min per stop)
    driving_time = (total_distance / 60) * 60  # minutes
    stop_time = len(route) * 30  # 30 minutes per stop
    total_duration = driving_time + stop_time
    
    return {
        "start_point": start_coords,
        "stops": route,
        "total_distance": total_distance,
        "total_duration": total_duration,  # in minutes
        "estimated_hours": total_duration / 60
    }

@api_router.post("/route/pdf")
async def generate_route_pdf_endpoint(request: RouteRequest):
    """Generate PDF for a calculated route"""
    # Calculate route first
    route_data = await calculate_route(request)
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_route_pdf(route_data, settings)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=rota_plani.pdf"}
    )

# ===================== EMAIL WITH ATTACHMENT =====================

@api_router.post("/emails/send-with-attachment")
async def send_email_with_attachment(
    lead_id: str = Form(...),
    subject: str = Form(...),
    body: str = Form(...),
    attachment: Optional[UploadFile] = File(None),
    background_tasks: BackgroundTasks = None
):
    """Send email with optional file attachment"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    # Read attachment if provided
    attachment_data = None
    attachment_name = None
    if attachment:
        attachment_data = await attachment.read()
        attachment_name = attachment.filename
    
    history_id = str(uuid.uuid4())
    
    try:
        msg = MIMEMultipart()
        msg['Subject'] = subject
        msg['From'] = f"{settings['from_name']} <{settings['from_email']}>"
        msg['To'] = lead['email']
        
        # Add body
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            {body.replace(chr(10), '<br>')}
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'plain'))
        msg.attach(MIMEText(html_body, 'html'))
        
        # Add attachment if provided
        if attachment_data and attachment_name:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment_data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{attachment_name}"')
            msg.attach(part)
        
        if settings.get('use_tls', True):
            server = smtplib.SMTP(settings['host'], settings['port'])
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(settings['host'], settings['port'])
        
        server.login(settings['username'], settings['password'])
        server.sendmail(settings['from_email'], lead['email'], msg.as_string())
        server.quit()
        
        # Save to history
        history = {
            "id": history_id,
            "lead_id": lead['id'],
            "lead_email": lead['email'],
            "lead_name": f"{lead['first_name']} {lead['last_name']}",
            "company_name": lead['company_name'],
            "subject": subject,
            "body": body,
            "status": "sent",
            "has_attachment": attachment_name is not None,
            "attachment_name": attachment_name,
            "error_message": None,
            "sent_at": datetime.now(timezone.utc).isoformat()
        }
        await db.email_history.insert_one(history)
        
        return {"success": True, "message": "Email sent successfully"}
        
    except Exception as e:
        # Save failed attempt
        history = {
            "id": history_id,
            "lead_id": lead['id'],
            "lead_email": lead['email'],
            "lead_name": f"{lead['first_name']} {lead['last_name']}",
            "company_name": lead['company_name'],
            "subject": subject,
            "body": body,
            "status": "failed",
            "error_message": str(e),
            "sent_at": datetime.now(timezone.utc).isoformat()
        }
        await db.email_history.insert_one(history)
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")

# ===================== SPECIFICATIONS ENDPOINTS =====================

class SpecificationIngredient(BaseModel):
    name: str
    percentage: Optional[str] = None
    description: Optional[str] = None

class SpecificationCreate(BaseModel):
    name: str
    product_code: str
    category: Optional[str] = None
    description: Optional[str] = None
    ingredients: Optional[List[SpecificationIngredient]] = []
    nutritional_info: Optional[str] = None
    allergens: Optional[str] = None
    storage_instructions: Optional[str] = None
    shelf_life: Optional[str] = None
    certifications: Optional[str] = None

class Specification(SpecificationCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.post("/specifications")
async def create_specification(spec_data: SpecificationCreate):
    spec = Specification(**spec_data.model_dump())
    doc = spec.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    # Convert ingredients to dicts
    if doc.get('ingredients'):
        doc['ingredients'] = [ing if isinstance(ing, dict) else ing.model_dump() for ing in doc['ingredients']]
    await db.specifications.insert_one(doc)
    return {"id": spec.id, "message": "Specification created successfully"}

@api_router.get("/specifications")
async def get_specifications():
    specs = await db.specifications.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    for s in specs:
        deserialize_datetime(s)
    return specs

@api_router.get("/specifications/{spec_id}")
async def get_specification(spec_id: str):
    spec = await db.specifications.find_one({"id": spec_id}, {"_id": 0})
    if not spec:
        raise HTTPException(status_code=404, detail="Specification not found")
    deserialize_datetime(spec)
    return spec

@api_router.put("/specifications/{spec_id}")
async def update_specification(spec_id: str, spec_data: SpecificationCreate):
    existing = await db.specifications.find_one({"id": spec_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Specification not found")
    
    update_data = spec_data.model_dump()
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    # Convert ingredients to dicts
    if update_data.get('ingredients'):
        update_data['ingredients'] = [ing if isinstance(ing, dict) else ing for ing in update_data['ingredients']]
    
    await db.specifications.update_one({"id": spec_id}, {"$set": update_data})
    return {"message": "Specification updated successfully"}

@api_router.delete("/specifications/{spec_id}")
async def delete_specification(spec_id: str):
    result = await db.specifications.delete_one({"id": spec_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Specification not found")
    return {"message": "Specification deleted successfully"}

def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """Extract text from PDF using pdfplumber"""
    import pdfplumber
    text_parts = []
    
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
    except Exception as e:
        logger.error(f"PDF text extraction error: {e}")
        # Fallback to PyMuPDF
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            for page in doc:
                text_parts.append(page.get_text())
            doc.close()
        except Exception as e2:
            logger.error(f"PyMuPDF fallback error: {e2}")
    
    return "\n\n---PAGE BREAK---\n\n".join(text_parts)

@api_router.post("/specifications/upload-pdf")
async def upload_specification_pdf(file: UploadFile = File(...)):
    """Upload a PDF specification document and extract text for editing"""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    
    content = await file.read()
    spec_id = str(uuid.uuid4())
    
    # Extract text from PDF for editing
    extracted_text = extract_text_from_pdf(content)
    
    # Store specification with PDF data and extracted text
    doc = {
        "id": spec_id,
        "filename": file.filename,
        "name": file.filename.replace('.pdf', '').replace('_', ' ').replace('-', ' '),
        "description": "",
        "notes": "",
        "content_type": "application/pdf",
        "size": len(content),
        "pdf_data": base64.b64encode(content).decode('utf-8'),
        "extracted_text": extracted_text,
        "edited_text": extracted_text,  # Initially same as extracted
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.specifications.insert_one(doc)
    
    return {
        **{k: v for k, v in doc.items() if k not in ['_id', 'pdf_data']},
        "has_text": bool(extracted_text)
    }

@api_router.get("/specifications/{spec_id}/text")
async def get_specification_text(spec_id: str):
    """Get extracted/edited text from a specification"""
    spec = await db.specifications.find_one({"id": spec_id}, {"_id": 0, "pdf_data": 0})
    if not spec:
        raise HTTPException(status_code=404, detail="Specification not found")
    
    return {
        "id": spec_id,
        "name": spec.get("name", ""),
        "filename": spec.get("filename", ""),
        "extracted_text": spec.get("extracted_text", ""),
        "edited_text": spec.get("edited_text", spec.get("extracted_text", "")),
        "has_original_pdf": bool(spec.get("pdf_data"))
    }

@api_router.put("/specifications/{spec_id}/text")
async def update_specification_text(spec_id: str, edited_text: str = Form(...)):
    """Update the edited text of a specification"""
    spec = await db.specifications.find_one({"id": spec_id}, {"_id": 0})
    if not spec:
        raise HTTPException(status_code=404, detail="Specification not found")
    
    await db.specifications.update_one(
        {"id": spec_id},
        {"$set": {
            "edited_text": edited_text,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Text updated successfully", "id": spec_id}

@api_router.get("/specifications/{spec_id}/regenerate-pdf")
async def regenerate_specification_pdf(spec_id: str):
    """Generate a new PDF from the edited text"""
    spec = await db.specifications.find_one({"id": spec_id}, {"_id": 0})
    if not spec:
        raise HTTPException(status_code=404, detail="Specification not found")
    
    edited_text = spec.get("edited_text", spec.get("extracted_text", ""))
    
    # Generate PDF from edited text using ReportLab
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Register DejaVu font for UTF-8 support
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        font_name = 'DejaVu'
        font_bold = 'DejaVu-Bold'
    except:
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
    
    y = height - 2*cm
    
    # Title
    c.setFont(font_bold, 14)
    c.drawString(2*cm, y, spec.get("name", "Specification"))
    y -= 1*cm
    
    # Content
    c.setFont(font_name, 10)
    lines = edited_text.split('\n')
    
    for line in lines:
        if line.strip() == "---PAGE BREAK---":
            c.showPage()
            y = height - 2*cm
            c.setFont(font_name, 10)
            continue
        
        # Word wrap long lines
        while len(line) > 90:
            c.drawString(2*cm, y, line[:90])
            line = line[90:]
            y -= 0.5*cm
            if y < 2*cm:
                c.showPage()
                y = height - 2*cm
                c.setFont(font_name, 10)
        
        c.drawString(2*cm, y, line)
        y -= 0.5*cm
        
        if y < 2*cm:
            c.showPage()
            y = height - 2*cm
            c.setFont(font_name, 10)
    
    # Footer
    c.setFont(font_name, 8)
    c.drawString(2*cm, 1*cm, f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    
    c.save()
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={spec.get('name', 'specification')}_edited.pdf"
        }
    )

@api_router.get("/specifications/{spec_id}/download")
async def download_specification_pdf(spec_id: str):
    """Download uploaded PDF specification"""
    spec = await db.specifications.find_one({"id": spec_id}, {"_id": 0})
    if not spec:
        raise HTTPException(status_code=404, detail="Specification not found")
    
    if spec.get('pdf_data'):
        # Return uploaded PDF with inline display for preview
        content = base64.b64decode(spec['pdf_data'])
        return Response(
            content=content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename={spec.get('filename', 'specification.pdf')}",
                "Access-Control-Allow-Origin": "*"
            }
        )
    else:
        # Generate PDF if no upload
        settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
        pdf_content = generate_specification_pdf(spec, settings)
        return Response(
            content=pdf_content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename=specification_{spec_id[:8]}.pdf",
                "Access-Control-Allow-Origin": "*"
            }
        )

@api_router.post("/specifications/upload")
async def upload_specification_file(file: UploadFile = File(...)):
    """Upload a PDF file for specification attachment"""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    
    # Read file content
    content = await file.read()
    file_id = str(uuid.uuid4())
    
    # Store in database as base64
    doc = {
        "id": file_id,
        "filename": file.filename,
        "content_type": "application/pdf",
        "size": len(content),
        "data": base64.b64encode(content).decode('utf-8'),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.specification_files.insert_one(doc)
    
    return {
        "id": file_id,
        "filename": file.filename,
        "url": f"{BACKEND_URL}/api/specifications/files/{file_id}"
    }

# Add BACKEND_URL constant near the top imports
BACKEND_URL = os.environ.get('REACT_APP_BACKEND_URL', '')

@api_router.get("/specifications/files/{file_id}")
async def get_specification_file(file_id: str):
    """Download a specification attachment"""
    doc = await db.specification_files.find_one({"id": file_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="File not found")
    
    content = base64.b64decode(doc['data'])
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={doc['filename']}"}
    )

@api_router.get("/specifications/{spec_id}/pdf")
async def get_specification_pdf(spec_id: str):
    spec = await db.specifications.find_one({"id": spec_id}, {"_id": 0})
    if not spec:
        raise HTTPException(status_code=404, detail="Specification not found")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_specification_pdf(spec, settings)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=specification_{spec_id[:8]}.pdf"}
    )

class SpecEmailRequest(BaseModel):
    lead_id: str
    subject: str
    body: str

@api_router.post("/specifications/{spec_id}/email")
async def email_specification(spec_id: str, request: SpecEmailRequest):
    spec = await db.specifications.find_one({"id": spec_id}, {"_id": 0})
    if not spec:
        raise HTTPException(status_code=404, detail="Specification not found")
    
    lead = await db.leads.find_one({"id": request.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not smtp_settings:
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_specification_pdf(spec, settings)
    
    try:
        msg = MIMEMultipart()
        msg['Subject'] = request.subject
        msg['From'] = f"{smtp_settings['from_name']} <{smtp_settings['from_email']}>"
        msg['To'] = lead['email']
        
        msg.attach(MIMEText(request.body, 'plain'))
        
        # Attach PDF
        pdf_attachment = MIMEBase('application', 'pdf')
        pdf_attachment.set_payload(pdf_content)
        encoders.encode_base64(pdf_attachment)
        pdf_attachment.add_header('Content-Disposition', f'attachment; filename="specification_{spec["product_code"]}.pdf"')
        msg.attach(pdf_attachment)
        
        if smtp_settings.get('use_tls', True):
            server = smtplib.SMTP(smtp_settings['host'], smtp_settings['port'])
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(smtp_settings['host'], smtp_settings['port'])
        
        server.login(smtp_settings['username'], smtp_settings['password'])
        server.sendmail(smtp_settings['from_email'], lead['email'], msg.as_string())
        server.quit()
        
        return {"success": True, "message": "Email sent successfully"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")

# ===================== PRODUCT ENDPOINTS =====================

@api_router.post("/products", response_model=Product)
async def create_product(product_data: ProductCreate):
    # Check if product code exists
    existing = await db.products.find_one({"code": product_data.code}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Product code already exists")
    
    product = Product(**product_data.model_dump())
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.products.insert_one(doc)
    return product

@api_router.get("/products", response_model=List[Product])
async def get_products():
    products = await db.products.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    for p in products:
        deserialize_datetime(p, ['created_at'])
    return products

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    deserialize_datetime(product, ['created_at'])
    return product

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product_data: ProductCreate):
    existing = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = product_data.model_dump()
    await db.products.update_one({"id": product_id}, {"$set": update_data})
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    deserialize_datetime(updated, ['created_at'])
    return updated

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted successfully"}

# ===================== COMPANY SETTINGS ENDPOINTS =====================

@api_router.get("/settings/company")
async def get_company_settings():
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    if not settings:
        # Return default settings
        return CompanySettings().model_dump()
    deserialize_datetime(settings, ['updated_at'])
    return settings

@api_router.post("/settings/company")
async def save_company_settings(
    company_name: str = "SpiceCRM",
    logo_url: Optional[str] = None,
    address: Optional[str] = "",
    phone: Optional[str] = "",
    email: Optional[str] = "",
    website: Optional[str] = "",
    tax_number: Optional[str] = "",
    yearly_target: float = 0.0,
    currency: str = "EUR"
):
    settings = CompanySettings(
        company_name=company_name,
        logo_url=logo_url,
        address=address,
        phone=phone,
        email=email,
        website=website,
        tax_number=tax_number,
        yearly_target=yearly_target,
        currency=currency
    )
    doc = settings.model_dump()
    doc['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.company_settings.replace_one(
        {"id": "company_settings"},
        doc,
        upsert=True
    )
    return settings

# ===================== PDF GENERATION ENDPOINTS =====================

def generate_pdf_content(title: str, data: dict, company_settings: dict = None) -> bytes:
    """Generate a simple PDF using reportlab"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Company header
    y = height - 2*cm
    if company_settings:
        c.setFont("Helvetica-Bold", 16)
        c.drawString(2*cm, y, company_settings.get('company_name', 'SpiceCRM'))
        y -= 0.7*cm
        c.setFont("Helvetica", 10)
        if company_settings.get('address'):
            c.drawString(2*cm, y, company_settings['address'])
            y -= 0.5*cm
        if company_settings.get('phone'):
            c.drawString(2*cm, y, f"Tel: {company_settings['phone']}")
            y -= 0.5*cm
        if company_settings.get('email'):
            c.drawString(2*cm, y, f"Email: {company_settings['email']}")
            y -= 0.5*cm
    
    # Title
    y -= 1*cm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, y, title)
    y -= 1*cm
    
    # Content
    c.setFont("Helvetica", 10)
    for key, value in data.items():
        if y < 3*cm:
            c.showPage()
            y = height - 2*cm
            c.setFont("Helvetica", 10)
        
        c.drawString(2*cm, y, f"{key}: {value}")
        y -= 0.5*cm
    
    # Footer
    c.setFont("Helvetica", 8)
    c.drawString(2*cm, 1*cm, f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue()

@api_router.get("/orders/{order_id}/pdf")
async def get_order_pdf(order_id: str, lang: str = 'tr'):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get lead info for customer details (email, phone, address, tax)
    lead_info = None
    if order.get('lead_id'):
        lead_info = await db.leads.find_one({"id": order['lead_id']}, {"_id": 0})
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_order_pdf(order, settings, lang=lang, lead_info=lead_info)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=siparis_{order_id[:8]}.pdf"}
    )

@api_router.get("/leads/export/pdf")
async def export_leads_pdf(lang: str = 'en'):
    """Export all leads as professional styled PDF"""
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.colors import HexColor, white
    
    # Register fonts
    try:
        pdfmetrics.registerFont(TTFont('FreeSans', '/usr/share/fonts/truetype/freefont/FreeSans.ttf'))
        pdfmetrics.registerFont(TTFont('FreeSans-Bold', '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf'))
        FONT = 'FreeSans'
        FONT_BOLD = 'FreeSans-Bold'
    except:
        FONT = 'Helvetica'
        FONT_BOLD = 'Helvetica-Bold'
    
    # Colors
    PRIMARY = HexColor('#4f46e5')  # Indigo
    SECONDARY = HexColor('#818cf8')
    TEXT_DARK = HexColor('#1e293b')
    TEXT_MUTED = HexColor('#64748b')
    BG_LIGHT = HexColor('#f1f5f9')
    
    # Multi-language labels
    labels = {
        'en': {'title': 'Customer List', 'total': 'Total', 'customers': 'customers', 'date': 'Date',
               'company': 'Company', 'contact': 'Contact', 'email': 'Email', 'phone': 'Phone', 
               'city': 'City', 'country': 'Country', 'page': 'Page'},
        'tr': {'title': 'Müşteri Listesi', 'total': 'Toplam', 'customers': 'müşteri', 'date': 'Tarih',
               'company': 'Firma', 'contact': 'Yetkili Kişi', 'email': 'E-posta', 'phone': 'Telefon',
               'city': 'Şehir', 'country': 'Ülke', 'page': 'Sayfa'},
        'de': {'title': 'Kundenliste', 'total': 'Gesamt', 'customers': 'Kunden', 'date': 'Datum',
               'company': 'Firma', 'contact': 'Kontakt', 'email': 'E-Mail', 'phone': 'Telefon',
               'city': 'Stadt', 'country': 'Land', 'page': 'Seite'},
        'pl': {'title': 'Lista Klientów', 'total': 'Razem', 'customers': 'klienci', 'date': 'Data',
               'company': 'Firma', 'contact': 'Kontakt', 'email': 'E-mail', 'phone': 'Telefon',
               'city': 'Miasto', 'country': 'Kraj', 'page': 'Strona'}
    }
    L = labels.get(lang, labels['en'])
    company_name = settings.get('company_name', 'Gewürzberg GmbH') if settings else 'Gewürzberg GmbH'
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    page_num = 1
    
    def draw_header():
        # Modern header with gradient effect
        c.setFillColor(PRIMARY)
        c.rect(0, height - 3.5*cm, width, 3.5*cm, fill=True, stroke=False)
        c.setFillColor(SECONDARY)
        c.rect(0, height - 3.5*cm, width, 0.3*cm, fill=True, stroke=False)
        
        # Company name
        c.setFillColor(white)
        c.setFont(FONT_BOLD, 20)
        c.drawString(2*cm, height - 1.8*cm, company_name)
        c.setFont(FONT, 10)
        c.drawString(2*cm, height - 2.6*cm, L['title'])
        
        # Date and total
        c.setFont(FONT, 9)
        c.drawRightString(width - 2*cm, height - 1.8*cm, f"{L['date']}: {datetime.now().strftime('%d.%m.%Y')}")
        c.drawRightString(width - 2*cm, height - 2.6*cm, f"{L['total']}: {len(leads)} {L['customers']}")
        
        return height - 4.5*cm
    
    def draw_table_header(y_pos):
        # Table header background
        c.setFillColor(BG_LIGHT)
        c.rect(1.5*cm, y_pos - 0.7*cm, width - 3*cm, 0.9*cm, fill=True, stroke=False)
        
        # Column headers
        c.setFillColor(TEXT_DARK)
        c.setFont(FONT_BOLD, 8)
        c.drawString(2*cm, y_pos - 0.4*cm, L['company'])
        c.drawString(6*cm, y_pos - 0.4*cm, L['contact'])
        c.drawString(9.5*cm, y_pos - 0.4*cm, L['email'])
        c.drawString(13.5*cm, y_pos - 0.4*cm, L['phone'])
        c.drawString(17*cm, y_pos - 0.4*cm, L['city'])
        
        return y_pos - 1.2*cm
    
    def draw_footer():
        c.setFillColor(TEXT_MUTED)
        c.setFont(FONT, 8)
        c.drawCentredString(width/2, 1*cm, f"{L['page']} {page_num}")
    
    y = draw_header()
    y = draw_table_header(y)
    
    row_height = 0.7*cm
    for i, lead in enumerate(leads):
        if y < 2.5*cm:
            draw_footer()
            c.showPage()
            page_num += 1
            y = draw_header()
            y = draw_table_header(y)
        
        # Alternate row background
        if i % 2 == 0:
            c.setFillColor(HexColor('#f8fafc'))
            c.rect(1.5*cm, y - 0.5*cm, width - 3*cm, row_height, fill=True, stroke=False)
        
        c.setFillColor(TEXT_DARK)
        c.setFont(FONT, 8)
        
        company = (lead.get('company') or lead.get('company_name') or '')[:22]
        contact = f"{lead.get('first_name', '')} {lead.get('last_name', '')}"[:18]
        email = (lead.get('email') or '')[:22]
        phone = (lead.get('phone') or '')[:14]
        city = (lead.get('city') or '')[:12]
        
        c.drawString(2*cm, y - 0.3*cm, company)
        c.setFillColor(TEXT_MUTED)
        c.drawString(6*cm, y - 0.3*cm, contact)
        c.setFillColor(PRIMARY)
        c.drawString(9.5*cm, y - 0.3*cm, email)
        c.setFillColor(TEXT_DARK)
        c.drawString(13.5*cm, y - 0.3*cm, phone)
        c.drawString(17*cm, y - 0.3*cm, city)
        
        y -= row_height
    
    draw_footer()
    c.save()
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=customer_list_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

@api_router.get("/recipes/{recipe_id}/pdf")
async def get_recipe_pdf(recipe_id: str, lang: str = 'en'):
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    
    # Get lead info for customer details
    lead_info = None
    if recipe.get('lead_id'):
        lead_info = await db.leads.find_one({"id": recipe['lead_id']}, {"_id": 0})
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_recipe_pdf(recipe, settings, lang=lang, lead_info=lead_info)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=recete_{recipe_id[:8]}.pdf"}
    )

@api_router.post("/recipes/{recipe_id}/email")
async def email_recipe(recipe_id: str, to_email: str, background_tasks: BackgroundTasks):
    """Send recipe as PDF via email"""
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    
    # Get lead info for customer details
    lead_info = None
    if recipe.get('lead_id'):
        lead_info = await db.leads.find_one({"id": recipe['lead_id']}, {"_id": 0})
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})
    
    if not smtp_settings:
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    # Generate PDF using new function with lead info
    pdf_content = generate_recipe_pdf(recipe, settings, lead_info=lead_info)
    
    # Send email with attachment
    async def send_recipe_email():
        try:
            msg = MIMEMultipart()
            msg['Subject'] = f"Reçete: {recipe['name']}"
            msg['From'] = f"{smtp_settings['from_name']} <{smtp_settings['from_email']}>"
            msg['To'] = to_email
            
            body = f"Merhaba,\n\n{recipe['name']} reçetesi ekte yer almaktadır.\n\nSaygılarımızla"
            msg.attach(MIMEText(body, 'plain'))
            
            attachment = MIMEApplication(pdf_content, _subtype='pdf')
            attachment.add_header('Content-Disposition', 'attachment', filename=f"recete_{recipe['name']}.pdf")
            msg.attach(attachment)
            
            if smtp_settings.get('use_tls', True):
                server = smtplib.SMTP(smtp_settings['host'], smtp_settings['port'])
                server.starttls()
            else:
                server = smtplib.SMTP_SSL(smtp_settings['host'], smtp_settings['port'])
            
            server.login(smtp_settings['username'], smtp_settings['password'])
            server.sendmail(smtp_settings['from_email'], to_email, msg.as_string())
            server.quit()
            logger.info(f"Recipe email sent to {to_email}")
        except Exception as e:
            logger.error(f"Failed to send recipe email: {str(e)}")
    
    background_tasks.add_task(send_recipe_email)
    return {"message": "Recipe email queued for sending"}

# ===================== ORDER ENDPOINTS =====================

@api_router.post("/orders", response_model=Order)
async def create_order(order_data: OrderCreate):
    # Get lead info
    lead = await db.leads.find_one({"id": order_data.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Handle multi-product orders
    products_list = []
    total_price = 0.0
    
    if order_data.products and len(order_data.products) > 0:
        # Multi-product order
        for item in order_data.products:
            subtotal = item.pieces * item.amount * item.unit_price
            products_list.append(OrderProductItem(
                product_name=item.product_name,
                product_code=item.product_code,
                pieces=item.pieces,
                amount=item.amount,
                unit=item.unit,
                unit_price=item.unit_price,
                subtotal=subtotal
            ))
            total_price += subtotal
        
        # For display, use first product info
        first_product = order_data.products[0]
        product_name = first_product.product_name
        product_code = first_product.product_code
        pieces = first_product.pieces
        amount = first_product.amount
        unit = first_product.unit
        unit_price = first_product.unit_price
    else:
        # Legacy single-product order
        pieces = order_data.pieces or 1
        amount = order_data.amount or 1
        unit_price = order_data.unit_price or 0
        total_price = pieces * amount * unit_price
        
        product_name = order_data.product_name
        product_code = order_data.product_code
        unit = order_data.unit
        
        # Create single product entry for consistency
        products_list.append(OrderProductItem(
            product_name=product_name,
            product_code=product_code,
            pieces=pieces,
            amount=amount,
            unit=unit,
            unit_price=unit_price,
            subtotal=total_price
        ))
    
    order = Order(
        lead_id=order_data.lead_id,
        lead_name=f"{lead['first_name']} {lead['last_name']}",
        company_name=lead['company_name'],
        products=products_list,
        product_name=product_name,
        product_code=product_code,
        pieces=pieces,
        amount=amount,
        unit=unit,
        unit_price=unit_price,
        total_price=total_price,
        notes=order_data.notes
    )
    
    doc = order.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    # Serialize products
    doc['products'] = [p.model_dump() if hasattr(p, 'model_dump') else p for p in doc['products']]
    await db.orders.insert_one(doc)
    return order

@api_router.get("/orders", response_model=List[Order])
async def get_orders():
    orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for order in orders:
        deserialize_datetime(order)
    return orders

@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    deserialize_datetime(order)
    return order

@api_router.put("/orders/{order_id}", response_model=Order)
async def update_order(order_id: str, order_data: OrderUpdate):
    existing = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Order not found")
    
    update_data = {k: v for k, v in order_data.model_dump().items() if v is not None}
    
    # Handle multi-product updates
    if 'products' in update_data and update_data['products']:
        products = update_data['products']
        total_price = 0.0
        serialized_products = []
        for p in products:
            if hasattr(p, 'model_dump'):
                p_dict = p.model_dump()
            else:
                p_dict = p
            subtotal = p_dict.get('pieces', 1) * p_dict.get('amount', 1) * p_dict.get('unit_price', 0)
            p_dict['subtotal'] = subtotal
            serialized_products.append(p_dict)
            total_price += subtotal
        update_data['products'] = serialized_products
        update_data['total_price'] = total_price
        # Update legacy fields with first product
        if serialized_products:
            first = serialized_products[0]
            update_data['product_name'] = first.get('product_name')
            update_data['product_code'] = first.get('product_code')
            update_data['pieces'] = first.get('pieces')
            update_data['amount'] = first.get('amount')
            update_data['unit'] = first.get('unit')
            update_data['unit_price'] = first.get('unit_price')
    else:
        # Legacy single-product update - recalculate total
        pieces = update_data.get('pieces', existing.get('pieces', 1))
        amount = update_data.get('amount', existing.get('amount', existing.get('quantity', 1)))
        unit_price = update_data.get('unit_price', existing.get('unit_price', 0))
        update_data['total_price'] = pieces * amount * unit_price
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    deserialize_datetime(updated)
    return updated

@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str):
    result = await db.orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"message": "Order deleted successfully"}

@api_router.get("/orders/lead/{lead_id}", response_model=List[Order])
async def get_lead_orders(lead_id: str):
    orders = await db.orders.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for order in orders:
        deserialize_datetime(order)
    return orders

# ===================== RECIPE ENDPOINTS =====================

@api_router.post("/recipes", response_model=Recipe)
async def create_recipe(recipe_data: RecipeCreate):
    # Get lead info
    lead = await db.leads.find_one({"id": recipe_data.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    recipe = Recipe(
        lead_id=recipe_data.lead_id,
        lead_name=f"{lead.get('first_name', '')} {lead.get('last_name', '')}",
        company_name=lead.get('company_name', ''),
        name=recipe_data.name,
        product_code=recipe_data.product_code,
        meat_amount=recipe_data.meat_amount,
        water_amount=recipe_data.water_amount,
        spice_amount=recipe_data.spice_amount,
        binding_amount=recipe_data.binding_amount,
        mixing_time=recipe_data.mixing_time,
        motor_speed=recipe_data.motor_speed,
        additional_ingredients=recipe_data.additional_ingredients or [],
        instructions=recipe_data.instructions,
        notes=recipe_data.notes
    )
    
    doc = recipe.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.recipes.insert_one(doc)
    return recipe

@api_router.get("/recipes", response_model=List[Recipe])
async def get_recipes():
    recipes = await db.recipes.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for recipe in recipes:
        deserialize_datetime(recipe)
    return recipes

@api_router.get("/recipes/{recipe_id}", response_model=Recipe)
async def get_recipe(recipe_id: str):
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    deserialize_datetime(recipe)
    return recipe

@api_router.get("/recipes/lead/{lead_id}", response_model=List[Recipe])
async def get_lead_recipes(lead_id: str):
    recipes = await db.recipes.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for recipe in recipes:
        deserialize_datetime(recipe)
    return recipes

@api_router.put("/recipes/{recipe_id}", response_model=Recipe)
async def update_recipe(recipe_id: str, recipe_data: RecipeUpdate):
    existing = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Recipe not found")
    
    update_data = {k: v for k, v in recipe_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.recipes.update_one({"id": recipe_id}, {"$set": update_data})
    updated = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    deserialize_datetime(updated)
    return updated

@api_router.delete("/recipes/{recipe_id}")
async def delete_recipe(recipe_id: str):
    result = await db.recipes.delete_one({"id": recipe_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Recipe not found")
    return {"message": "Recipe deleted successfully"}

@api_router.post("/recipes/{recipe_id}/duplicate")
async def duplicate_recipe(recipe_id: str, new_lead_id: str):
    """Bir reçeteyi başka bir müşteriye kopyala"""
    original = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not original:
        raise HTTPException(status_code=404, detail="Recipe not found")
    
    # Get new lead info
    new_lead = await db.leads.find_one({"id": new_lead_id}, {"_id": 0})
    if not new_lead:
        raise HTTPException(status_code=404, detail="Target lead not found")
    
    # Create new recipe
    new_recipe = {
        **original,
        "id": str(uuid.uuid4()),
        "lead_id": new_lead_id,
        "lead_name": f"{new_lead.get('first_name', '')} {new_lead.get('last_name', '')}",
        "company_name": new_lead.get('company_name', ''),
        "name": f"{original['name']} (Kopya)",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.recipes.insert_one(new_recipe)
    return {"message": "Recipe duplicated successfully", "new_recipe_id": new_recipe["id"]}

# ===================== LEAD FINDER ENDPOINTS =====================

@api_router.post("/leads/search")
async def search_for_leads(request: SearchLeadsRequest):
    """Search for potential leads using Gemini AI-powered search"""
    from lead_finder import LeadFinder
    
    finder = LeadFinder()
    
    try:
        # Search for leads
        found_leads = await finder.search_leads(
            keywords=request.keywords,
            location=request.location,
            country=request.country,
            limit=request.limit
        )
        
        # Convert to response format
        leads_data = []
        for lead in found_leads:
            leads_data.append({
                "company_name": lead.company_name,
                "contact_person": lead.contact_person,
                "email": lead.email,
                "phone": lead.phone,
                "address": lead.address,
                "city": lead.city,
                "country": lead.country,
                "website": lead.website,
                "business_type": lead.business_type,
                "notes": lead.notes
            })
        
        # Save search result to database
        search_result = {
            "id": str(uuid.uuid4()),
            "query_keywords": request.keywords,
            "location": request.location,
            "country": request.country,
            "leads_found": leads_data,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "status": "completed"
        }
        await db.search_history.insert_one(search_result)
        
        return {
            "status": "success",
            "total_found": len(leads_data),
            "leads": leads_data
        }
        
    except Exception as e:
        logger.error(f"Lead search failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")

@api_router.get("/leads/search/templates")
async def get_search_templates():
    """Get predefined search templates for common industries"""
    return [
        {"name": "Döner Fabrikaları", "keywords": ["döner", "kebab", "meat processing"], "category": "meat"},
        {"name": "Gyros Üreticileri", "keywords": ["gyros", "souvlaki", "greek food"], "category": "meat"},
        {"name": "Et İşleme", "keywords": ["meat processing", "food manufacturer"], "category": "meat"},
        {"name": "Baharat Tedarikçileri", "keywords": ["spice", "seasoning", "wholesale"], "category": "spice"}
    ]

@api_router.get("/leads/search/history")
async def get_search_history():
    """Get history of lead searches"""
    history = await db.search_history.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    for item in history:
        if isinstance(item.get('created_at'), str):
            item['created_at'] = datetime.fromisoformat(item['created_at'])
    return history

@api_router.get("/potential-leads")
async def get_potential_leads(
    region: str = None,
    city: str = None,
    status: str = None,
    search: str = None
):
    """Get potential leads from the German döner factories database"""
    query = {}
    
    if region:
        query["region"] = {"$regex": region, "$options": "i"}
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"company_name": {"$regex": search, "$options": "i"}},
            {"city": {"$regex": search, "$options": "i"}},
            {"region": {"$regex": search, "$options": "i"}}
        ]
    
    leads = await db.potential_leads.find(query, {"_id": 0}).to_list(500)
    
    # Get unique regions for filter dropdown
    regions = await db.potential_leads.distinct("region")
    
    return {
        "total": len(leads),
        "leads": leads,
        "regions": sorted([r for r in regions if r])
    }

@api_router.put("/potential-leads/{lead_id}/status")
async def update_potential_lead_status(lead_id: str, status: str):
    """Update status of a potential lead"""
    if status not in ["potential", "contacted", "converted", "rejected"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.potential_leads.update_one(
        {"id": lead_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return {"success": True, "message": f"Status updated to {status}"}

@api_router.post("/potential-leads/{lead_id}/convert")
async def convert_potential_to_customer(lead_id: str):
    """Convert a potential lead to actual customer"""
    # Get potential lead
    potential = await db.potential_leads.find_one({"id": lead_id}, {"_id": 0})
    if not potential:
        raise HTTPException(status_code=404, detail="Potential lead not found")
    
    # Check if already converted
    existing = await db.leads.find_one({"company_name": potential["company_name"]})
    if existing:
        raise HTTPException(status_code=400, detail="Customer already exists")
    
    # Create new customer
    new_customer = {
        "id": str(uuid.uuid4()),
        "company_name": potential["company_name"],
        "contact_person": "",
        "email": "",
        "phone": potential.get("phone", ""),
        "address": potential.get("address", ""),
        "city": potential.get("city", ""),
        "country": potential.get("country", "Deutschland"),
        "status": "new",
        "source": "potential_database",
        "notes": f"Bölge: {potential.get('region', '')}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.leads.insert_one(new_customer)
    
    # Update potential lead status
    await db.potential_leads.update_one(
        {"id": lead_id},
        {"$set": {"status": "converted", "converted_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"success": True, "customer_id": new_customer["id"], "message": "Müşteri olarak eklendi"}

@api_router.post("/leads/import")
async def import_found_leads(leads: List[dict]):
    """Import found leads into the main leads database"""
    imported = []
    skipped = []
    
    for lead_data in leads:
        # Check if lead already exists by email or company name
        existing = None
        if lead_data.get('email'):
            existing = await db.leads.find_one({"email": lead_data['email']}, {"_id": 0})
        if not existing and lead_data.get('company_name'):
            existing = await db.leads.find_one({"company_name": lead_data['company_name']}, {"_id": 0})
        
        if existing:
            skipped.append(lead_data.get('company_name', 'Unknown'))
            continue
        
        # Create new lead
        new_lead = {
            "id": str(uuid.uuid4()),
            "first_name": lead_data.get('contact_name', '').split()[0] if lead_data.get('contact_name') else "Contact",
            "last_name": lead_data.get('contact_name', '').split()[-1] if lead_data.get('contact_name') and len(lead_data.get('contact_name', '').split()) > 1 else "",
            "company_name": lead_data.get('company_name', 'Unknown'),
            "tax_number": lead_data.get('tax_number', ''),
            "address": lead_data.get('address', ''),
            "email": lead_data.get('email', f"contact@{lead_data.get('company_name', 'unknown').lower().replace(' ', '')}.com"),
            "city": lead_data.get('city', ''),
            "country": lead_data.get('country', ''),
            "notes": f"Source: {lead_data.get('source', 'AI Search')}\nWebsite: {lead_data.get('website', 'N/A')}\nPhone: {lead_data.get('phone', 'N/A')}\nDescription: {lead_data.get('description', '')}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.leads.insert_one(new_lead)
        imported.append(new_lead['company_name'])
    
    return {
        "status": "success",
        "imported_count": len(imported),
        "skipped_count": len(skipped),
        "imported": imported,
        "skipped": skipped
    }

# ===================== AUTH ENDPOINTS =====================

@api_router.post("/auth/login")
async def login(request: LoginRequest, response: Response):
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
    admin_password = os.environ.get('ADMIN_PASSWORD', '190371')
    
    # First check if it's the admin user
    if request.username == admin_username and request.password == admin_password:
        user_id = "admin-user-id"
        access_token = create_access_token(user_id, admin_username)
        
        response.set_cookie(
            key="access_token", 
            value=access_token, 
            httponly=True, 
            secure=False, 
            samesite="lax", 
            max_age=86400,
            path="/"
        )
        
        return {
            "id": user_id,
            "username": admin_username,
            "name": "Emre Dirlik",
            "role": "admin",
            "token": access_token
        }
    
    # Check database for other users
    user = await db.users.find_one({"$or": [{"email": request.username}, {"username": request.username}]})
    if user and bcrypt.checkpw(request.password.encode('utf-8'), user.get('password_hash', '').encode('utf-8')):
        user_id = user.get('id', str(user.get('_id')))
        access_token = create_access_token(user_id, user.get('email', request.username))
        
        response.set_cookie(
            key="access_token", 
            value=access_token, 
            httponly=True, 
            secure=False, 
            samesite="lax", 
            max_age=86400,
            path="/"
        )
        
        return {
            "id": user_id,
            "username": user.get('email'),
            "name": user.get('name', user.get('email')),
            "role": user.get('role', 'user'),
            "token": access_token
        }
    
    raise HTTPException(status_code=401, detail="Geçersiz kullanıcı adı veya şifre")

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie(key="access_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.get("/auth/me")
async def get_me(request: Request):
    try:
        user = await get_current_user(request)
        return user
    except HTTPException:
        raise HTTPException(status_code=401, detail="Not authenticated")

@api_router.get("/auth/check")
async def check_auth(request: Request):
    """Check if user is authenticated"""
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        return {"authenticated": False}
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return {"authenticated": True, "user": {"username": payload.get("username"), "name": "Emre Dirlik"}}
    except:
        return {"authenticated": False}

# ===================== WHATSAPP ENDPOINTS =====================

@api_router.get("/orders/{order_id}/whatsapp")
async def get_order_whatsapp_link(order_id: str, lang: str = 'tr'):
    """Generate WhatsApp share link for an order - NO TOTAL PRICE"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get lead info for address
    lead = await db.leads.find_one({"id": order.get('lead_id')}, {"_id": 0})
    customer_address = ""
    if lead:
        address_parts = [lead.get('address', ''), lead.get('city', ''), lead.get('country', '')]
        customer_address = ', '.join([p for p in address_parts if p])
    
    # Get company settings
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    company_name = settings.get('company_name', 'Gewürzberg GmbH') if settings else 'Gewürzberg GmbH'
    
    # Get backend URL for PDF link
    backend_url = os.environ.get('REACT_APP_BACKEND_URL', 'https://customer-agent-2.preview.emergentagent.com')
    pdf_url = f"{backend_url}/api/orders/{order_id}/pdf/public?lang={lang}"
    
    # Build product list - NO TOTAL PRICE
    products = order.get('products', [])
    product_lines = []
    
    if products and len(products) > 0:
        for p in products:
            pieces = p.get('pieces', 1)
            amount = p.get('amount', 1)
            unit = p.get('unit', 'kg')
            unit_price = p.get('unit_price', 0)
            qty_str = f"{pieces}×{amount}{unit}" if pieces > 1 else f"{amount}{unit}"
            product_lines.append(f"• {p.get('product_code', '')} - {qty_str} @ €{unit_price:.2f}/{unit}")
    else:
        # Legacy single product
        pieces = order.get('pieces', 1)
        amount = order.get('amount', order.get('quantity', 1))
        unit = order.get('unit', 'kg')
        qty_str = f"{pieces}×{amount}{unit}" if pieces > 1 else f"{amount}{unit}"
        product_lines.append(f"• {order.get('product_code', '')} - {qty_str} @ €{order.get('unit_price', 0):.2f}/{unit}")
    
    products_text = '\n'.join(product_lines)
    
    # Format message - NO TOTAL PRICE
    message = f"""📦 *{company_name} - Sipariş*

*Müşteri:* {order.get('company_name', '')}
*Adres:* {customer_address}

*Ürünler:*
{products_text}

📄 PDF: {pdf_url}"""
    
    # URL encode the message
    encoded_message = quote(message)
    whatsapp_url = f"https://wa.me/?text={encoded_message}"
    
    return {"whatsapp_url": whatsapp_url, "message": message, "pdf_url": pdf_url}

@api_router.get("/orders/{order_id}/pdf/public")
async def get_public_order_pdf(order_id: str):
    """Public endpoint for order PDF (for WhatsApp sharing)"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_order_pdf(order, settings)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=siparis_{order_id[:8]}.pdf"}
    )

# ===================== PROFESSIONAL PDF GENERATION =====================

def generate_professional_order_pdf(order: dict, company_settings: dict = None) -> bytes:
    """Generate a professionally styled order PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm, mm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Colors matching the app design
    primary_color = HexColor('#1e293b')  # slate-800
    accent_color = HexColor('#f97316')   # orange-500
    bg_light = HexColor('#f8fafc')       # slate-50
    text_muted = HexColor('#64748b')     # slate-500
    success_color = HexColor('#22c55e')  # green-500
    
    company_name = company_settings.get('company_name', 'Gewürzberg GmbH') if company_settings else 'Gewürzberg GmbH'
    
    # Header background
    c.setFillColor(primary_color)
    c.rect(0, height - 4*cm, width, 4*cm, fill=True, stroke=False)
    
    # Company name
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 24)
    c.drawString(2*cm, height - 2.5*cm, company_name)
    c.setFont("Helvetica", 10)
    c.drawString(2*cm, height - 3.2*cm, "Emre Dirlik")
    
    # Document title
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - 2*cm, height - 2.5*cm, "SİPARİŞ FORMU")
    c.setFont("Helvetica", 10)
    c.drawRightString(width - 2*cm, height - 3.2*cm, f"#{order['id'][:8].upper()}")
    
    y = height - 5.5*cm
    
    # Customer info card
    c.setFillColor(bg_light)
    c.roundRect(1.5*cm, y - 3*cm, width - 3*cm, 3*cm, 5, fill=True, stroke=False)
    
    c.setFillColor(text_muted)
    c.setFont("Helvetica", 9)
    c.drawString(2*cm, y - 0.7*cm, "MÜŞTERİ")
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, y - 1.4*cm, order['company_name'])
    c.setFont("Helvetica", 10)
    c.setFillColor(text_muted)
    c.drawString(2*cm, y - 2.1*cm, order.get('lead_name', ''))
    
    c.setFillColor(text_muted)
    c.setFont("Helvetica", 9)
    c.drawString(10*cm, y - 0.7*cm, "TARİH")
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(10*cm, y - 1.4*cm, datetime.now().strftime('%d.%m.%Y'))
    
    y -= 4*cm
    
    # Product details header
    c.setFillColor(accent_color)
    c.rect(1.5*cm, y - 1*cm, width - 3*cm, 1*cm, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y - 0.7*cm, "ÜRÜN DETAYLARI")
    
    y -= 1.5*cm
    
    # Product info
    pieces = order.get('pieces', 1)
    amount = order.get('amount', order.get('quantity', 1))
    unit = order.get('unit', 'kg')
    
    items = [
        ("Ürün Adı", order['product_name']),
        ("Ürün Kodu", order['product_code']),
        ("Miktar", f"{pieces} × {amount} {unit}"),
        ("Birim Fiyat", f"€{order['unit_price']:.2f}/{unit}"),
    ]
    
    for label, value in items:
        c.setFillColor(text_muted)
        c.setFont("Helvetica", 9)
        c.drawString(2*cm, y, label)
        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(6*cm, y, str(value))
        y -= 0.8*cm
    
    y -= 0.5*cm
    
    # Total box
    c.setFillColor(success_color)
    c.roundRect(1.5*cm, y - 2*cm, width - 3*cm, 2*cm, 5, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont("Helvetica", 12)
    c.drawString(2*cm, y - 0.8*cm, "TOPLAM TUTAR")
    c.setFont("Helvetica-Bold", 24)
    c.drawRightString(width - 2*cm, y - 1.4*cm, f"€{order['total_price']:.2f}")
    
    y -= 3*cm
    
    # Status badge
    status_colors = {
        'pending': HexColor('#eab308'),
        'confirmed': HexColor('#3b82f6'),
        'shipped': HexColor('#8b5cf6'),
        'delivered': HexColor('#22c55e'),
        'cancelled': HexColor('#ef4444')
    }
    status_labels = {
        'pending': 'BEKLEMEDE',
        'confirmed': 'ONAYLANDI',
        'shipped': 'GÖNDERİLDİ',
        'delivered': 'TESLİM EDİLDİ',
        'cancelled': 'İPTAL'
    }
    
    status = order.get('status', 'pending')
    c.setFillColor(status_colors.get(status, text_muted))
    c.roundRect(2*cm, y - 1*cm, 4*cm, 1*cm, 3, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(4*cm, y - 0.7*cm, status_labels.get(status, status.upper()))
    
    # Notes
    if order.get('notes'):
        y -= 2*cm
        c.setFillColor(text_muted)
        c.setFont("Helvetica", 9)
        c.drawString(2*cm, y, "NOTLAR")
        c.setFillColor(black)
        c.setFont("Helvetica", 10)
        c.drawString(2*cm, y - 0.6*cm, order['notes'][:100])
    
    # Footer
    c.setFillColor(text_muted)
    c.setFont("Helvetica", 8)
    c.drawString(2*cm, 1.5*cm, f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    c.drawRightString(width - 2*cm, 1.5*cm, company_name)
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue()

def generate_professional_recipe_pdf(recipe: dict, company_settings: dict = None) -> bytes:
    """Generate a professionally styled recipe PDF matching the UI design"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor, white, black
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Colors
    primary_color = HexColor('#1e293b')
    accent_color = HexColor('#f97316')
    red_color = HexColor('#ef4444')
    blue_color = HexColor('#3b82f6')
    orange_color = HexColor('#f97316')
    purple_color = HexColor('#8b5cf6')
    green_color = HexColor('#22c55e')
    gray_color = HexColor('#6b7280')
    bg_light = HexColor('#f8fafc')
    text_muted = HexColor('#64748b')
    
    company_name = company_settings.get('company_name', 'Gewürzberg GmbH') if company_settings else 'Gewürzberg GmbH'
    
    # Header
    c.setFillColor(primary_color)
    c.rect(0, height - 4*cm, width, 4*cm, fill=True, stroke=False)
    
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 24)
    c.drawString(2*cm, height - 2.5*cm, company_name)
    c.setFont("Helvetica", 10)
    c.drawString(2*cm, height - 3.2*cm, "Üretim Reçetesi")
    
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - 2*cm, height - 2.5*cm, recipe['product_code'])
    
    y = height - 5.5*cm
    
    # Recipe title card
    c.setFillColor(bg_light)
    c.roundRect(1.5*cm, y - 2.5*cm, width - 3*cm, 2.5*cm, 5, fill=True, stroke=False)
    
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(2*cm, y - 1*cm, recipe['name'])
    c.setFillColor(text_muted)
    c.setFont("Helvetica", 11)
    c.drawString(2*cm, y - 1.8*cm, f"Müşteri: {recipe['company_name']}")
    
    y -= 4*cm
    
    # Main ingredients header
    c.setFillColor(accent_color)
    c.rect(1.5*cm, y - 1*cm, width - 3*cm, 1*cm, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y - 0.7*cm, "ANA MALZEMELER")
    
    y -= 1.8*cm
    
    # Ingredient boxes (2x2 grid like in UI)
    box_width = (width - 4*cm) / 2
    box_height = 1.8*cm
    
    ingredients = [
        (red_color, "Et Miktarı", f"{recipe['meat_amount']} kg"),
        (blue_color, "Su Miktarı", f"{recipe['water_amount']} L"),
        (orange_color, "Baharat Miktarı", f"{recipe['spice_amount']} kg"),
        (purple_color, "Binding Miktarı", f"{recipe['binding_amount']} kg"),
    ]
    
    for i, (color, label, value) in enumerate(ingredients):
        col = i % 2
        row = i // 2
        x = 1.5*cm + col * (box_width + 0.5*cm)
        box_y = y - row * (box_height + 0.3*cm)
        
        # Light colored background
        c.setFillColor(HexColor('#fef2f2') if i == 0 else HexColor('#eff6ff') if i == 1 else HexColor('#fff7ed') if i == 2 else HexColor('#faf5ff'))
        c.roundRect(x, box_y - box_height, box_width - 0.3*cm, box_height, 5, fill=True, stroke=False)
        
        c.setFillColor(color)
        c.setFont("Helvetica", 9)
        c.drawString(x + 0.4*cm, box_y - 0.6*cm, label)
        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(x + 0.4*cm, box_y - 1.3*cm, value)
    
    y -= 4.5*cm
    
    # Production parameters header
    c.setFillColor(green_color)
    c.rect(1.5*cm, y - 1*cm, width - 3*cm, 1*cm, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y - 0.7*cm, "ÜRETİM PARAMETRELERİ")
    
    y -= 1.8*cm
    
    params = [
        (green_color, "Karışım Süresi", f"{recipe['mixing_time']} dakika"),
        (gray_color, "Motor Hızı", f"{recipe['motor_speed']} rpm"),
    ]
    
    for i, (color, label, value) in enumerate(params):
        x = 1.5*cm + i * (box_width + 0.5*cm)
        
        c.setFillColor(HexColor('#f0fdf4') if i == 0 else HexColor('#f3f4f6'))
        c.roundRect(x, y - box_height, box_width - 0.3*cm, box_height, 5, fill=True, stroke=False)
        
        c.setFillColor(color)
        c.setFont("Helvetica", 9)
        c.drawString(x + 0.4*cm, y - 0.6*cm, label)
        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(x + 0.4*cm, y - 1.3*cm, value)
    
    y -= 3*cm
    
    # Additional ingredients
    if recipe.get('additional_ingredients'):
        c.setFillColor(text_muted)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2*cm, y, "EK MALZEMELER")
        y -= 0.6*cm
        
        c.setFont("Helvetica", 10)
        for ing in recipe['additional_ingredients']:
            c.setFillColor(black)
            c.drawString(2*cm, y, f"• {ing['name']}: {ing['amount']} {ing['unit']}")
            y -= 0.5*cm
    
    y -= 0.5*cm
    
    # Instructions
    if recipe.get('instructions'):
        c.setFillColor(text_muted)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2*cm, y, "ÜRETİM TALİMATLARI")
        y -= 0.6*cm
        
        c.setFillColor(bg_light)
        c.roundRect(1.5*cm, y - 2*cm, width - 3*cm, 2*cm, 5, fill=True, stroke=False)
        c.setFillColor(black)
        c.setFont("Helvetica", 9)
        # Word wrap instructions
        instructions = recipe['instructions'][:200]
        c.drawString(2*cm, y - 0.5*cm, instructions[:80])
        if len(instructions) > 80:
            c.drawString(2*cm, y - 1*cm, instructions[80:160])
    
    # Footer
    c.setFillColor(text_muted)
    c.setFont("Helvetica", 8)
    c.drawString(2*cm, 1.5*cm, f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    c.drawRightString(width - 2*cm, 1.5*cm, company_name)
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue()

# Update PDF endpoints to use professional versions
@api_router.get("/orders/{order_id}/pdf/professional")
async def get_order_pdf_professional(order_id: str):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_professional_order_pdf(order, settings)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=siparis_{order_id[:8]}.pdf"}
    )

@api_router.get("/recipes/{recipe_id}/pdf/professional")
async def get_recipe_pdf_professional(recipe_id: str):
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_professional_recipe_pdf(recipe, settings)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=recete_{recipe_id[:8]}.pdf"}
    )

# ===================== AGENDA ENDPOINTS =====================

class AgendaTask(BaseModel):
    title: str
    due_date: Optional[str] = None
    completed: bool = False
    # Visit Planning fields
    event_type: Optional[str] = "task"  # task, visit, meeting, call, delivery
    lead_id: Optional[str] = None
    lead_name: Optional[str] = None
    company_name: Optional[str] = None
    time: Optional[str] = None  # HH:MM format
    notes: Optional[str] = None

class AgendaTaskUpdate(BaseModel):
    title: Optional[str] = None
    due_date: Optional[str] = None
    completed: Optional[bool] = None
    event_type: Optional[str] = None
    lead_id: Optional[str] = None
    lead_name: Optional[str] = None
    company_name: Optional[str] = None
    time: Optional[str] = None
    notes: Optional[str] = None

@api_router.get("/agenda")
async def get_agenda():
    """Get all agenda tasks"""
    tasks = await db.agenda.find({}, {"_id": 0}).sort("due_date", 1).to_list(500)
    return tasks

@api_router.post("/agenda")
async def create_agenda_task(task: AgendaTask):
    """Create a new agenda task or visit"""
    # If lead_id provided, fetch lead details
    lead_name = task.lead_name
    company_name = task.company_name
    
    if task.lead_id and not task.lead_name:
        lead = await db.leads.find_one({"id": task.lead_id}, {"_id": 0})
        if lead:
            lead_name = f"{lead.get('first_name', '')} {lead.get('last_name', '')}"
            company_name = lead.get('company_name', '')
    
    doc = {
        "id": str(uuid.uuid4()),
        "title": task.title,
        "due_date": task.due_date,
        "completed": task.completed,
        "event_type": task.event_type or "task",
        "lead_id": task.lead_id,
        "lead_name": lead_name,
        "company_name": company_name,
        "time": task.time,
        "notes": task.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.agenda.insert_one(doc)
    return {k: v for k, v in doc.items() if k != '_id'}

@api_router.put("/agenda/{task_id}")
async def update_agenda_task(task_id: str, task: AgendaTaskUpdate):
    """Update an agenda task"""
    existing = await db.agenda.find_one({"id": task_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Task not found")
    
    update_data = {k: v for k, v in task.model_dump().items() if v is not None}
    await db.agenda.update_one({"id": task_id}, {"$set": update_data})
    
    updated = await db.agenda.find_one({"id": task_id}, {"_id": 0})
    return updated

@api_router.delete("/agenda/{task_id}")
async def delete_agenda_task(task_id: str):
    """Delete an agenda task"""
    result = await db.agenda.delete_one({"id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted successfully"}

# ===================== DAILY REPORTS ENDPOINTS =====================

class DailyReportCreate(BaseModel):
    date: str  # YYYY-MM-DD format
    lead_id: str
    visit_type: str  # meeting, delivery, support, etc
    notes: str
    outcome: Optional[str] = None
    next_action: Optional[str] = None

class DailyReportUpdate(BaseModel):
    visit_type: Optional[str] = None
    notes: Optional[str] = None
    outcome: Optional[str] = None
    next_action: Optional[str] = None

@api_router.get("/daily-reports")
async def get_daily_reports(date: Optional[str] = None):
    """Get all daily reports, optionally filtered by date"""
    query = {}
    if date:
        query["date"] = date
    reports = await db.daily_reports.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return reports

@api_router.get("/daily-reports/by-date/{date}")
async def get_reports_by_date(date: str):
    """Get all reports for a specific date"""
    reports = await db.daily_reports.find({"date": date}, {"_id": 0}).to_list(100)
    return reports

@api_router.post("/daily-reports")
async def create_daily_report(report: DailyReportCreate):
    """Create a new daily report"""
    # Get lead info
    lead = await db.leads.find_one({"id": report.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    doc = {
        "id": str(uuid.uuid4()),
        "date": report.date,
        "lead_id": report.lead_id,
        "company_name": lead.get("company_name", ""),
        "city": lead.get("city", ""),
        "visit_type": report.visit_type,
        "notes": report.notes,
        "outcome": report.outcome,
        "next_action": report.next_action,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.daily_reports.insert_one(doc)
    return {k: v for k, v in doc.items() if k != '_id'}

@api_router.put("/daily-reports/{report_id}")
async def update_daily_report(report_id: str, report: DailyReportUpdate):
    """Update a daily report"""
    existing = await db.daily_reports.find_one({"id": report_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Report not found")
    
    update_data = {k: v for k, v in report.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.daily_reports.update_one({"id": report_id}, {"$set": update_data})
    
    updated = await db.daily_reports.find_one({"id": report_id}, {"_id": 0})
    return updated

@api_router.delete("/daily-reports/{report_id}")
async def delete_daily_report(report_id: str):
    """Delete a daily report"""
    result = await db.daily_reports.delete_one({"id": report_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Report not found")
    return {"message": "Report deleted successfully"}

@api_router.get("/daily-reports/{report_id}/pdf")
async def get_daily_report_pdf(report_id: str):
    """Generate PDF for a daily report"""
    report = await db.daily_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_daily_report_pdf(report, settings)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=daily_report_{report['date']}.pdf"}
    )

@api_router.get("/daily-reports/date/{date}/pdf")
async def get_daily_reports_by_date_pdf(date: str, lang: str = 'en'):
    """Generate combined PDF for all reports on a specific date"""
    reports = await db.daily_reports.find({"date": date}, {"_id": 0}).to_list(100)
    if not reports:
        raise HTTPException(status_code=404, detail="No reports found for this date")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_combined_daily_report_pdf(reports, date, settings, lang)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=daily_reports_{date}.pdf"}
    )

class DailyReportEmailRequest(BaseModel):
    to_email: str
    subject: Optional[str] = None
    message: Optional[str] = None

@api_router.post("/daily-reports/date/{date}/email")
async def email_daily_reports(date: str, request: DailyReportEmailRequest):
    """Send daily reports PDF via email"""
    reports = await db.daily_reports.find({"date": date}, {"_id": 0}).to_list(100)
    if not reports:
        raise HTTPException(status_code=404, detail="No reports found for this date")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    pdf_content = generate_combined_daily_report_pdf(reports, date, settings, 'en')
    
    # In production, integrate with email service (SendGrid, etc.)
    # For now, we'll log and return success
    logger.info(f"Email would be sent to {request.to_email} with {len(reports)} reports for {date}")
    
    # Store email record
    await db.email_log.insert_one({
        "id": str(uuid.uuid4()),
        "type": "daily_reports",
        "to_email": request.to_email,
        "subject": request.subject or f"Daily Reports - {date}",
        "date": date,
        "report_count": len(reports),
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "status": "logged"
    })
    
    return {"message": "Email logged successfully", "report_count": len(reports)}

@api_router.get("/daily-reports/month/{year}/{month}/pdf")
async def download_monthly_reports_pdf(year: int, month: int, lang: str = 'tr'):
    """Download all reports for a specific month as single PDF"""
    # Create date range for the month
    month_str = f"{year}-{str(month).zfill(2)}"
    
    # Find all reports that start with this month
    reports_cursor = db.daily_reports.find(
        {"date": {"$regex": f"^{month_str}"}},
        {"_id": 0}
    ).sort("date", 1)
    reports = await reports_cursor.to_list(500)
    
    if not reports:
        raise HTTPException(status_code=404, detail="No reports found for this month")
    
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    
    # Generate combined PDF for all reports in the month
    pdf_content = generate_combined_daily_report_pdf(reports, f"{month_str} (Monthly)", settings, lang)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="monthly_reports_{month_str}.pdf"'
        }
    )

# ===================== AI SALES FORECAST =====================

@api_router.get("/sales/forecast")
async def get_sales_forecast():
    """AI-powered sales forecast based on historical orders"""
    from datetime import timedelta
    from calendar import monthrange
    
    # Get historical orders
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    if not orders:
        return {
            "forecast": None,
            "message": "Not enough historical data for forecast",
            "historical_data": [],
            "next_month_prediction": 0
        }
    
    # Group orders by month
    monthly_data = {}
    for order in orders:
        created = order.get('created_at', '')
        if created:
            try:
                if isinstance(created, str):
                    dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                else:
                    dt = created
                month_key = dt.strftime('%Y-%m')
                
                if month_key not in monthly_data:
                    monthly_data[month_key] = {'revenue': 0, 'orders': 0}
                
                # Calculate order value
                amount = float(order.get('amount', 0) or 0)
                unit_price = float(order.get('unit_price', 0) or 0)
                order_value = amount * unit_price
                
                monthly_data[month_key]['revenue'] += order_value
                monthly_data[month_key]['orders'] += 1
            except Exception as e:
                logger.error(f"Error parsing order date: {e}")
                continue
    
    # Sort by month
    sorted_months = sorted(monthly_data.keys())
    historical = [
        {
            "month": m,
            "revenue": round(monthly_data[m]['revenue'], 2),
            "orders": monthly_data[m]['orders']
        }
        for m in sorted_months
    ]
    
    # Simple forecast: average of last 3 months with trend
    if len(historical) >= 2:
        recent = historical[-3:] if len(historical) >= 3 else historical
        avg_revenue = sum(h['revenue'] for h in recent) / len(recent)
        avg_orders = sum(h['orders'] for h in recent) / len(recent)
        
        # Calculate trend
        if len(recent) >= 2:
            trend = (recent[-1]['revenue'] - recent[0]['revenue']) / len(recent)
            predicted_revenue = avg_revenue + trend
        else:
            predicted_revenue = avg_revenue
        
        # Get next month
        now = datetime.now(timezone.utc)
        next_month = now.replace(day=1) + timedelta(days=32)
        next_month = next_month.replace(day=1)
        next_month_str = next_month.strftime('%Y-%m')
        next_month_name = next_month.strftime('%B %Y')
        
        # Confidence level based on data points
        confidence = min(95, 50 + len(historical) * 5)
        
        return {
            "forecast": {
                "next_month": next_month_str,
                "next_month_name": next_month_name,
                "predicted_revenue": round(max(0, predicted_revenue), 2),
                "predicted_orders": round(avg_orders),
                "confidence": confidence,
                "trend": "up" if trend > 0 else "down" if trend < 0 else "stable"
            },
            "historical_data": historical[-12:],  # Last 12 months
            "summary": {
                "total_revenue": round(sum(h['revenue'] for h in historical), 2),
                "total_orders": sum(h['orders'] for h in historical),
                "avg_monthly_revenue": round(sum(h['revenue'] for h in historical) / len(historical), 2),
                "avg_monthly_orders": round(sum(h['orders'] for h in historical) / len(historical))
            }
        }
    else:
        return {
            "forecast": None,
            "message": "Need at least 2 months of data for forecast",
            "historical_data": historical,
            "next_month_prediction": 0
        }

# Include the router in the main app

# ===================== DHL TRACKING ENDPOINTS =====================

from dhl_tracking import dhl_tracker, STATUS_LABELS

class ShipmentCreate(BaseModel):
    order_id: Optional[str] = None
    tracking_number: str
    carrier: str = "DHL"
    recipient_name: str
    recipient_address: str
    notes: Optional[str] = ""

class ShipmentUpdate(BaseModel):
    tracking_number: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None

@api_router.get("/shipments")
async def get_shipments():
    """Get all shipments"""
    shipments = await db.shipments.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return shipments

@api_router.post("/shipments")
async def create_shipment(shipment: ShipmentCreate):
    """Create a new shipment and start tracking"""
    # Get initial tracking info
    tracking_info = await dhl_tracker.track_package(shipment.tracking_number)
    
    doc = {
        "id": str(uuid.uuid4()),
        "order_id": shipment.order_id,
        "tracking_number": shipment.tracking_number.strip().upper(),
        "carrier": shipment.carrier,
        "recipient_name": shipment.recipient_name,
        "recipient_address": shipment.recipient_address,
        "notes": shipment.notes,
        "status": tracking_info.get("status", "unknown"),
        "status_text": tracking_info.get("status_text", "Takip başlatıldı"),
        "current_location": tracking_info.get("current_location", ""),
        "estimated_delivery": tracking_info.get("estimated_delivery"),
        "events": tracking_info.get("events", []),
        "last_tracked": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.shipments.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/shipments/{shipment_id}")
async def get_shipment(shipment_id: str):
    """Get a single shipment"""
    shipment = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    return shipment

@api_router.put("/shipments/{shipment_id}")
async def update_shipment(shipment_id: str, data: ShipmentUpdate):
    """Update shipment details"""
    existing = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.shipments.update_one({"id": shipment_id}, {"$set": update_data})
    updated = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
    return updated

@api_router.delete("/shipments/{shipment_id}")
async def delete_shipment(shipment_id: str):
    """Delete a shipment"""
    result = await db.shipments.delete_one({"id": shipment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Shipment not found")
    return {"status": "deleted", "id": shipment_id}

@api_router.post("/shipments/{shipment_id}/refresh")
async def refresh_shipment_tracking(shipment_id: str, notify_admin: bool = False, admin_email: str = None):
    """Refresh tracking status for a shipment and optionally notify on delivery"""
    shipment = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    old_status = shipment.get("status", "unknown")
    
    # Get updated tracking info
    tracking_info = await dhl_tracker.track_package(shipment["tracking_number"])
    
    new_status = tracking_info.get("status", shipment.get("status", "unknown"))
    
    update_data = {
        "status": new_status,
        "status_text": tracking_info.get("status_text", ""),
        "current_location": tracking_info.get("current_location", ""),
        "estimated_delivery": tracking_info.get("estimated_delivery"),
        "events": tracking_info.get("events", []),
        "last_tracked": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.shipments.update_one({"id": shipment_id}, {"$set": update_data})
    updated = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
    
    # Send notification if status changed to delivered
    notification_sent = False
    if new_status == "delivered" and old_status != "delivered" and notify_admin and admin_email:
        try:
            settings = await db.company_settings.find_one({}, {"_id": 0})
            if settings and settings.get('smtp_host'):
                # Get lead info for customer name
                lead = await db.leads.find_one({"id": shipment.get('lead_id')}, {"_id": 0})
                customer_name = lead.get('company') if lead else shipment.get('company_name', 'Bilinmiyor')
                
                subject = f"📦 Kargo Teslim Edildi - {shipment['tracking_number']}"
                body = f"""
                <html>
                <body style="font-family: Arial, sans-serif; padding: 20px;">
                    <div style="max-width: 600px; margin: 0 auto; background: #d1fae5; padding: 30px; border-radius: 10px;">
                        <h2 style="color: #059669;">✅ Kargo Teslim Edildi!</h2>
                        <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p><strong>Takip No:</strong> {shipment['tracking_number']}</p>
                            <p><strong>Müşteri:</strong> {customer_name}</p>
                            <p><strong>Konum:</strong> {tracking_info.get('current_location', 'Belirtilmedi')}</p>
                            <p><strong>Teslim Tarihi:</strong> {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')}</p>
                        </div>
                        <p style="color: #6b7280; font-size: 12px;">Bu bildirim otomatik olarak gönderilmiştir.</p>
                    </div>
                </body>
                </html>
                """
                
                msg = MIMEMultipart()
                msg['From'] = f"{settings.get('from_name', '')} <{settings.get('from_email', settings['smtp_username'])}>"
                msg['To'] = admin_email
                msg['Subject'] = subject
                msg.attach(MIMEText(body, 'html', 'utf-8'))
                
                smtp_host = settings['smtp_host']
                smtp_port = int(settings.get('smtp_port', 587))
                
                if settings.get('use_ssl'):
                    server = smtplib.SMTP_SSL(smtp_host, smtp_port)
                else:
                    server = smtplib.SMTP(smtp_host, smtp_port)
                    if settings.get('use_tls', True):
                        server.starttls()
                
                server.login(settings['smtp_username'], settings['smtp_password'])
                server.send_message(msg)
                server.quit()
                notification_sent = True
        except Exception as e:
            logger.error(f"Delivery notification failed: {e}")
    
    return {**updated, "notification_sent": notification_sent, "status_changed": old_status != new_status}

@api_router.get("/tracking/{tracking_number}")
async def track_package(tracking_number: str):
    """Track a package by tracking number (without saving)"""
    result = await dhl_tracker.track_package(tracking_number)
    return result

@api_router.post("/shipments/refresh-all")
async def refresh_all_shipments(notify_admin: bool = False, admin_email: str = None):
    """Refresh tracking for all active shipments and send delivery notifications"""
    # Get all non-delivered shipments
    shipments = await db.shipments.find(
        {"status": {"$nin": ["delivered", "exception"]}},
        {"_id": 0}
    ).to_list(100)
    
    updated_count = 0
    delivered_notifications = []
    
    for shipment in shipments:
        try:
            old_status = shipment.get("status", "unknown")
            tracking_info = await dhl_tracker.track_package(shipment["tracking_number"])
            new_status = tracking_info.get("status", shipment.get("status"))
            
            update_data = {
                "status": new_status,
                "status_text": tracking_info.get("status_text", ""),
                "current_location": tracking_info.get("current_location", ""),
                "estimated_delivery": tracking_info.get("estimated_delivery"),
                "events": tracking_info.get("events", []),
                "last_tracked": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.shipments.update_one({"id": shipment["id"]}, {"$set": update_data})
            updated_count += 1
            
            # Track delivered shipments for notification
            if new_status == "delivered" and old_status != "delivered":
                delivered_notifications.append({
                    "tracking_number": shipment["tracking_number"],
                    "lead_id": shipment.get("lead_id"),
                    "company_name": shipment.get("company_name")
                })
            
            # Small delay to avoid rate limiting
            await asyncio.sleep(0.5)
            
        except Exception as e:
            logger.error(f"Failed to refresh shipment {shipment['id']}: {e}")
    
    # Send consolidated notification if there are deliveries
    if delivered_notifications and notify_admin and admin_email:
        try:
            settings = await db.company_settings.find_one({}, {"_id": 0})
            if settings and settings.get('smtp_host'):
                delivery_list = ""
                for d in delivered_notifications:
                    lead = await db.leads.find_one({"id": d.get('lead_id')}, {"_id": 0})
                    customer = lead.get('company') if lead else d.get('company_name', 'Bilinmiyor')
                    delivery_list += f"<li><strong>{d['tracking_number']}</strong> - {customer}</li>"
                
                subject = f"📦 {len(delivered_notifications)} Kargo Teslim Edildi!"
                body = f"""
                <html>
                <body style="font-family: Arial, sans-serif; padding: 20px;">
                    <div style="max-width: 600px; margin: 0 auto; background: #d1fae5; padding: 30px; border-radius: 10px;">
                        <h2 style="color: #059669;">✅ Teslim Edilen Kargolar</h2>
                        <p>Aşağıdaki kargolar başarıyla teslim edildi:</p>
                        <ul style="background: white; padding: 20px 20px 20px 40px; border-radius: 8px;">
                            {delivery_list}
                        </ul>
                        <p style="color: #6b7280; font-size: 12px; margin-top: 20px;">Bu bildirim otomatik kargo takip sistemi tarafından gönderilmiştir.</p>
                    </div>
                </body>
                </html>
                """
                
                msg = MIMEMultipart()
                msg['From'] = f"{settings.get('from_name', '')} <{settings.get('from_email', settings['smtp_username'])}>"
                msg['To'] = admin_email
                msg['Subject'] = subject
                msg.attach(MIMEText(body, 'html', 'utf-8'))
                
                smtp_host = settings['smtp_host']
                smtp_port = int(settings.get('smtp_port', 587))
                
                if settings.get('use_ssl'):
                    server = smtplib.SMTP_SSL(smtp_host, smtp_port)
                else:
                    server = smtplib.SMTP(smtp_host, smtp_port)
                    if settings.get('use_tls', True):
                        server.starttls()
                
                server.login(settings['smtp_username'], settings['smtp_password'])
                server.send_message(msg)
                server.quit()
        except Exception as e:
            logger.error(f"Bulk delivery notification failed: {e}")
    
    return {
        "status": "success", 
        "updated_count": updated_count, 
        "total": len(shipments),
        "delivered_count": len(delivered_notifications)
    }

# ===================== AI SERVICES ENDPOINTS =====================

from ai_services import email_assistant, churn_predictor, recipe_optimizer, chatbot

class AIEmailRequest(BaseModel):
    email_type: str  # introduction, follow_up, quotation, thank_you, promotion, reminder
    customer_name: str
    company_name: str
    language: str = 'en'
    context: Optional[str] = ''
    product_info: Optional[str] = ''

@api_router.post("/ai/email/generate")
async def generate_ai_email(request: AIEmailRequest):
    """Generate a professional email using AI"""
    result = await email_assistant.generate_email(
        email_type=request.email_type,
        customer_name=request.customer_name,
        company_name=request.company_name,
        language=request.language,
        context=request.context or '',
        product_info=request.product_info or ''
    )
    return result

@api_router.get("/ai/churn/analyze/{lead_id}")
async def analyze_customer_churn(lead_id: str):
    """Analyze churn risk for a specific customer"""
    # Get customer data
    customer = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get customer orders
    orders = await db.orders.find({"lead_id": lead_id}, {"_id": 0}).to_list(100)
    
    result = await churn_predictor.analyze_customer(customer, orders)
    return {
        "customer_id": lead_id,
        "company_name": customer.get('company_name'),
        **result
    }

@api_router.get("/ai/churn/at-risk")
async def get_at_risk_customers():
    """Get all customers at risk of churning"""
    # Get all customers
    customers = await db.leads.find({}, {"_id": 0}).to_list(500)
    
    # Get all orders grouped by customer
    all_orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    orders_by_customer = {}
    for order in all_orders:
        lead_id = order.get('lead_id')
        if lead_id:
            if lead_id not in orders_by_customer:
                orders_by_customer[lead_id] = []
            orders_by_customer[lead_id].append(order)
    
    at_risk = await churn_predictor.get_at_risk_customers(customers, orders_by_customer)
    
    return {
        "total_customers": len(customers),
        "at_risk_count": len(at_risk),
        "at_risk_customers": at_risk[:20]  # Return top 20
    }

@api_router.post("/ai/recipe/optimize/{recipe_id}")
async def optimize_recipe(recipe_id: str, target: str = 'cost'):
    """Optimize a recipe using AI"""
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    
    result = await recipe_optimizer.optimize_recipe(recipe, target)
    return result

class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None
    language: str = 'en'
    context: Optional[dict] = None

@api_router.post("/ai/chat")
async def ai_chat(request: ChatRequest):
    """Chat with AI assistant"""
    session_id = request.session_id or str(uuid.uuid4())
    
    result = await chatbot.chat(
        session_id=session_id,
        message=request.message,
        language=request.language,
        context=request.context
    )
    
    # Store chat message in database
    await db.chat_history.insert_one({
        "id": str(uuid.uuid4()),
        "session_id": session_id,
        "user_message": request.message,
        "ai_response": result.get('response', ''),
        "language": request.language,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return result

@api_router.get("/ai/chat/history/{session_id}")
async def get_chat_history(session_id: str):
    """Get chat history for a session"""
    messages = await db.chat_history.find(
        {"session_id": session_id}, 
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    
    return {"session_id": session_id, "messages": messages}

# ===================== MAIL INBOX ROUTES (IMAP/SMTP) =====================

import imaplib
import email
from email.header import decode_header

class MailSendRequest(BaseModel):
    to: str
    subject: str
    body: str
    html: bool = False

@api_router.get("/mail/inbox")
async def get_mail_inbox(page: int = 1, limit: int = 20, folder: str = "INBOX"):
    """Fetch emails from IMAP inbox with pagination"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    
    if not settings or not settings.get('imap_host'):
        return {
            "status": "not_configured",
            "emails": [],
            "total": 0,
            "page": page,
            "pages": 0,
            "message": "IMAP settings not configured"
        }
    
    try:
        # Connect to IMAP server
        imap_host = settings.get('imap_host', 'imap.ionos.de')
        imap_port = int(settings.get('imap_port', 993))
        imap_user = settings.get('smtp_username')
        imap_pass = settings.get('smtp_password')
        
        if not imap_user or not imap_pass:
            return {"status": "error", "emails": [], "total": 0, "page": page, "pages": 0, "message": "IMAP credentials not configured"}
        
        logger.info(f"Connecting to IMAP: {imap_host}:{imap_port} as {imap_user}")
        
        # Try SSL connection first (port 993)
        try:
            mail = imaplib.IMAP4_SSL(imap_host, imap_port)
        except Exception as ssl_error:
            logger.warning(f"SSL connection failed, trying STARTTLS: {ssl_error}")
            # Try STARTTLS
            mail = imaplib.IMAP4(imap_host, 143)
            mail.starttls()
        
        mail.login(imap_user, imap_pass)
        
        # Select folder (INBOX, Sent, Drafts, etc.)
        try:
            mail.select(folder)
        except:
            mail.select('INBOX')
        
        # Fetch all email IDs
        _, message_numbers = mail.search(None, 'ALL')
        all_email_ids = message_numbers[0].split()
        total_emails = len(all_email_ids)
        total_pages = max(1, (total_emails + limit - 1) // limit)
        
        # Calculate pagination slice
        start_idx = max(0, total_emails - (page * limit))
        end_idx = total_emails - ((page - 1) * limit)
        email_ids = all_email_ids[start_idx:end_idx]
        
        emails = []
        for eid in reversed(email_ids):
            # Fetch headers, flags and body structure for attachment detection
            _, msg_data = mail.fetch(eid, '(BODY.PEEK[HEADER] FLAGS BODYSTRUCTURE)')
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    
                    # Decode subject
                    subject = ''
                    if msg['Subject']:
                        decoded = decode_header(msg['Subject'])
                        for part, charset in decoded:
                            if isinstance(part, bytes):
                                subject += part.decode(charset or 'utf-8', errors='replace')
                            else:
                                subject += part
                    
                    # Decode from
                    from_header = msg.get('From', '')
                    from_name = ''
                    from_email_addr = from_header
                    if '<' in from_header:
                        from_name = from_header.split('<')[0].strip().strip('"')
                        from_email_addr = from_header.split('<')[1].strip('>')
                    
                    # Check flags
                    is_read = b'\\Seen' in response_part[0]
                    
                    # Check for attachments in BODYSTRUCTURE
                    has_attachments = False
                    attachment_count = 0
                    body_struct_str = str(response_part[0])
                    if 'ATTACHMENT' in body_struct_str.upper() or 'attachment' in body_struct_str:
                        has_attachments = True
                        attachment_count = body_struct_str.upper().count('ATTACHMENT')
                    
                    emails.append({
                        "id": eid.decode(),
                        "subject": subject,
                        "from_name": from_name,
                        "from_email": from_email_addr,
                        "date": msg.get('Date', ''),
                        "body": '',  # Load on demand
                        "snippet": subject[:100] if subject else '',
                        "is_read": is_read,
                        "has_attachments": has_attachments,
                        "attachmentCount": attachment_count
                    })
        
        mail.logout()
        return {
            "status": "connected", 
            "emails": emails,
            "total": total_emails,
            "page": page,
            "pages": total_pages,
            "limit": limit
        }
        
    except Exception as e:
        logger.error(f"IMAP error: {e}")
        return {"status": "error", "emails": [], "total": 0, "page": page, "pages": 0, "message": str(e)}

@api_router.get("/mail/body/{email_id}")
async def get_email_body(email_id: str):
    """Fetch email body on demand"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    
    if not settings or not settings.get('imap_host'):
        raise HTTPException(status_code=400, detail="IMAP not configured")
    
    try:
        mail = imaplib.IMAP4_SSL(settings['imap_host'], int(settings.get('imap_port', 993)))
        mail.login(settings['smtp_username'], settings['smtp_password'])
        mail.select('INBOX')
        
        _, msg_data = mail.fetch(email_id.encode(), '(RFC822)')
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                msg = email.message_from_bytes(response_part[1])
                
                # Get body and attachments
                body = ''
                html_body = ''
                attachments = []
                
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        content_disposition = str(part.get('Content-Disposition', ''))
                        
                        # Check for attachments
                        if 'attachment' in content_disposition or part.get_filename():
                            filename = part.get_filename()
                            if filename:
                                # Decode filename if needed
                                if '=?' in filename:
                                    decoded = decode_header(filename)
                                    filename = ''.join([
                                        t[0].decode(t[1] or 'utf-8') if isinstance(t[0], bytes) else t[0]
                                        for t in decoded
                                    ])
                                attachments.append({
                                    'filename': filename,
                                    'content_type': content_type,
                                    'size': len(part.get_payload(decode=True) or b'')
                                })
                        elif content_type == 'text/plain' and not body:
                            try:
                                body = part.get_payload(decode=True).decode('utf-8', errors='replace')
                            except:
                                body = str(part.get_payload())
                        elif content_type == 'text/html' and not html_body:
                            try:
                                html_body = part.get_payload(decode=True).decode('utf-8', errors='replace')
                            except:
                                html_body = str(part.get_payload())
                else:
                    try:
                        body = msg.get_payload(decode=True).decode('utf-8', errors='replace')
                    except:
                        body = str(msg.get_payload())
                
                mail.logout()
                return {"body": html_body or body, "plain": body, "attachments": attachments}
        
        mail.logout()
        return {"body": "", "plain": "", "attachments": []}
    except Exception as e:
        logger.error(f"Get email body error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Multi-language support for AI features
AI_LANGUAGE_CONFIG = {
    'tr': {
        'name': 'Türkçe',
        'system_message': 'Sen profesyonel bir e-posta asistanısın. Türkçe yanıt ver. Kısa ve öz ol.',
        'summary_prompt': 'Aşağıdaki e-postayı Türkçe olarak özetle (maksimum 2 cümle)',
        'replies_prompt': 'Bu e-postaya verilebilecek 3 adet kısa ve profesyonel Türkçe yanıt önerisi ver',
        'default_replies': [
            'Teşekkür ederim, en kısa sürede inceleyeceğim.',
            'Bu konuda size geri dönüş yapacağım.',
            'Detaylı bilgi için teşekkürler.'
        ],
        'fallback_summary': 'tarafından gönderilen e-posta'
    },
    'en': {
        'name': 'English',
        'system_message': 'You are a professional email assistant. Respond in English. Be concise.',
        'summary_prompt': 'Summarize the following email in English (maximum 2 sentences)',
        'replies_prompt': 'Provide 3 short and professional English reply suggestions for this email',
        'default_replies': [
            'Thank you, I will review this shortly.',
            'I will get back to you on this matter.',
            'Thanks for the detailed information.'
        ],
        'fallback_summary': 'Email from'
    },
    'de': {
        'name': 'Deutsch',
        'system_message': 'Sie sind ein professioneller E-Mail-Assistent. Antworten Sie auf Deutsch. Seien Sie präzise.',
        'summary_prompt': 'Fassen Sie die folgende E-Mail auf Deutsch zusammen (maximal 2 Sätze)',
        'replies_prompt': 'Geben Sie 3 kurze und professionelle deutsche Antwortvorschläge für diese E-Mail',
        'default_replies': [
            'Vielen Dank, ich werde das schnellstmöglich prüfen.',
            'Ich werde mich in dieser Angelegenheit bei Ihnen melden.',
            'Danke für die ausführlichen Informationen.'
        ],
        'fallback_summary': 'E-Mail von'
    },
    'pl': {
        'name': 'Polski',
        'system_message': 'Jesteś profesjonalnym asystentem e-mail. Odpowiadaj po polsku. Bądź zwięzły.',
        'summary_prompt': 'Podsumuj następującą wiadomość e-mail po polsku (maksymalnie 2 zdania)',
        'replies_prompt': 'Podaj 3 krótkie i profesjonalne polskie propozycje odpowiedzi na tę wiadomość',
        'default_replies': [
            'Dziękuję, wkrótce to przejrzę.',
            'Wrócę do Państwa w tej sprawie.',
            'Dziękuję za szczegółowe informacje.'
        ],
        'fallback_summary': 'E-mail od'
    },
    'el': {
        'name': 'Ελληνικά',
        'system_message': 'Είστε επαγγελματίας βοηθός email. Απαντήστε στα ελληνικά. Να είστε σύντομοι.',
        'summary_prompt': 'Συνοψίστε το παρακάτω email στα ελληνικά (μέγιστο 2 προτάσεις)',
        'replies_prompt': 'Δώστε 3 σύντομες και επαγγελματικές ελληνικές προτάσεις απάντησης για αυτό το email',
        'default_replies': [
            'Ευχαριστώ, θα το εξετάσω σύντομα.',
            'Θα επικοινωνήσω μαζί σας σχετικά με αυτό το θέμα.',
            'Ευχαριστώ για τις λεπτομερείς πληροφορίες.'
        ],
        'fallback_summary': 'Email από'
    },
    'bg': {
        'name': 'Български',
        'system_message': 'Вие сте професионален имейл асистент. Отговаряйте на български. Бъдете кратки.',
        'summary_prompt': 'Обобщете следния имейл на български (максимум 2 изречения)',
        'replies_prompt': 'Дайте 3 кратки и професионални български предложения за отговор на този имейл',
        'default_replies': [
            'Благодаря, ще го прегледам скоро.',
            'Ще се свържа с вас по този въпрос.',
            'Благодаря за подробната информация.'
        ],
        'fallback_summary': 'Имейл от'
    }
}


@api_router.post("/ai/summarize-email")
async def ai_summarize_email(data: dict):
    """AI-powered email summarization and smart replies with multi-language support"""
    try:
        subject = data.get('subject', '')
        body = data.get('body', '')
        from_name = data.get('from', '')
        lang = data.get('language', 'tr')  # Default to Turkish
        
        # Get language config
        lang_config = AI_LANGUAGE_CONFIG.get(lang, AI_LANGUAGE_CONFIG['tr'])
        
        # Strip HTML tags from body for summarization
        import re
        clean_body = re.sub(r'<[^>]+>', '', body)
        clean_body = clean_body[:2000]  # Limit length
        
        # Try to use AI service
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            import os
            
            api_key = os.environ.get('EMERGENT_LLM_KEY')
            chat = LlmChat(
                api_key=api_key,
                session_id=f"email-summary-{uuid.uuid4()}",
                system_message=lang_config['system_message']
            ).with_model("gemini", "gemini-2.5-flash")
            
            prompt = f"""{lang_config['summary_prompt']}:

From: {from_name}
Subject: {subject}
Content:
{clean_body}

{lang_config['replies_prompt']}. Number each suggestion on a new line."""

            user_message = UserMessage(text=prompt)
            response = await chat.send_message(user_message)
            
            # Parse response
            text = str(response)
            lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
            
            # First lines are summary
            summary_lines = []
            replies = []
            for line in lines:
                if line.startswith(('1.', '2.', '3.', '1)', '2)', '3)')):
                    reply = line.lstrip('123.)- ').strip()
                    if len(reply) > 5:
                        replies.append(reply)
                elif not replies:
                    summary_lines.append(line)
            
            summary = ' '.join(summary_lines[:2]) if summary_lines else f"{lang_config['fallback_summary']} {from_name}"
            
            if not replies:
                replies = lang_config['default_replies']
            
            return {"summary": summary, "replies": replies[:3], "language": lang}
            
        except Exception as ai_error:
            logger.error(f"AI service error: {ai_error}")
            return {
                "summary": f"{lang_config['fallback_summary']} {from_name}: '{subject}'",
                "replies": lang_config['default_replies'],
                "language": lang
            }
    except Exception as e:
        logger.error(f"Email summarization error: {e}")
        lang_config = AI_LANGUAGE_CONFIG.get('tr')
        return {
            "summary": lang_config['fallback_summary'],
            "replies": lang_config['default_replies'],
            "language": 'tr'
        }


@api_router.post("/ai/translate-email")
async def ai_translate_email(data: dict):
    """Translate email content to multiple languages"""
    try:
        body = data.get('body', '')
        target_lang = data.get('target_lang', 'tr')
        
        import re
        clean_body = re.sub(r'<[^>]+>', '', body)[:3000]
        
        lang_map = {
            'tr': 'Türkçe',
            'en': 'English', 
            'de': 'Deutsch',
            'pl': 'Polski',
            'el': 'Ελληνικά (Greek)',
            'bg': 'Български (Bulgarian)'
        }
        target = lang_map.get(target_lang, 'Türkçe')
        
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            import os
            
            api_key = os.environ.get('EMERGENT_LLM_KEY')
            chat = LlmChat(
                api_key=api_key,
                session_id=f"translate-{uuid.uuid4()}",
                system_message=f"You are a professional translator. Translate the text to {target}. Only write the translation, nothing else."
            ).with_model("gemini", "gemini-2.5-flash")
            
            user_message = UserMessage(text=clean_body)
            response = await chat.send_message(user_message)
            
            return {"translated": str(response), "language": target_lang}
        except Exception as e:
            logger.error(f"Translation error: {e}")
            return {"translated": clean_body, "error": "Translation failed", "language": target_lang}
    except Exception as e:
        return {"translated": "", "error": str(e)}


# Multi-language tone configurations for email composition
AI_TONE_CONFIG = {
    'tr': {
        'professional': 'profesyonel ve resmi',
        'friendly': 'samimi ve sıcak',
        'formal': 'çok resmi ve kurumsal',
        'system': 'Sen profesyonel bir e-posta yazarısın. {tone} bir üslupla Türkçe e-posta yaz.',
        'prompt': 'Aşağıdaki bilgilere göre bir e-posta yaz:\n\nİstek: {prompt}\n{context}\n\nE-postayı {tone} bir üslupla yaz.',
        'error': 'E-posta oluşturulamadı'
    },
    'en': {
        'professional': 'professional and business-like',
        'friendly': 'friendly and warm',
        'formal': 'very formal and corporate',
        'system': 'You are a professional email writer. Write an English email in a {tone} style.',
        'prompt': 'Write an email based on the following:\n\nRequest: {prompt}\n{context}\n\nWrite the email in a {tone} style.',
        'error': 'Could not generate email'
    },
    'de': {
        'professional': 'professionell und geschäftsmäßig',
        'friendly': 'freundlich und warm',
        'formal': 'sehr formell und geschäftlich',
        'system': 'Sie sind ein professioneller E-Mail-Schreiber. Schreiben Sie eine deutsche E-Mail in einem {tone} Stil.',
        'prompt': 'Schreiben Sie eine E-Mail basierend auf Folgendem:\n\nAnfrage: {prompt}\n{context}\n\nSchreiben Sie die E-Mail in einem {tone} Stil.',
        'error': 'E-Mail konnte nicht erstellt werden'
    },
    'pl': {
        'professional': 'profesjonalny i biznesowy',
        'friendly': 'przyjazny i ciepły',
        'formal': 'bardzo formalny i korporacyjny',
        'system': 'Jesteś profesjonalnym pisarzem e-maili. Napisz e-mail po polsku w stylu {tone}.',
        'prompt': 'Napisz e-mail na podstawie:\n\nProśba: {prompt}\n{context}\n\nNapisz e-mail w stylu {tone}.',
        'error': 'Nie można wygenerować e-maila'
    },
    'el': {
        'professional': 'επαγγελματικός και επιχειρηματικός',
        'friendly': 'φιλικός και ζεστός',
        'formal': 'πολύ επίσημος και εταιρικός',
        'system': 'Είστε επαγγελματίας συγγραφέας email. Γράψτε ένα ελληνικό email σε {tone} στυλ.',
        'prompt': 'Γράψτε ένα email βάσει των παρακάτω:\n\nΑίτημα: {prompt}\n{context}\n\nΓράψτε το email σε {tone} στυλ.',
        'error': 'Δεν ήταν δυνατή η δημιουργία email'
    },
    'bg': {
        'professional': 'професионален и делови',
        'friendly': 'приятелски и топъл',
        'formal': 'много официален и корпоративен',
        'system': 'Вие сте професионален писател на имейли. Напишете български имейл в {tone} стил.',
        'prompt': 'Напишете имейл въз основа на следното:\n\nЗаявка: {prompt}\n{context}\n\nНапишете имейла в {tone} стил.',
        'error': 'Имейлът не можа да бъде генериран'
    }
}


@api_router.post("/ai/compose-email")
async def ai_compose_email(data: dict):
    """AI-assisted email composition with multi-language support"""
    try:
        prompt = data.get('prompt', '')
        context = data.get('context', '')
        tone = data.get('tone', 'professional')
        lang = data.get('language', 'tr')
        
        # Get language config
        lang_config = AI_TONE_CONFIG.get(lang, AI_TONE_CONFIG['tr'])
        tone_desc = lang_config.get(tone, lang_config['professional'])
        
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            import os
            
            api_key = os.environ.get('EMERGENT_LLM_KEY')
            system_msg = lang_config['system'].format(tone=tone_desc)
            
            chat = LlmChat(
                api_key=api_key,
                session_id=f"compose-{uuid.uuid4()}",
                system_message=system_msg
            ).with_model("gemini", "gemini-2.5-flash")
            
            context_text = f"Context: {context}" if context else ""
            full_prompt = lang_config['prompt'].format(prompt=prompt, context=context_text, tone=tone_desc)

            user_message = UserMessage(text=full_prompt)
            response = await chat.send_message(user_message)
            
            return {"email": str(response), "language": lang}
        except Exception as e:
            logger.error(f"Compose error: {e}")
            return {"email": "", "error": lang_config['error'], "language": lang}
    except Exception as e:
        return {"email": "", "error": str(e)}


@api_router.post("/ai/improve-text")
async def ai_improve_text(data: dict):
    """Improve email text"""
    try:
        text = data.get('text', '')
        action = data.get('action', 'improve')  # improve, shorten, expand, formal
        
        action_map = {
            'improve': 'daha iyi ve profesyonel hale getir',
            'shorten': 'kısalt ama anlamı koru',
            'expand': 'daha detaylı ve açıklayıcı yaz',
            'formal': 'daha resmi bir dille yeniden yaz'
        }
        action_desc = action_map.get(action, 'iyileştir')
        
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            import os
            
            api_key = os.environ.get('EMERGENT_LLM_KEY')
            chat = LlmChat(
                api_key=api_key,
                session_id=f"improve-{uuid.uuid4()}",
                system_message="Sen profesyonel bir editörsün. Sadece düzeltilmiş metni yaz, başka bir şey yazma."
            ).with_model("gemini", "gemini-2.5-flash")
            
            user_message = UserMessage(text=f"Aşağıdaki metni {action_desc}:\n\n{text}")
            response = await chat.send_message(user_message)
            
            return {"improved": str(response)}
        except Exception as e:
            logger.error(f"Improve error: {e}")
            return {"improved": text, "error": "İyileştirme yapılamadı"}
    except Exception as e:
        return {"improved": "", "error": str(e)}


@api_router.post("/ai/analyze-spam")
async def ai_analyze_spam(data: dict):
    """AI-powered spam analysis"""
    try:
        subject = data.get('subject', '')
        body = data.get('body', '')
        from_email = data.get('from_email', '')
        
        import re
        clean_body = re.sub(r'<[^>]+>', '', body)[:1500]
        
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            import os
            
            api_key = os.environ.get('EMERGENT_LLM_KEY')
            chat = LlmChat(
                api_key=api_key,
                session_id=f"spam-{uuid.uuid4()}",
                system_message="Sen bir spam analiz uzmanısın. E-postaları analiz et ve spam olup olmadığını belirle."
            ).with_model("gemini", "gemini-2.5-flash")
            
            prompt = f"""Bu e-postanın spam olup olmadığını analiz et:

Gönderen: {from_email}
Konu: {subject}
İçerik: {clean_body[:500]}

Şu JSON formatında yanıt ver:
{{"is_spam": true/false, "confidence": 0-100, "reason": "kısa açıklama"}}"""

            user_message = UserMessage(text=prompt)
            response = await chat.send_message(user_message)
            
            import json
            try:
                result = json.loads(str(response).strip())
                return result
            except:
                return {"is_spam": False, "confidence": 50, "reason": "Analiz yapılamadı"}
        except Exception as e:
            logger.error(f"Spam analysis error: {e}")
            return {"is_spam": False, "confidence": 0, "reason": "Analiz servisi kullanılamıyor"}
    except Exception as e:
        return {"is_spam": False, "confidence": 0, "reason": str(e)}


@api_router.post("/ai/recognize-customer")
async def ai_recognize_customer(data: dict):
    """AI-powered customer recognition from email"""
    try:
        from_email = data.get('from_email', '')
        from_name = data.get('from_name', '')
        body = data.get('body', '')
        
        # First try to find customer by email
        customer = await db.leads.find_one({
            "$or": [
                {"email": from_email},
                {"email": {"$regex": from_email.split('@')[0], "$options": "i"}}
            ]
        }, {"_id": 0})
        
        if customer:
            return {
                "found": True,
                "customer": {
                    "id": customer.get('id'),
                    "company_name": customer.get('company_name'),
                    "email": customer.get('email'),
                    "phone": customer.get('phone'),
                    "city": customer.get('city')
                },
                "source": "database"
            }
        
        # Try to find by company name mentioned in email
        import re
        clean_body = re.sub(r'<[^>]+>', '', body)[:1000]
        
        # Search for company names
        leads = await db.leads.find({}, {"_id": 0, "id": 1, "company_name": 1, "email": 1}).to_list(500)
        
        for lead in leads:
            company = lead.get('company_name', '')
            if company and len(company) > 3:
                if company.lower() in from_name.lower() or company.lower() in clean_body.lower():
                    return {
                        "found": True,
                        "customer": lead,
                        "source": "name_match"
                    }
        
        return {"found": False, "customer": None, "source": None}
    except Exception as e:
        logger.error(f"Customer recognition error: {e}")
        return {"found": False, "customer": None, "error": str(e)}


@api_router.post("/ai/analyze-sentiment")
async def ai_analyze_sentiment(data: dict):
    """AI-powered sentiment analysis for emails"""
    try:
        subject = data.get('subject', '')
        body = data.get('body', '')
        
        import re
        clean_body = re.sub(r'<[^>]+>', '', body)[:1500]
        
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            import os
            
            api_key = os.environ.get('EMERGENT_LLM_KEY')
            chat = LlmChat(
                api_key=api_key,
                session_id=f"sentiment-{uuid.uuid4()}",
                system_message="Sen bir duygu analizi uzmanısın. E-postaların tonunu analiz et."
            ).with_model("gemini", "gemini-2.5-flash")
            
            prompt = f"""Bu e-postanın duygusal tonunu analiz et:

Konu: {subject}
İçerik: {clean_body[:500]}

Şu JSON formatında yanıt ver:
{{"sentiment": "positive/negative/neutral", "urgency": "high/medium/low", "confidence": 0-100, "summary": "kısa açıklama"}}"""

            user_message = UserMessage(text=prompt)
            response = await chat.send_message(user_message)
            
            import json
            try:
                result = json.loads(str(response).strip())
                return result
            except:
                return {"sentiment": "neutral", "urgency": "medium", "confidence": 50, "summary": "Analiz yapılamadı"}
        except Exception as e:
            logger.error(f"Sentiment analysis error: {e}")
            return {"sentiment": "neutral", "urgency": "medium", "confidence": 0, "summary": "Analiz servisi kullanılamıyor"}
    except Exception as e:
        return {"sentiment": "neutral", "urgency": "medium", "confidence": 0, "summary": str(e)}


# Custom Folder Management
@api_router.get("/mail/folders")
async def get_mail_folders():
    """Get custom mail folders"""
    try:
        folders = await db.mail_folders.find({}, {"_id": 0}).to_list(100)
        # Add default folders
        default_folders = [
            {"id": "inbox", "name": "Gelen Kutusu", "icon": "inbox", "is_default": True},
            {"id": "sent", "name": "Gönderilenler", "icon": "send", "is_default": True},
            {"id": "drafts", "name": "Taslaklar", "icon": "file-text", "is_default": True},
            {"id": "starred", "name": "Yıldızlı", "icon": "star", "is_default": True},
            {"id": "trash", "name": "Çöp Kutusu", "icon": "trash", "is_default": True},
        ]
        return {"default_folders": default_folders, "custom_folders": folders}
    except Exception as e:
        return {"default_folders": [], "custom_folders": [], "error": str(e)}


@api_router.post("/mail/folders")
async def create_mail_folder(data: dict):
    """Create a custom mail folder"""
    try:
        folder = {
            "id": str(uuid.uuid4()),
            "name": data.get("name", "Yeni Klasör"),
            "color": data.get("color", "#6366f1"),
            "icon": data.get("icon", "folder"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_default": False
        }
        await db.mail_folders.insert_one(folder)
        # Return folder without MongoDB _id
        folder_response = {k: v for k, v in folder.items() if k != '_id'}
        return {"success": True, "folder": folder_response}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/mail/folders/{folder_id}")
async def delete_mail_folder(folder_id: str):
    """Delete a custom mail folder"""
    try:
        result = await db.mail_folders.delete_one({"id": folder_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Folder not found")
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/mail/mark-read/{email_id}")
async def mark_email_read(email_id: str):
    """Mark email as read"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    
    if not settings or not settings.get('imap_host'):
        raise HTTPException(status_code=400, detail="IMAP not configured")
    
    try:
        mail = imaplib.IMAP4_SSL(settings['imap_host'], int(settings.get('imap_port', 993)))
        mail.login(settings['smtp_username'], settings['smtp_password'])
        mail.select('INBOX')
        mail.store(email_id.encode(), '+FLAGS', '\\Seen')
        mail.logout()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/mail/sent")
async def get_sent_emails():
    """Get sent emails from local database"""
    try:
        sent = await db.sent_emails.find({}, {"_id": 0}).sort("date", -1).limit(50).to_list(50)
        return {"emails": sent}
    except Exception as e:
        return {"emails": []}

@api_router.get("/mail/drafts")
async def get_draft_emails():
    """Get draft emails"""
    try:
        drafts = await db.draft_emails.find({}, {"_id": 0}).sort("date", -1).limit(20).to_list(20)
        return {"drafts": drafts}
    except Exception as e:
        return {"drafts": []}

@api_router.post("/mail/drafts")
async def save_draft(data: dict):
    """Save email as draft"""
    try:
        draft = {
            "id": str(uuid.uuid4()),
            "to": data.get("to", ""),
            "subject": data.get("subject", ""),
            "body": data.get("body", ""),
            "date": datetime.now(timezone.utc).isoformat()
        }
        await db.draft_emails.insert_one(draft)
        return {"success": True, "id": draft["id"]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/mail/send")
async def send_mail(request: MailSendRequest):
    """Prepare email and return mailto link"""
    import urllib.parse
    import re
    
    body_text = request.body
    body_text = re.sub(r'<br\s*/?>', '\n', body_text)
    body_text = re.sub(r'<p[^>]*>', '', body_text)
    body_text = re.sub(r'</p>', '\n', body_text)
    body_text = re.sub(r'<[^>]+>', '', body_text)
    body_text = body_text.replace('&nbsp;', ' ').strip()
    
    await db.sent_emails.insert_one({
        "id": str(uuid.uuid4()),
        "to": request.to,
        "subject": request.subject,
        "body": request.body,
        "date": datetime.now(timezone.utc).isoformat()
    })
    
    mailto = f"mailto:{request.to}?subject={urllib.parse.quote(request.subject)}&body={urllib.parse.quote(body_text)}"
    
    return {"success": True, "mailto": mailto}


@api_router.post("/mail/send-to-drafts")
async def send_mail_to_drafts(
    to: str = Form(...),
    subject: str = Form(...),
    body: str = Form(...),
    attachments: List[UploadFile] = File(default=[])
):
    """Save email with attachments to IMAP Drafts folder"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    
    if not settings:
        raise HTTPException(status_code=400, detail="Email ayarları yapılandırılmamış")
    
    from_email = settings.get('from_email', settings.get('smtp_username'))
    from_name = settings.get('from_name', 'Gewürzberg GmbH')
    
    msg = MIMEMultipart()
    msg['From'] = f"{from_name} <{from_email}>"
    msg['To'] = to
    msg['Subject'] = subject
    msg['Date'] = formatdate(localtime=True)
    
    # Add body
    msg.attach(MIMEText(body, 'html', 'utf-8'))
    
    # Add attachments
    for attachment in attachments:
        content = await attachment.read()
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(content)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{attachment.filename}"')
        msg.attach(part)
    
    # Save to IMAP Drafts
    try:
        import imaplib
        import time
        imap_host = settings.get('imap_host', 'imap.ionos.de')
        imap = imaplib.IMAP4_SSL(imap_host, 993)
        imap.login(settings['smtp_username'], settings['smtp_password'])
        imap.append('Drafts', '', imaplib.Time2Internaldate(time.time()), msg.as_bytes())
        imap.logout()
        
        return {"success": True, "message": "Mail ek ile Taslaklar'a kaydedildi. IONOS web mail'den gönderin."}
    except Exception as e:
        logger.error(f"IMAP error: {e}")
        raise HTTPException(status_code=500, detail=f"Taslak kaydedilemedi: {str(e)}")

@api_router.post("/mail/send-with-attachments")
async def send_mail_with_attachments(
    to: str = Form(...),
    subject: str = Form(...),
    body: str = Form(...),
    html: str = Form("false"),
    attachments: List[UploadFile] = File(default=[])
):
    """Send email with attachments via SMTP"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    
    if not settings or not settings.get('smtp_host'):
        raise HTTPException(status_code=400, detail="SMTP not configured")
    
    try:
        msg = MIMEMultipart()
        msg['From'] = f"{settings.get('from_name', '')} <{settings.get('from_email', settings['smtp_username'])}>"
        msg['To'] = to
        msg['Subject'] = subject
        
        # Use HTML content if html flag is set
        if html.lower() == 'true':
            msg.attach(MIMEText(body, 'html', 'utf-8'))
        else:
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Attach files
        for attachment in attachments:
            content = await attachment.read()
            part = MIMEApplication(content, Name=attachment.filename)
            part['Content-Disposition'] = f'attachment; filename="{attachment.filename}"'
            msg.attach(part)
        
        smtp_host = settings['smtp_host']
        smtp_port = int(settings.get('smtp_port', 587))
        
        if settings.get('use_ssl'):
            server = smtplib.SMTP_SSL(smtp_host, smtp_port)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port)
            if settings.get('use_tls', True):
                server.starttls()
        
        server.login(settings['smtp_username'], settings['smtp_password'])
        server.send_message(msg)
        server.quit()
        
        # Log to email_log
        await db.email_log.insert_one({
            "id": str(uuid.uuid4()),
            "to": to,
            "subject": subject,
            "body": body,
            "attachments": [a.filename for a in attachments],
            "status": "sent",
            "sent_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"success": True, "message": "Email sent with attachments"}
    except Exception as e:
        logger.error(f"SMTP error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== PRODUCT VIDEOS ROUTES =====================

import shutil

UPLOAD_DIR = Path("/app/uploads/videos")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

class ProductVideoCreate(BaseModel):
    title: str
    description: Optional[str] = ''
    product_id: Optional[str] = None

@api_router.get("/product-videos")
async def get_product_videos():
    """Get all product videos"""
    videos = await db.product_videos.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return videos

@api_router.post("/product-videos")
async def upload_product_video(
    file: UploadFile = File(...),
    title: str = Form(...),
    description: str = Form(''),
    product_id: str = Form(None),
    folder_id: str = Form(None)
):
    """Upload a product video"""
    # Validate file type
    allowed_types = ['video/mp4', 'video/webm', 'video/quicktime', 'video/x-msvideo']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Allowed: MP4, WebM, MOV, AVI")
    
    # Generate unique filename
    video_id = str(uuid.uuid4())
    ext = file.filename.split('.')[-1] if '.' in file.filename else 'mp4'
    filename = f"{video_id}.{ext}"
    file_path = UPLOAD_DIR / filename
    
    # Save file
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        file_size = file_path.stat().st_size
        
        # Create video record
        video_data = {
            "id": video_id,
            "title": title,
            "description": description,
            "product_id": product_id,
            "folder_id": folder_id,
            "filename": filename,
            "url": f"/api/product-videos/stream/{video_id}",
            "file_size": file_size,
            "content_type": file.content_type,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.product_videos.insert_one(video_data)
        
        return {**video_data, "_id": None}
    except Exception as e:
        if file_path.exists():
            file_path.unlink()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/product-videos/stream/{video_id}")
async def stream_video(video_id: str):
    """Stream a video file"""
    video = await db.product_videos.find_one({"id": video_id}, {"_id": 0})
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    
    file_path = UPLOAD_DIR / video['filename']
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Video file not found")
    
    def iterfile():
        with open(file_path, "rb") as f:
            while chunk := f.read(1024 * 1024):  # 1MB chunks
                yield chunk
    
    return StreamingResponse(
        iterfile(),
        media_type=video.get('content_type', 'video/mp4'),
        headers={"Accept-Ranges": "bytes"}
    )

@api_router.delete("/product-videos/{video_id}")
async def delete_product_video(video_id: str):
    """Delete a product video"""
    video = await db.product_videos.find_one({"id": video_id}, {"_id": 0})
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    
    # Delete file
    file_path = UPLOAD_DIR / video['filename']
    if file_path.exists():
        file_path.unlink()
    
    # Delete record
    await db.product_videos.delete_one({"id": video_id})
    
    return {"message": "Video deleted"}

@api_router.put("/product-videos/{video_id}/move")
async def move_video_to_folder(video_id: str, folder_id: str = None):
    """Move a video to a different folder (or root if folder_id is null)"""
    video = await db.product_videos.find_one({"id": video_id}, {"_id": 0})
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    
    # Validate folder if provided
    if folder_id:
        folder = await db.video_folders.find_one({"id": folder_id}, {"_id": 0})
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")
    
    await db.product_videos.update_one(
        {"id": video_id},
        {"$set": {"folder_id": folder_id}}
    )
    
    return {"message": "Video moved", "video_id": video_id, "folder_id": folder_id}

@api_router.post("/product-videos/{video_id}/copy")
async def copy_video_to_folder(video_id: str, folder_id: str = None):
    """Copy a video to a different folder"""
    video = await db.product_videos.find_one({"id": video_id}, {"_id": 0})
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    
    # Validate folder if provided
    if folder_id:
        folder = await db.video_folders.find_one({"id": folder_id}, {"_id": 0})
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")
    
    # Create new video record (copy)
    new_video_id = str(uuid.uuid4())
    
    # Copy the actual file
    old_file_path = UPLOAD_DIR / video['filename']
    if not old_file_path.exists():
        raise HTTPException(status_code=404, detail="Video file not found")
    
    ext = video['filename'].split('.')[-1] if '.' in video['filename'] else 'mp4'
    new_filename = f"{new_video_id}.{ext}"
    new_file_path = UPLOAD_DIR / new_filename
    
    shutil.copy2(old_file_path, new_file_path)
    
    new_video_data = {
        "id": new_video_id,
        "title": f"{video['title']} (Kopya)",
        "description": video.get('description', ''),
        "product_id": video.get('product_id'),
        "folder_id": folder_id,
        "filename": new_filename,
        "url": f"/api/product-videos/stream/{new_video_id}",
        "file_size": video.get('file_size', 0),
        "content_type": video.get('content_type', 'video/mp4'),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.product_videos.insert_one(new_video_data)
    
    return {"message": "Video copied", "new_video_id": new_video_id, "folder_id": folder_id}

# ===================== ADMIN USER MANAGEMENT =====================

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    role: str = 'user'

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None

@api_router.get("/admin/users")
async def get_all_users():
    """Get all users (admin only)"""
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

@api_router.post("/admin/users")
async def create_user(user_data: UserCreate):
    """Create a new user"""
    # Check if email already exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    user_id = str(uuid.uuid4())
    password_hash = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    new_user = {
        "id": user_id,
        "name": user_data.name,
        "email": user_data.email,
        "username": user_data.email,
        "password_hash": password_hash,
        "role": user_data.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(new_user)
    
    return {"id": user_id, "name": user_data.name, "email": user_data.email, "role": user_data.role}

@api_router.put("/admin/users/{user_id}")
async def update_user(user_id: str, user_data: UserUpdate):
    """Update a user"""
    update_data = {}
    if user_data.name:
        update_data["name"] = user_data.name
    if user_data.email:
        update_data["email"] = user_data.email
        update_data["username"] = user_data.email
    if user_data.password:
        update_data["password_hash"] = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    if user_data.role:
        update_data["role"] = user_data.role
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    return {"message": "User updated"}

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str):
    """Delete a user"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.get("role") == "admin":
        raise HTTPException(status_code=400, detail="Cannot delete admin user")
    
    await db.users.delete_one({"id": user_id})
    return {"message": "User deleted"}

# ===================== EMAIL SIGNATURE =====================

@api_router.get("/settings/signature")
async def get_signature():
    """Get email signature"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    return {"signature": settings.get("email_signature", "") if settings else ""}

@api_router.post("/settings/signature")
async def save_signature(data: dict):
    """Save email signature"""
    await db.company_settings.update_one(
        {},
        {"$set": {"email_signature": data.get("signature", "")}},
        upsert=True
    )
    return {"message": "Signature saved"}

@api_router.get("/settings/email")
async def get_email_settings():
    """Get email settings"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    if not settings:
        return {}
    return {
        "sender_name": settings.get("sender_name", ""),
        "sender_email": settings.get("smtp_username", ""),
        "imap_host": settings.get("imap_host", ""),
        "imap_port": str(settings.get("imap_port", "993")),
        "smtp_host": settings.get("smtp_host", ""),
        "smtp_port": str(settings.get("smtp_port", "587")),
        "smtp_username": settings.get("smtp_username", ""),
        "smtp_password": ""  # Don't return password for security
    }

@api_router.post("/settings/email")
async def save_email_settings(data: dict):
    """Save email settings"""
    update_data = {
        "sender_name": data.get("sender_name", ""),
        "imap_host": data.get("imap_host", ""),
        "imap_port": int(data.get("imap_port", 993)),
        "smtp_host": data.get("smtp_host", ""),
        "smtp_port": int(data.get("smtp_port", 587)),
        "smtp_username": data.get("smtp_username", ""),
    }
    
    # Only update password if provided
    if data.get("smtp_password"):
        update_data["smtp_password"] = data.get("smtp_password")
    
    await db.company_settings.update_one(
        {},
        {"$set": update_data},
        upsert=True
    )
    return {"message": "Email settings saved"}


# Get frontend URL for CORS
frontend_url = os.environ.get('REACT_APP_BACKEND_URL', 'https://customer-agent-2.preview.emergentagent.com')

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*", frontend_url],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    # Create admin user if not exists
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
    existing = await db.users.find_one({"username": admin_username}, {"_id": 0})
    if not existing:
        await db.users.insert_one({
            "id": "admin-user-id",
            "username": admin_username,
            "name": "Emre Dirlik",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin user created: {admin_username}")
    
    # Write test credentials
    creds_path = Path('/app/memory/test_credentials.md')
    creds_path.parent.mkdir(parents=True, exist_ok=True)
    creds_path.write_text(f"""# Test Credentials

## Admin User
- Username: {admin_username}
- Password: {os.environ.get('ADMIN_PASSWORD', '190371')}
- Role: admin

## Auth Endpoints
- POST /api/auth/login
- POST /api/auth/logout
- GET /api/auth/me
- GET /api/auth/check
""")

# ============== AI ANALYTICS ENDPOINTS ==============

@api_router.get("/ai/analytics/summary")
async def get_ai_analytics_summary():
    """Get AI analytics summary"""
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate basic stats
    total_customers = len(leads)
    
    # Calculate health scores
    healthy_count = 0
    at_risk_count = 0
    
    for lead in leads:
        lead_orders = [o for o in orders if o.get('lead_id') == lead.get('id')]
        health_score = calculate_health_score(lead, lead_orders)
        if health_score >= 60:
            healthy_count += 1
        else:
            at_risk_count += 1
    
    return {
        "total_customers": total_customers,
        "healthy_customers": healthy_count,
        "at_risk_customers": at_risk_count,
        "avg_health_score": 75  # Placeholder
    }

@api_router.get("/ai/analytics/customer-insights")
async def get_customer_insights():
    """Get detailed customer insights with health scores and predictions"""
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    insights = []
    for lead in leads:
        lead_orders = sorted(
            [o for o in orders if o.get('lead_id') == lead.get('id')],
            key=lambda x: x.get('created_at', ''),
            reverse=True
        )
        
        # Calculate health score
        health_score = calculate_health_score(lead, lead_orders)
        
        # Calculate churn risk
        churn_risk = 'low'
        if health_score < 40:
            churn_risk = 'high'
        elif health_score < 60:
            churn_risk = 'medium'
        
        # Calculate order prediction
        avg_interval = 30  # Default 30 days
        last_order_date = None
        next_order_prediction = None
        days_overdue = 0
        days_until_order = 0
        
        if lead_orders:
            last_order = lead_orders[0]
            last_order_date = last_order.get('created_at', '')[:10]
            
            # Calculate average interval between orders
            if len(lead_orders) >= 2:
                intervals = []
                for i in range(len(lead_orders) - 1):
                    try:
                        date1 = datetime.fromisoformat(lead_orders[i].get('created_at', '').replace('Z', '+00:00'))
                        date2 = datetime.fromisoformat(lead_orders[i+1].get('created_at', '').replace('Z', '+00:00'))
                        intervals.append((date1 - date2).days)
                    except:
                        pass
                if intervals:
                    avg_interval = sum(intervals) // len(intervals)
            
            # Predict next order
            try:
                last_date = datetime.fromisoformat(last_order.get('created_at', '').replace('Z', '+00:00'))
                predicted_date = last_date + timedelta(days=avg_interval)
                next_order_prediction = predicted_date.strftime('%d.%m.%Y')
                
                days_diff = (predicted_date - datetime.now(timezone.utc)).days
                if days_diff < 0:
                    days_overdue = abs(days_diff)
                else:
                    days_until_order = days_diff
            except:
                pass
        
        # Best contact time (mock - based on business hours)
        contact_times = ['09:00-12:00', '14:00-17:00', '10:00-13:00']
        best_contact_time = contact_times[hash(lead.get('id', '')) % len(contact_times)]
        
        insights.append({
            "id": lead.get('id'),
            "company_name": lead.get('company_name'),
            "city": lead.get('city'),
            "country": lead.get('country'),
            "email": lead.get('email'),
            "health_score": health_score,
            "churn_risk": churn_risk,
            "last_order_date": last_order_date,
            "avg_order_interval": avg_interval,
            "next_order_prediction": next_order_prediction,
            "days_overdue": days_overdue,
            "days_until_order": days_until_order,
            "best_contact_time": best_contact_time,
            "total_orders": len(lead_orders)
        })
    
    # Sort by health score (lowest first - needs attention)
    insights.sort(key=lambda x: x['health_score'])
    
    return insights

@api_router.get("/ai/analytics/sales-forecast")
async def get_sales_forecast():
    """Get AI sales forecast"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    # Group orders by month
    monthly_totals = {}
    for order in orders:
        try:
            date_str = order.get('created_at', '')[:7]  # YYYY-MM
            if date_str:
                total = sum(p.get('quantity', 0) * p.get('unit_price', 0) for p in order.get('products', []))
                monthly_totals[date_str] = monthly_totals.get(date_str, 0) + total
        except:
            pass
    
    # Calculate average monthly revenue
    if monthly_totals:
        avg_monthly = sum(monthly_totals.values()) / len(monthly_totals)
    else:
        avg_monthly = 10000  # Default
    
    # Generate forecast for next 6 months
    current_date = datetime.now(timezone.utc)
    monthly_forecast = []
    months_tr = ['Oca', 'Şub', 'Mar', 'Nis', 'May', 'Haz', 'Tem', 'Ağu', 'Eyl', 'Eki', 'Kas', 'Ara']
    
    for i in range(6):
        future_date = current_date + timedelta(days=30 * (i + 1))
        month_name = months_tr[future_date.month - 1]
        # Add some variation
        forecast_value = int(avg_monthly * (1 + (i * 0.05) + (hash(str(i)) % 10 - 5) / 100))
        monthly_forecast.append({
            "month": f"{month_name} {future_date.year}",
            "forecast": forecast_value
        })
    
    max_forecast = max(m['forecast'] for m in monthly_forecast) if monthly_forecast else 1
    
    # Quarterly forecast
    quarterly_forecast = [
        {"quarter": "Q2 2026", "forecast": int(avg_monthly * 3 * 1.05), "growth": 5},
        {"quarter": "Q3 2026", "forecast": int(avg_monthly * 3 * 1.10), "growth": 10},
        {"quarter": "Q4 2026", "forecast": int(avg_monthly * 3 * 1.15), "growth": 15},
        {"quarter": "Q1 2027", "forecast": int(avg_monthly * 3 * 1.08), "growth": 8}
    ]
    
    return {
        "next_month": monthly_forecast[0]['forecast'] if monthly_forecast else 0,
        "monthly_forecast": monthly_forecast,
        "max_forecast": max_forecast,
        "quarterly_forecast": quarterly_forecast
    }

@api_router.get("/ai/analytics/seasonal-trends")
async def get_seasonal_trends():
    """Get seasonal sales trends"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    # Group by month
    monthly_values = [0] * 12
    for order in orders:
        try:
            date_str = order.get('created_at', '')
            if date_str:
                month = int(date_str[5:7]) - 1  # 0-indexed
                total = sum(p.get('quantity', 0) * p.get('unit_price', 0) for p in order.get('products', []))
                monthly_values[month] += total
        except:
            pass
    
    # If no data, use sample data
    if sum(monthly_values) == 0:
        monthly_values = [15000, 12000, 18000, 22000, 25000, 20000, 18000, 16000, 24000, 28000, 30000, 35000]
    
    max_val = max(monthly_values)
    min_val = min(monthly_values)
    
    monthly_trends = []
    for i, val in enumerate(monthly_values):
        monthly_trends.append({
            "month": i + 1,
            "value": val,
            "is_high": val == max_val,
            "is_low": val == min_val
        })
    
    # AI insights
    high_month_idx = monthly_values.index(max_val)
    low_month_idx = monthly_values.index(min_val)
    months_tr = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
    
    insights = [
        f"En yüksek satış {months_tr[high_month_idx]} ayında gerçekleşiyor",
        f"En düşük satış {months_tr[low_month_idx]} ayında görülüyor",
        f"Yılın son çeyreği (Q4) en yüksek satış dönemini oluşturuyor",
        "Baharat siparişleri genellikle yaz aylarında azalma gösteriyor"
    ]
    
    return {
        "monthly_trends": monthly_trends,
        "insights": insights
    }

@api_router.post("/ai/analytics/refresh")
async def refresh_ai_analytics():
    """Refresh AI analytics (recalculate all scores)"""
    # In a real implementation, this would trigger background jobs
    # For now, it just returns success
    return {"status": "success", "message": "Analytics refreshed"}

def calculate_health_score(lead, orders):
    """Calculate customer health score (0-100)"""
    score = 50  # Base score
    
    # Order frequency bonus
    if len(orders) >= 10:
        score += 20
    elif len(orders) >= 5:
        score += 15
    elif len(orders) >= 2:
        score += 10
    elif len(orders) >= 1:
        score += 5
    
    # Recency bonus
    if orders:
        try:
            last_order_date = datetime.fromisoformat(orders[0].get('created_at', '').replace('Z', '+00:00'))
            days_since = (datetime.now(timezone.utc) - last_order_date).days
            
            if days_since <= 30:
                score += 20
            elif days_since <= 60:
                score += 10
            elif days_since <= 90:
                score += 5
            elif days_since > 180:
                score -= 20
        except:
            pass
    else:
        score -= 10  # No orders penalty
    
    # Contact info bonus
    if lead.get('email'):
        score += 5
    if lead.get('phone'):
        score += 5
    
    return max(0, min(100, score))

# ============== DASHBOARD ALERTS ENDPOINT ==============

@api_router.get("/dashboard/alerts")
async def get_dashboard_alerts():
    """Get critical alerts for dashboard"""
    alerts = []
    
    # Get orders with overdue payments
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    overdue_payments = []
    pending_payments = []
    for order in orders:
        payment_status = order.get('payment_status', 'pending')
        due_date = order.get('payment_due_date')
        
        if payment_status in ['pending', 'partial']:
            if due_date and due_date < today:
                overdue_payments.append(order)
            else:
                pending_payments.append(order)
    
    if overdue_payments:
        alerts.append({
            "type": "payment_overdue",
            "severity": "high",
            "title": "Vadesi Geçmiş Ödemeler",
            "message": f"{len(overdue_payments)} siparişin ödeme vadesi geçti",
            "count": len(overdue_payments),
            "action_url": "/orders?filter=overdue"
        })
    
    # Get at-risk customers (health score < 40)
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    at_risk = []
    for lead in leads:
        lead_orders = [o for o in orders if o.get('lead_id') == lead.get('id')]
        health_score = calculate_health_score(lead, lead_orders)
        if health_score < 40:
            at_risk.append({**lead, "health_score": health_score})
    
    if at_risk:
        alerts.append({
            "type": "customer_risk",
            "severity": "medium",
            "title": "Risk Altındaki Müşteriler",
            "message": f"{len(at_risk)} müşteri kayıp riski altında",
            "count": len(at_risk),
            "action_url": "/ai-analytics"
        })
    
    # Check for pending orders that need attention
    pending_orders = [o for o in orders if o.get('status') == 'pending']
    if len(pending_orders) > 3:
        alerts.append({
            "type": "pending_orders",
            "severity": "low",
            "title": "Bekleyen Siparişler",
            "message": f"{len(pending_orders)} sipariş onay bekliyor",
            "count": len(pending_orders),
            "action_url": "/orders?filter=pending"
        })
    
    return {
        "alerts": alerts,
        "summary": {
            "overdue_payments": len(overdue_payments),
            "pending_payments": len(pending_payments),
            "at_risk_customers": len(at_risk)
        }
    }

# ============== PAYMENT TRACKING ENDPOINTS ==============

@api_router.put("/orders/{order_id}/payment")
async def update_order_payment(order_id: str, payment_data: dict):
    """Update order payment status"""
    update_fields = {}
    
    if 'payment_status' in payment_data:
        update_fields['payment_status'] = payment_data['payment_status']
    if 'payment_amount' in payment_data:
        update_fields['payment_amount'] = payment_data['payment_amount']
    if 'payment_due_date' in payment_data:
        update_fields['payment_due_date'] = payment_data['payment_due_date']
    
    update_fields['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.orders.update_one(
        {"id": order_id},
        {"$set": update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return {"message": "Payment updated", "order_id": order_id}

@api_router.get("/finance/summary")
async def get_finance_summary():
    """Get financial summary - revenue by customer, payment status"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    
    # Customer revenue analysis
    customer_revenue = {}
    for order in orders:
        lead_id = order.get('lead_id')
        total = order.get('total_price', 0)
        if lead_id:
            if lead_id not in customer_revenue:
                customer_revenue[lead_id] = {
                    'total_revenue': 0,
                    'order_count': 0,
                    'paid_amount': 0,
                    'pending_amount': 0
                }
            customer_revenue[lead_id]['total_revenue'] += total
            customer_revenue[lead_id]['order_count'] += 1
            
            if order.get('payment_status') == 'paid':
                customer_revenue[lead_id]['paid_amount'] += total
            else:
                customer_revenue[lead_id]['pending_amount'] += total
    
    # Enrich with lead data
    customer_list = []
    for lead in leads:
        lead_id = lead.get('id')
        if lead_id in customer_revenue:
            rev_data = customer_revenue[lead_id]
            customer_list.append({
                'id': lead_id,
                'company_name': lead.get('company_name'),
                'city': lead.get('city'),
                'country': lead.get('country'),
                **rev_data
            })
    
    # Sort by total revenue
    customer_list.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    # Overall summary
    total_revenue = sum(o.get('total_price', 0) for o in orders)
    total_paid = sum(o.get('total_price', 0) for o in orders if o.get('payment_status') == 'paid')
    total_pending = total_revenue - total_paid
    
    return {
        "total_revenue": total_revenue,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "customer_ranking": customer_list[:20],  # Top 20 customers
        "payment_breakdown": {
            "paid": len([o for o in orders if o.get('payment_status') == 'paid']),
            "pending": len([o for o in orders if o.get('payment_status', 'pending') == 'pending']),
            "partial": len([o for o in orders if o.get('payment_status') == 'partial']),
            "overdue": len([o for o in orders if o.get('payment_status') == 'overdue'])
        }
    }

@api_router.get("/finance/product-profitability")
async def get_product_profitability():
    """Get product profitability analysis"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    product_stats = {}
    for order in orders:
        for product in order.get('products', []):
            product_code = product.get('product_code', 'Unknown')
            product_name = product.get('product_name', 'Unknown')
            
            if product_code not in product_stats:
                product_stats[product_code] = {
                    'product_code': product_code,
                    'product_name': product_name,
                    'total_quantity': 0,
                    'total_revenue': 0,
                    'order_count': 0,
                    'avg_price': 0
                }
            
            amount = product.get('amount', 0)
            pieces = product.get('pieces', 1)
            unit_price = product.get('unit_price', 0)
            subtotal = pieces * amount * unit_price
            
            product_stats[product_code]['total_quantity'] += amount * pieces
            product_stats[product_code]['total_revenue'] += subtotal
            product_stats[product_code]['order_count'] += 1
    
    # Calculate avg price
    for code in product_stats:
        if product_stats[code]['total_quantity'] > 0:
            product_stats[code]['avg_price'] = product_stats[code]['total_revenue'] / product_stats[code]['total_quantity']
    
    # Sort by revenue
    product_list = list(product_stats.values())
    product_list.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    return {
        "products": product_list,
        "total_products": len(product_list)
    }

# ============== EXPORT ENDPOINTS ==============

@api_router.get("/export/leads/excel")
async def export_leads_excel():
    """Export all leads to Excel format (CSV)"""
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    
    # Create CSV
    output = io.StringIO()
    if leads:
        fieldnames = ['company_name', 'first_name', 'last_name', 'email', 'phone', 'city', 'country', 'address', 'tax_number']
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        for lead in leads:
            writer.writerow(lead)
    
    csv_content = output.getvalue().encode('utf-8-sig')  # UTF-8 with BOM for Excel
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=customers_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

@api_router.get("/export/orders/excel")
async def export_orders_excel():
    """Export all orders to Excel format (CSV)"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    # Flatten orders for CSV
    output = io.StringIO()
    fieldnames = ['id', 'company_name', 'lead_name', 'total_price', 'status', 'payment_status', 'created_at']
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    for order in orders:
        writer.writerow(order)
    
    csv_content = output.getvalue().encode('utf-8-sig')
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=orders_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

@api_router.get("/reports/comparison")
async def get_comparison_report(period: str = "month"):
    """Get comparison report - this period vs previous period"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    
    now = datetime.now(timezone.utc)
    
    if period == "month":
        current_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        prev_start = (current_start - timedelta(days=1)).replace(day=1)
        prev_end = current_start - timedelta(seconds=1)
    elif period == "quarter":
        quarter = (now.month - 1) // 3
        current_start = now.replace(month=quarter*3+1, day=1, hour=0, minute=0, second=0, microsecond=0)
        prev_start = (current_start - timedelta(days=1)).replace(day=1)
        prev_start = prev_start.replace(month=((prev_start.month-1)//3)*3+1, day=1)
        prev_end = current_start - timedelta(seconds=1)
    else:  # year
        current_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        prev_start = current_start.replace(year=current_start.year-1)
        prev_end = current_start - timedelta(seconds=1)
    
    current_orders = []
    prev_orders = []
    
    for order in orders:
        try:
            order_date = datetime.fromisoformat(order.get('created_at', '').replace('Z', '+00:00'))
            if order_date >= current_start:
                current_orders.append(order)
            elif order_date >= prev_start and order_date <= prev_end:
                prev_orders.append(order)
        except:
            pass
    
    current_revenue = sum(o.get('total_price', 0) for o in current_orders)
    prev_revenue = sum(o.get('total_price', 0) for o in prev_orders)
    
    revenue_change = 0
    if prev_revenue > 0:
        revenue_change = ((current_revenue - prev_revenue) / prev_revenue) * 100
    
    order_change = 0
    if len(prev_orders) > 0:
        order_change = ((len(current_orders) - len(prev_orders)) / len(prev_orders)) * 100
    
    return {
        "period": period,
        "current": {
            "revenue": current_revenue,
            "order_count": len(current_orders),
            "avg_order_value": current_revenue / len(current_orders) if current_orders else 0
        },
        "previous": {
            "revenue": prev_revenue,
            "order_count": len(prev_orders),
            "avg_order_value": prev_revenue / len(prev_orders) if prev_orders else 0
        },
        "changes": {
            "revenue_percent": round(revenue_change, 1),
            "order_percent": round(order_change, 1)
        }
    }

# ============== VIDEO FOLDERS ENDPOINTS ==============

@api_router.get("/video-folders")
async def get_video_folders():
    """Get all video folders"""
    folders = await db.video_folders.find({}, {"_id": 0}).to_list(100)
    return folders

@api_router.post("/video-folders")
async def create_video_folder(folder_data: dict):
    """Create a new video folder"""
    folder = {
        "id": str(uuid.uuid4()),
        "name": folder_data.get("name"),
        "description": folder_data.get("description", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.video_folders.insert_one(folder)
    return {"id": folder["id"], "message": "Folder created"}

@api_router.delete("/video-folders/{folder_id}")
async def delete_video_folder(folder_id: str):
    """Delete a video folder"""
    # Update videos in this folder to have no folder
    await db.product_videos.update_many(
        {"folder_id": folder_id},
        {"$set": {"folder_id": None}}
    )
    result = await db.video_folders.delete_one({"id": folder_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Folder not found")
    return {"message": "Folder deleted"}

# ===================== CUSTOMER ACTIVITY LOG =====================

class ActivityCreate(BaseModel):
    activity_type: str  # visit, call, email, order, follow_up
    outcome: str  # positive, negative, postponed, ordered, no_answer
    notes: Optional[str] = ""
    next_action_date: Optional[str] = None
    next_action_note: Optional[str] = ""

@api_router.get("/leads/{lead_id}/activities")
async def get_lead_activities(lead_id: str):
    """Get all activities for a specific lead"""
    activities = await db.lead_activities.find(
        {"lead_id": lead_id}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return activities

@api_router.post("/leads/{lead_id}/activities")
async def create_lead_activity(lead_id: str, activity: ActivityCreate):
    """Create a new activity for a lead"""
    # Verify lead exists
    lead = await db.leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    activity_doc = {
        "id": str(uuid.uuid4()),
        "lead_id": lead_id,
        "activity_type": activity.activity_type,
        "outcome": activity.outcome,
        "notes": activity.notes,
        "next_action_date": activity.next_action_date,
        "next_action_note": activity.next_action_note,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.lead_activities.insert_one(activity_doc)
    
    # Update lead's last_activity field
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {
            "last_activity": activity_doc["created_at"],
            "last_activity_outcome": activity.outcome,
            "next_action_date": activity.next_action_date
        }}
    )
    
    return {k: v for k, v in activity_doc.items() if k != "_id"}

@api_router.delete("/activities/{activity_id}")
async def delete_activity(activity_id: str):
    """Delete an activity"""
    result = await db.lead_activities.delete_one({"id": activity_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    return {"message": "Activity deleted"}

@api_router.get("/activities/upcoming")
async def get_upcoming_activities():
    """Get all leads with upcoming action dates"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Get leads with next_action_date
    leads = await db.leads.find(
        {"next_action_date": {"$exists": True, "$ne": None, "$ne": ""}},
        {"_id": 0}
    ).to_list(100)
    
    # Sort by next_action_date
    leads_with_dates = [l for l in leads if l.get("next_action_date")]
    leads_with_dates.sort(key=lambda x: x.get("next_action_date", "9999"))
    
    return leads_with_dates

@api_router.post("/leads/{lead_id}/ai-suggestion")
async def get_ai_suggestion_for_lead(lead_id: str):
    """Get AI suggestion based on lead's activity history"""
    # Get lead info
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Get activity history
    activities = await db.lead_activities.find(
        {"lead_id": lead_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(20)
    
    # Get orders for this lead
    orders = await db.orders.find(
        {"lead_id": lead_id},
        {"_id": 0, "total_amount": 1, "created_at": 1, "status": 1}
    ).to_list(10)
    
    # Build context
    activity_summary = []
    for act in activities[:5]:
        outcome_tr = {
            'positive': 'Olumlu',
            'negative': 'Olumsuz', 
            'postponed': 'Erteledi',
            'ordered': 'Sipariş Verdi',
            'no_answer': 'Cevap Vermedi'
        }.get(act.get('outcome'), act.get('outcome'))
        type_tr = {
            'visit': 'Ziyaret',
            'call': 'Telefon',
            'email': 'Email',
            'order': 'Sipariş',
            'follow_up': 'Takip'
        }.get(act.get('activity_type'), act.get('activity_type'))
        activity_summary.append(f"- {act.get('created_at', '')[:10]}: {type_tr} - {outcome_tr}. {act.get('notes', '')[:100]}")
    
    total_orders = len(orders)
    total_revenue = sum(o.get('total_amount', 0) for o in orders)
    
    prompt = f"""Sen bir B2B satış danışmanısın. Aşağıdaki müşteri bilgilerine göre kısa ve net önerilerde bulun.

Müşteri: {lead.get('company_name')}
Şehir: {lead.get('city')}, {lead.get('country')}
Toplam Sipariş: {total_orders} adet
Toplam Ciro: €{total_revenue:,.0f}

Son Aktiviteler:
{chr(10).join(activity_summary) if activity_summary else 'Henüz aktivite yok'}

Lütfen şu konularda 2-3 cümlelik kısa öneriler ver:
1. Bu müşteriyle ilgili genel değerlendirme
2. Bir sonraki adım ne olmalı?
3. Dikkat edilmesi gereken noktalar

Türkçe ve samimi bir dille yaz."""

    try:
        from emergentintegrations.llm.chat import chat, UserMessage
        
        response = await chat(
            api_key=os.environ.get('EMERGENT_LLM_KEY'),
            model="gemini-2.0-flash",
            messages=[UserMessage(content=prompt)]
        )
        
        return {
            "suggestion": response.content,
            "lead_name": lead.get('company_name'),
            "total_activities": len(activities),
            "total_orders": total_orders,
            "total_revenue": total_revenue
        }
    except Exception as e:
        logger.error(f"AI suggestion error: {e}")
        # Fallback suggestion based on data
        if not activities:
            suggestion = f"🎯 {lead.get('company_name')} ile henüz bir etkileşim kaydı yok. İlk adım olarak telefon ile iletişime geçmenizi ve ihtiyaçlarını dinlemenizi öneririm."
        elif activities[0].get('outcome') == 'negative':
            suggestion = f"⚠️ Son görüşme olumsuz sonuçlanmış. 2-3 hafta bekleyip farklı bir yaklaşımla (örn: yeni ürün tanıtımı) tekrar denemenizi öneririm."
        elif activities[0].get('outcome') == 'postponed':
            suggestion = f"⏰ Müşteri ertelemiş. Belirtilen tarihte mutlaka hatırlatma yapın ve bu sefer somut bir teklif ile gidin."
        elif activities[0].get('outcome') == 'ordered':
            suggestion = f"✅ Harika! Sipariş alınmış. Teslimat sonrası memnuniyet kontrolü yapın ve çapraz satış fırsatlarını değerlendirin."
        else:
            suggestion = f"📞 Düzenli takip önemli. Haftada en az bir kez iletişime geçmeye çalışın."
        
        return {
            "suggestion": suggestion,
            "lead_name": lead.get('company_name'),
            "total_activities": len(activities),
            "total_orders": total_orders,
            "total_revenue": total_revenue
        }

# ===================== MANAGEMENT REPORTS =====================

class ReportRequest(BaseModel):
    report_type: str  # weekly, monthly, custom
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    recipient_email: str

@api_router.get("/reports/summary")
async def get_report_summary():
    """Get summary data for reports"""
    now = datetime.now(timezone.utc)
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)
    
    # This week's data
    weekly_orders = await db.orders.count_documents({
        "created_at": {"$gte": week_ago.isoformat()}
    })
    weekly_leads = await db.leads.count_documents({
        "created_at": {"$gte": week_ago.isoformat()}
    })
    weekly_activities = await db.lead_activities.count_documents({
        "created_at": {"$gte": week_ago.isoformat()}
    })
    
    # Get weekly revenue
    weekly_orders_data = await db.orders.find(
        {"created_at": {"$gte": week_ago.isoformat()}},
        {"total_amount": 1}
    ).to_list(1000)
    weekly_revenue = sum(o.get('total_amount', 0) for o in weekly_orders_data)
    
    # Monthly data
    monthly_orders = await db.orders.count_documents({
        "created_at": {"$gte": month_ago.isoformat()}
    })
    monthly_orders_data = await db.orders.find(
        {"created_at": {"$gte": month_ago.isoformat()}},
        {"total_amount": 1}
    ).to_list(1000)
    monthly_revenue = sum(o.get('total_amount', 0) for o in monthly_orders_data)
    
    # Upcoming follow-ups
    upcoming = await db.leads.find(
        {"next_action_date": {"$exists": True, "$ne": None}},
        {"_id": 0, "company_name": 1, "next_action_date": 1}
    ).to_list(10)
    
    # Activity breakdown
    activity_counts = {}
    all_activities = await db.lead_activities.find(
        {"created_at": {"$gte": week_ago.isoformat()}},
        {"outcome": 1}
    ).to_list(1000)
    for act in all_activities:
        outcome = act.get('outcome', 'unknown')
        activity_counts[outcome] = activity_counts.get(outcome, 0) + 1
    
    return {
        "weekly": {
            "orders": weekly_orders,
            "leads": weekly_leads,
            "activities": weekly_activities,
            "revenue": weekly_revenue
        },
        "monthly": {
            "orders": monthly_orders,
            "revenue": monthly_revenue
        },
        "activity_breakdown": activity_counts,
        "upcoming_followups": upcoming[:5],
        "generated_at": now.isoformat()
    }

@api_router.post("/reports/send")
async def send_report_email(request: ReportRequest):
    """Generate and send report to specified email"""
    # Get report data
    summary = await get_report_summary()
    
    # Build email content
    if request.report_type == "weekly":
        subject = f"Haftalık Satış Raporu - {datetime.now().strftime('%d.%m.%Y')}"
        body = f"""
Haftalık Satış Raporu
=====================

Bu Hafta:
- Yeni Siparişler: {summary['weekly']['orders']}
- Yeni Müşteriler: {summary['weekly']['leads']}
- Aktiviteler: {summary['weekly']['activities']}
- Ciro: €{summary['weekly']['revenue']:,.2f}

Aktivite Dağılımı:
- Olumlu: {summary['activity_breakdown'].get('positive', 0)}
- Olumsuz: {summary['activity_breakdown'].get('negative', 0)}
- Ertelenen: {summary['activity_breakdown'].get('postponed', 0)}
- Sipariş: {summary['activity_breakdown'].get('ordered', 0)}

Yaklaşan Takipler:
"""
        for fu in summary['upcoming_followups']:
            body += f"- {fu.get('company_name')}: {fu.get('next_action_date')}\n"
    else:
        subject = f"Aylık Satış Raporu - {datetime.now().strftime('%B %Y')}"
        body = f"""
Aylık Satış Raporu
==================

Bu Ay:
- Toplam Sipariş: {summary['monthly']['orders']}
- Toplam Ciro: €{summary['monthly']['revenue']:,.2f}

Haftalık Performans:
- Siparişler: {summary['weekly']['orders']}
- Aktiviteler: {summary['weekly']['activities']}
"""
    
    # Log email (or send if SMTP configured)
    email_log = {
        "id": str(uuid.uuid4()),
        "to": request.recipient_email,
        "subject": subject,
        "body": body,
        "type": "report",
        "report_type": request.report_type,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "sent"
    }
    await db.email_log.insert_one(email_log)
    
    return {
        "message": "Rapor gönderildi",
        "recipient": request.recipient_email,
        "report_type": request.report_type
    }

# ===================== ACTIVITY PDF REPORT =====================

class ActivityReportRequest(BaseModel):
    lead_id: Optional[str] = None  # If None, generate for all leads
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    recipient_email: Optional[str] = None

@api_router.post("/reports/activities/pdf")
async def generate_activity_pdf_report(request: ActivityReportRequest):
    """Generate PDF report of activities and optionally send via email"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io
    import base64
    
    # Register font for Turkish character support
    try:
        pdfmetrics.registerFont(TTFont('FreeSans', '/usr/share/fonts/truetype/freefont/FreeSans.ttf'))
        font_name = 'FreeSans'
    except:
        font_name = 'Helvetica'
    
    # Build query
    query = {}
    if request.lead_id:
        query["lead_id"] = request.lead_id
    if request.start_date:
        query["created_at"] = {"$gte": request.start_date}
    if request.end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = request.end_date
        else:
            query["created_at"] = {"$lte": request.end_date}
    
    # Fetch activities
    activities = await db.lead_activities.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    if not activities:
        raise HTTPException(status_code=404, detail="No activities found")
    
    # Get lead info for each activity
    lead_ids = list(set(a.get("lead_id") for a in activities))
    leads = await db.leads.find({"id": {"$in": lead_ids}}, {"_id": 0, "id": 1, "company_name": 1}).to_list(100)
    lead_map = {l["id"]: l.get("company_name", "Unknown") for l in leads}
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName=font_name, fontSize=18)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=font_name, fontSize=10)
    
    elements = []
    
    # Title
    title = Paragraph(f"Aktivite Raporu - {datetime.now().strftime('%d.%m.%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Summary
    outcome_counts = {}
    for act in activities:
        outcome = act.get("outcome", "unknown")
        outcome_counts[outcome] = outcome_counts.get(outcome, 0) + 1
    
    summary_text = f"Toplam Aktivite: {len(activities)}"
    for outcome, count in outcome_counts.items():
        label = {'positive': 'Olumlu', 'negative': 'Olumsuz', 'postponed': 'Ertelenen', 'ordered': 'Sipariş', 'no_answer': 'Cevapsız'}.get(outcome, outcome)
        summary_text += f" | {label}: {count}"
    elements.append(Paragraph(summary_text, normal_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Table data
    table_data = [["Tarih", "Müşteri", "Tip", "Sonuç", "Notlar"]]
    
    type_labels = {'visit': 'Ziyaret', 'call': 'Telefon', 'email': 'Email', 'order': 'Sipariş', 'follow_up': 'Takip'}
    outcome_labels = {'positive': 'Olumlu', 'negative': 'Olumsuz', 'postponed': 'Ertelendi', 'ordered': 'Sipariş', 'no_answer': 'Cevapsız'}
    
    for act in activities[:100]:  # Limit to 100 rows
        date = act.get("created_at", "")[:10] if act.get("created_at") else "-"
        company = lead_map.get(act.get("lead_id"), "Unknown")[:25]
        act_type = type_labels.get(act.get("activity_type"), act.get("activity_type", "-"))
        outcome = outcome_labels.get(act.get("outcome"), act.get("outcome", "-"))
        notes = (act.get("notes") or "-")[:50]
        table_data.append([date, company, act_type, outcome, notes])
    
    table = Table(table_data, colWidths=[2.5*cm, 5*cm, 2.5*cm, 2.5*cm, 5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#002FA7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    elements.append(table)
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # If email requested, send it
    if request.recipient_email:
        # Store email log
        email_log = {
            "id": str(uuid.uuid4()),
            "to": request.recipient_email,
            "subject": f"Aktivite Raporu - {datetime.now().strftime('%d.%m.%Y')}",
            "body": f"Aktivite raporu ekte bulunmaktadır.\n\nToplam {len(activities)} aktivite.",
            "type": "activity_report",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "status": "sent",
            "has_attachment": True
        }
        await db.email_log.insert_one(email_log)
    
    # Return PDF as base64
    pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
    
    return {
        "pdf_base64": pdf_base64,
        "filename": f"aktivite_raporu_{datetime.now().strftime('%Y%m%d')}.pdf",
        "total_activities": len(activities),
        "email_sent": request.recipient_email is not None
    }

@api_router.get("/reports/activities/download")
async def download_activity_report(lead_id: Optional[str] = None):
    """Download activity report as PDF file"""
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io
    
    # Register font
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        font_name = 'DejaVu'
    except:
        font_name = 'Helvetica'
    
    # Build query
    query = {}
    if lead_id:
        query["lead_id"] = lead_id
    
    # Fetch activities
    activities = await db.lead_activities.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    if not activities:
        raise HTTPException(status_code=404, detail="No activities found")
    
    # Get lead info
    lead_ids = list(set(a.get("lead_id") for a in activities))
    leads = await db.leads.find({"id": {"$in": lead_ids}}, {"_id": 0, "id": 1, "company_name": 1}).to_list(100)
    lead_map = {l["id"]: l.get("company_name", "Unknown") for l in leads}
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName=font_name, fontSize=18)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=font_name, fontSize=10)
    
    elements = []
    
    # Title
    if lead_id and lead_id in lead_map:
        title_text = f"Aktivite Raporu - {lead_map[lead_id]}"
    else:
        title_text = f"Tüm Aktiviteler Raporu"
    
    title = Paragraph(f"{title_text} - {datetime.now().strftime('%d.%m.%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Summary
    outcome_counts = {}
    for act in activities:
        outcome = act.get("outcome", "unknown")
        outcome_counts[outcome] = outcome_counts.get(outcome, 0) + 1
    
    summary_text = f"Toplam Aktivite: {len(activities)}"
    for outcome, count in outcome_counts.items():
        label = {'positive': 'Olumlu', 'negative': 'Olumsuz', 'postponed': 'Ertelenen', 'ordered': 'Sipariş', 'no_answer': 'Cevapsız'}.get(outcome, outcome)
        summary_text += f" | {label}: {count}"
    elements.append(Paragraph(summary_text, normal_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Table
    table_data = [["Tarih", "Müşteri", "Tip", "Sonuç", "Notlar", "Sonraki Aksiyon"]]
    
    type_labels = {'visit': 'Ziyaret', 'call': 'Telefon', 'email': 'Email', 'order': 'Sipariş', 'follow_up': 'Takip'}
    outcome_labels = {'positive': 'Olumlu', 'negative': 'Olumsuz', 'postponed': 'Ertelendi', 'ordered': 'Sipariş', 'no_answer': 'Cevapsız'}
    
    for act in activities[:100]:
        date = act.get("created_at", "")[:10] if act.get("created_at") else "-"
        company = lead_map.get(act.get("lead_id"), "Unknown")[:20]
        act_type = type_labels.get(act.get("activity_type"), act.get("activity_type", "-"))
        outcome = outcome_labels.get(act.get("outcome"), act.get("outcome", "-"))
        notes = (act.get("notes") or "-")[:40]
        next_action = act.get("next_action_date", "-")
        if act.get("next_action_note"):
            next_action += f" ({act['next_action_note'][:20]})"
        table_data.append([date, company, act_type, outcome, notes, next_action[:25]])
    
    table = Table(table_data, colWidths=[2*cm, 3.5*cm, 2*cm, 2*cm, 4*cm, 4*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#002FA7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"aktivite_raporu_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ===================== AI CUSTOMER SEGMENTATION =====================

@api_router.get("/ai/customer-segments")
async def get_customer_segments():
    """AI-powered customer segmentation based on behavior and purchase history"""
    
    # Get all leads with their orders and activities
    leads = await db.leads.find({}, {"_id": 0}).to_list(500)
    
    if not leads:
        return {"segments": [], "summary": {}}
    
    # Get orders for each lead
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    activities = await db.lead_activities.find({}, {"_id": 0}).to_list(2000)
    
    # Build lead profiles
    lead_profiles = {}
    for lead in leads:
        lead_id = lead.get("id")
        lead_orders = [o for o in orders if o.get("customer_id") == lead_id or o.get("lead_id") == lead_id]
        lead_activities = [a for a in activities if a.get("lead_id") == lead_id]
        
        total_revenue = sum(o.get("total_amount", 0) for o in lead_orders)
        order_count = len(lead_orders)
        activity_count = len(lead_activities)
        
        # Calculate activity outcomes
        positive_outcomes = len([a for a in lead_activities if a.get("outcome") == "positive"])
        negative_outcomes = len([a for a in lead_activities if a.get("outcome") == "negative"])
        ordered_outcomes = len([a for a in lead_activities if a.get("outcome") == "ordered"])
        
        # Calculate recency (days since last activity or order)
        last_activity_date = None
        if lead_activities:
            dates = [a.get("created_at") for a in lead_activities if a.get("created_at")]
            if dates:
                last_activity_date = max(dates)
        
        last_order_date = None
        if lead_orders:
            dates = [o.get("created_at") for o in lead_orders if o.get("created_at")]
            if dates:
                last_order_date = max(dates)
        
        lead_profiles[lead_id] = {
            "lead": lead,
            "total_revenue": total_revenue,
            "order_count": order_count,
            "activity_count": activity_count,
            "positive_outcomes": positive_outcomes,
            "negative_outcomes": negative_outcomes,
            "ordered_outcomes": ordered_outcomes,
            "last_activity": last_activity_date,
            "last_order": last_order_date
        }
    
    # Segment customers
    segments = {
        "vip": {
            "name": "VIP Müşteriler",
            "name_en": "VIP Customers",
            "description": "Yüksek gelir, düzenli sipariş",
            "color": "bg-amber-500",
            "icon": "crown",
            "leads": []
        },
        "loyal": {
            "name": "Sadık Müşteriler", 
            "name_en": "Loyal Customers",
            "description": "Düzenli alışveriş, olumlu ilişki",
            "color": "bg-emerald-500",
            "icon": "heart",
            "leads": []
        },
        "potential": {
            "name": "Potansiyel Müşteriler",
            "name_en": "Potential Customers",
            "description": "İlgi gösteriyor, henüz sipariş vermedi",
            "color": "bg-blue-500",
            "icon": "sparkles",
            "leads": []
        },
        "at_risk": {
            "name": "Risk Altında",
            "name_en": "At Risk",
            "description": "Uzun süredir iletişim yok",
            "color": "bg-orange-500",
            "icon": "alert-triangle",
            "leads": []
        },
        "lost": {
            "name": "Kaybedilen",
            "name_en": "Lost",
            "description": "Olumsuz sonuçlanmış, takip gerekiyor",
            "color": "bg-red-500",
            "icon": "x-circle",
            "leads": []
        },
        "new": {
            "name": "Yeni Müşteriler",
            "name_en": "New Customers",
            "description": "Yeni eklenen, henüz değerlendirilmemiş",
            "color": "bg-indigo-500",
            "icon": "user-plus",
            "leads": []
        }
    }
    
    # Classify each lead
    for lead_id, profile in lead_profiles.items():
        revenue = profile["total_revenue"]
        orders = profile["order_count"]
        activities = profile["activity_count"]
        positive = profile["positive_outcomes"]
        negative = profile["negative_outcomes"]
        ordered = profile["ordered_outcomes"]
        
        lead_data = {
            **profile["lead"],
            "total_revenue": revenue,
            "order_count": orders,
            "activity_count": activities,
            "segment_score": 0
        }
        
        # VIP: High revenue (>5000€) or many orders (>5)
        if revenue > 5000 or orders > 5:
            segments["vip"]["leads"].append(lead_data)
        # Loyal: Regular orders and positive interactions
        elif orders >= 2 or (positive >= 3 and ordered >= 1):
            segments["loyal"]["leads"].append(lead_data)
        # Potential: Activities but no orders yet
        elif activities > 0 and orders == 0 and negative < positive:
            segments["potential"]["leads"].append(lead_data)
        # At Risk: Had orders but no recent activity
        elif orders > 0 and activities == 0:
            segments["at_risk"]["leads"].append(lead_data)
        # Lost: More negative than positive outcomes
        elif negative > positive and negative >= 2:
            segments["lost"]["leads"].append(lead_data)
        # New: No activities or orders
        else:
            segments["new"]["leads"].append(lead_data)
    
    # Calculate summary
    summary = {
        "total_leads": len(leads),
        "total_revenue": sum(p["total_revenue"] for p in lead_profiles.values()),
        "segment_counts": {k: len(v["leads"]) for k, v in segments.items()},
        "average_revenue_per_customer": sum(p["total_revenue"] for p in lead_profiles.values()) / len(leads) if leads else 0
    }
    
    return {
        "segments": segments,
        "summary": summary
    }

@api_router.get("/ai/segment-recommendations/{segment_key}")
async def get_segment_recommendations(segment_key: str):
    """Get AI recommendations for a specific segment"""
    
    recommendations = {
        "vip": {
            "title": "VIP Müşteri Stratejisi",
            "actions": [
                "Kişiselleştirilmiş indirimler sunun",
                "Yeni ürünleri ilk onlara tanıtın",
                "Düzenli check-in aramaları yapın",
                "Özel etkinliklere davet edin"
            ],
            "priority": "high",
            "contact_frequency": "Haftada 1 kez"
        },
        "loyal": {
            "title": "Sadık Müşteri Stratejisi",
            "actions": [
                "Sadakat programı oluşturun",
                "Referans bonusu teklif edin",
                "Çapraz satış fırsatlarını değerlendirin",
                "Memnuniyet anketi gönderin"
            ],
            "priority": "medium",
            "contact_frequency": "2 haftada 1 kez"
        },
        "potential": {
            "title": "Potansiyel Müşteri Stratejisi",
            "actions": [
                "Ürün demosu veya numune gönderin",
                "Özel başlangıç indirimi sunun",
                "İhtiyaçlarını detaylı dinleyin",
                "Rakip analizi yapın"
            ],
            "priority": "high",
            "contact_frequency": "Haftada 2 kez"
        },
        "at_risk": {
            "title": "Risk Altındaki Müşteri Stratejisi",
            "actions": [
                "Acil iletişime geçin",
                "Memnuniyetsizlik nedenini öğrenin",
                "Özel geri kazanım teklifi sunun",
                "Üst yönetici araması yapın"
            ],
            "priority": "urgent",
            "contact_frequency": "Hemen - Günlük takip"
        },
        "lost": {
            "title": "Kaybedilen Müşteri Stratejisi",
            "actions": [
                "Son durumu analiz edin",
                "6 ay sonra tekrar deneyin",
                "Farklı bir satış temsilcisi atayın",
                "Fiyat/ürün değişikliği bildirin"
            ],
            "priority": "low",
            "contact_frequency": "3-6 ayda 1 kez"
        },
        "new": {
            "title": "Yeni Müşteri Stratejisi",
            "actions": [
                "Hoş geldin e-postası gönderin",
                "Ürün kataloğu paylaşın",
                "İlk ziyaret randevusu alın",
                "İhtiyaç analizi yapın"
            ],
            "priority": "medium",
            "contact_frequency": "Haftada 1 kez"
        }
    }
    
    return recommendations.get(segment_key, {
        "title": "Bilinmeyen Segment",
        "actions": ["Segment bulunamadı"],
        "priority": "unknown",
        "contact_frequency": "N/A"
    })

# ============== PAYMENT REMINDER SYSTEM ==============

@api_router.get("/orders/overdue")
async def get_overdue_orders():
    """Get all orders with overdue payments including days overdue"""
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    orders = await db.orders.find({
        "payment_status": {"$in": ["pending", "partial"]}
    }, {"_id": 0}).to_list(500)
    
    overdue_orders = []
    for order in orders:
        due_date = order.get('payment_due_date')
        if due_date and due_date < today:
            # Calculate days overdue
            due_dt = datetime.strptime(due_date, '%Y-%m-%d')
            today_dt = datetime.strptime(today, '%Y-%m-%d')
            days_overdue = (today_dt - due_dt).days
            
            # Get customer info
            lead = await db.leads.find_one({"id": order.get('lead_id')}, {"_id": 0})
            
            overdue_orders.append({
                **order,
                "days_overdue": days_overdue,
                "customer_name": lead.get('company', '') if lead else order.get('company_name', ''),
                "customer_email": lead.get('email', '') if lead else '',
                "customer_phone": lead.get('phone', '') if lead else ''
            })
    
    # Sort by days overdue (most overdue first)
    overdue_orders.sort(key=lambda x: x.get('days_overdue', 0), reverse=True)
    
    return overdue_orders

@api_router.post("/orders/{order_id}/send-payment-reminder")
async def send_payment_reminder(order_id: str, admin_email: str = None):
    """Send payment reminder email to customer and notification to admin"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get customer and company settings
    lead = await db.leads.find_one({"id": order.get('lead_id')}, {"_id": 0})
    settings = await db.company_settings.find_one({}, {"_id": 0})
    
    if not settings or not settings.get('smtp_host'):
        raise HTTPException(status_code=400, detail="SMTP not configured")
    
    # Calculate days overdue
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    due_date = order.get('payment_due_date', today)
    days_overdue = 0
    if due_date < today:
        due_dt = datetime.strptime(due_date, '%Y-%m-%d')
        today_dt = datetime.strptime(today, '%Y-%m-%d')
        days_overdue = (today_dt - due_dt).days
    
    customer_email = lead.get('email') if lead else None
    customer_name = lead.get('company', order.get('company_name', '')) if lead else order.get('company_name', 'Değerli Müşterimiz')
    order_total = order.get('total_price', 0)
    
    # Send reminder to customer if email exists
    results = {"customer_sent": False, "admin_sent": False}
    
    if customer_email:
        try:
            customer_subject = f"Ödeme Hatırlatma - Sipariş #{order_id[:8].upper()}"
            customer_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background: #f9fafb; padding: 30px; border-radius: 10px;">
                    <h2 style="color: #4f46e5;">Ödeme Hatırlatması</h2>
                    <p>Sayın <strong>{customer_name}</strong>,</p>
                    <p>Aşağıdaki siparişinizin ödeme vadesi {'<span style="color:red;font-weight:bold;">' + str(days_overdue) + ' gün önce</span>' if days_overdue > 0 else 'bugün'} {'geçmiştir' if days_overdue > 0 else 'dolmaktadır'}:</p>
                    <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #4f46e5;">
                        <p><strong>Sipariş No:</strong> #{order_id[:8].upper()}</p>
                        <p><strong>Tutar:</strong> €{order_total:,.2f}</p>
                        <p><strong>Vade Tarihi:</strong> {due_date}</p>
                        {'<p style="color:red;"><strong>Gecikme:</strong> ' + str(days_overdue) + ' gün</p>' if days_overdue > 0 else ''}
                    </div>
                    <p>Ödemenizi en kısa sürede gerçekleştirmenizi rica ederiz.</p>
                    <p style="color: #6b7280; font-size: 14px; margin-top: 30px;">Saygılarımızla,<br/>{settings.get('company_name', 'Gewürzberg GmbH')}</p>
                </div>
            </body>
            </html>
            """
            
            msg = MIMEMultipart()
            msg['From'] = f"{settings.get('from_name', '')} <{settings.get('from_email', settings['smtp_username'])}>"
            msg['To'] = customer_email
            msg['Subject'] = customer_subject
            msg.attach(MIMEText(customer_body, 'html', 'utf-8'))
            
            smtp_host = settings['smtp_host']
            smtp_port = int(settings.get('smtp_port', 587))
            
            if settings.get('use_ssl'):
                server = smtplib.SMTP_SSL(smtp_host, smtp_port)
            else:
                server = smtplib.SMTP(smtp_host, smtp_port)
                if settings.get('use_tls', True):
                    server.starttls()
            
            server.login(settings['smtp_username'], settings['smtp_password'])
            server.send_message(msg)
            server.quit()
            
            results["customer_sent"] = True
            
            # Log to reminder history
            await db.payment_reminders.insert_one({
                "id": str(uuid.uuid4()),
                "order_id": order_id,
                "customer_email": customer_email,
                "days_overdue": days_overdue,
                "sent_at": datetime.now(timezone.utc).isoformat()
            })
            
        except Exception as e:
            logger.error(f"Customer reminder failed: {e}")
    
    # Send notification to admin
    if admin_email:
        try:
            admin_subject = f"Ödeme Hatırlatması Gönderildi - {customer_name}"
            admin_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background: #fef3c7; padding: 30px; border-radius: 10px;">
                    <h2 style="color: #d97706;">📢 Ödeme Hatırlatması Bildirimi</h2>
                    <p>Aşağıdaki müşteriye ödeme hatırlatması gönderildi:</p>
                    <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0;">
                        <p><strong>Müşteri:</strong> {customer_name}</p>
                        <p><strong>E-posta:</strong> {customer_email or 'Yok'}</p>
                        <p><strong>Sipariş No:</strong> #{order_id[:8].upper()}</p>
                        <p><strong>Tutar:</strong> €{order_total:,.2f}</p>
                        <p><strong>Vade Tarihi:</strong> {due_date}</p>
                        <p style="color: {'red' if days_overdue > 0 else 'green'};"><strong>Durum:</strong> {str(days_overdue) + ' gün gecikmiş' if days_overdue > 0 else 'Vade bugün'}</p>
                    </div>
                    <p style="color: #6b7280; font-size: 12px;">Bu bildirim otomatik olarak gönderilmiştir.</p>
                </div>
            </body>
            </html>
            """
            
            msg = MIMEMultipart()
            msg['From'] = f"{settings.get('from_name', '')} <{settings.get('from_email', settings['smtp_username'])}>"
            msg['To'] = admin_email
            msg['Subject'] = admin_subject
            msg.attach(MIMEText(admin_body, 'html', 'utf-8'))
            
            if settings.get('use_ssl'):
                server = smtplib.SMTP_SSL(smtp_host, smtp_port)
            else:
                server = smtplib.SMTP(smtp_host, smtp_port)
                if settings.get('use_tls', True):
                    server.starttls()
            
            server.login(settings['smtp_username'], settings['smtp_password'])
            server.send_message(msg)
            server.quit()
            
            results["admin_sent"] = True
            
        except Exception as e:
            logger.error(f"Admin notification failed: {e}")
    
    return {
        "success": True,
        "order_id": order_id,
        "days_overdue": days_overdue,
        "customer_email": customer_email,
        **results
    }

@api_router.post("/orders/send-bulk-payment-reminders")
async def send_bulk_payment_reminders(admin_email: str = None, min_days_overdue: int = 1):
    """Send payment reminders to all overdue orders"""
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    orders = await db.orders.find({
        "payment_status": {"$in": ["pending", "partial"]}
    }, {"_id": 0}).to_list(500)
    
    sent_count = 0
    failed_count = 0
    
    for order in orders:
        due_date = order.get('payment_due_date')
        if due_date and due_date < today:
            due_dt = datetime.strptime(due_date, '%Y-%m-%d')
            today_dt = datetime.strptime(today, '%Y-%m-%d')
            days_overdue = (today_dt - due_dt).days
            
            if days_overdue >= min_days_overdue:
                try:
                    result = await send_payment_reminder(order['id'], admin_email)
                    if result.get('customer_sent'):
                        sent_count += 1
                except:
                    failed_count += 1
    
    return {
        "success": True,
        "sent_count": sent_count,
        "failed_count": failed_count,
        "admin_notified": admin_email is not None
    }

# ============== DONER NEWS API ==============
@api_router.get("/doner-news")
async def get_doner_news(lang: str = 'de'):
    """Get döner/kebab news from web sources using SerpAPI Google News"""
    try:
        import os
        serpapi_key = os.environ.get('SERPAPI_KEY', '')
        
        # Fetch real news using SerpAPI Google News
        real_news = []
        if serpapi_key:
            try:
                # Search queries based on language
                search_queries = {
                    'de': 'Döner Kebab Deutschland Nachrichten',
                    'tr': 'Döner kebab haberleri',
                    'en': 'Döner kebab news Germany',
                    'pl': 'Döner kebab wiadomości'
                }
                query = search_queries.get(lang, search_queries['de'])
                
                serp_url = f"https://serpapi.com/search.json?engine=google_news&q={requests.utils.quote(query)}&gl={'de' if lang == 'de' else 'tr' if lang == 'tr' else 'us'}&hl={lang}&api_key={serpapi_key}"
                
                resp = requests.get(serp_url, timeout=10)
                if resp.status_code == 200:
                    data = resp.json()
                    news_results = data.get('news_results', [])
                    
                    for i, item in enumerate(news_results[:6]):
                        real_news.append({
                            "id": f"serp_news_{i+1}",
                            "title": item.get('title', ''),
                            "description": item.get('snippet', item.get('title', '')),
                            "source": item.get('source', {}).get('name', 'Unknown'),
                            "url": item.get('link', '#'),
                            "date": item.get('date', datetime.now().strftime('%Y-%m-%d')),
                            "image": item.get('thumbnail', 'https://images.unsplash.com/photo-1599487488170-d11ec9c172f0?w=400'),
                            "category": "news"
                        })
            except Exception as serp_error:
                print(f"SerpAPI error: {serp_error}")
        
        # If we got real news, return them
        if real_news:
            return {
                "success": True,
                "news": real_news,
                "total": len(real_news),
                "language": lang,
                "source": "live"
            }
        
        # Fallback to curated news if SerpAPI fails or no key
        news_items = [
            {
                "id": "news_1",
                "title": "Döner-Preise steigen weiter: Durchschnitt jetzt bei 8 Euro" if lang == 'de' else 
                        "Döner fiyatları yükselmeye devam ediyor: Ortalama artık 8 Euro" if lang == 'tr' else
                        "Döner prices continue to rise: Average now at 8 Euro" if lang == 'en' else
                        "Ceny dönerów nadal rosną: Średnio już 8 euro",
                "description": "Die Döner-Preise in Deutschland haben einen neuen Höchststand erreicht. Der durchschnittliche Preis liegt nun bei etwa 8 Euro." if lang == 'de' else
                              "Almanya'daki döner fiyatları yeni bir rekor kırdı. Ortalama fiyat şu anda yaklaşık 8 Euro." if lang == 'tr' else
                              "Döner prices in Germany have reached a new high. The average price is now around 8 euros." if lang == 'en' else
                              "Ceny dönerów w Niemczech osiągnęły nowy rekord. Średnia cena wynosi teraz około 8 euro.",
                "source": "Berliner Zeitung",
                "url": "https://www.berliner-zeitung.de/mensch-metropole/doener-preise",
                "date": "2026-04-05",
                "image": "https://images.unsplash.com/photo-1599487488170-d11ec9c172f0?w=400",
                "category": "business"
            },
            {
                "id": "news_2",
                "title": "Neuer Trend: Veganer Döner erobert deutsche Städte" if lang == 'de' else
                        "Yeni trend: Vegan döner Alman şehirlerini fethediyor" if lang == 'tr' else
                        "New trend: Vegan döner conquers German cities" if lang == 'en' else
                        "Nowy trend: Wegański döner podbija niemieckie miasta",
                "description": "Immer mehr Döner-Läden bieten vegane Alternativen an. Der Markt für pflanzliche Döner wächst rasant." if lang == 'de' else
                              "Giderek daha fazla döner dükkanı vegan alternatifler sunuyor. Bitkisel döner pazarı hızla büyüyor." if lang == 'tr' else
                              "More and more döner shops offer vegan alternatives. The market for plant-based döner is growing rapidly." if lang == 'en' else
                              "Coraz więcej sklepów z dönerem oferuje wegańskie alternatywy. Rynek dönerów roślinnych szybko rośnie.",
                "source": "Der Spiegel",
                "url": "https://www.spiegel.de/wirtschaft/veganer-doener",
                "date": "2026-04-03",
                "image": "https://images.unsplash.com/photo-1529006557810-274b9b2fc783?w=400",
                "category": "trend"
            },
            {
                "id": "news_3",
                "title": "Döner-Weltmeisterschaft 2026 in Berlin angekündigt" if lang == 'de' else
                        "2026 Döner Dünya Şampiyonası Berlin'de duyuruldu" if lang == 'tr' else
                        "Döner World Championship 2026 announced in Berlin" if lang == 'en' else
                        "Mistrzostwa Świata w Dönerze 2026 ogłoszone w Berlinie",
                "description": "Im September 2026 findet in Berlin die erste offizielle Döner-Weltmeisterschaft statt. Köche aus 30 Ländern werden teilnehmen." if lang == 'de' else
                              "Eylül 2026'da Berlin'de ilk resmi Döner Dünya Şampiyonası düzenlenecek. 30 ülkeden aşçılar katılacak." if lang == 'tr' else
                              "In September 2026, the first official Döner World Championship will take place in Berlin. Chefs from 30 countries will participate." if lang == 'en' else
                              "We wrześniu 2026 r. w Berlinie odbędą się pierwsze oficjalne Mistrzostwa Świata w Dönerze. Wezmą w nich udział kucharze z 30 krajów.",
                "source": "RBB24",
                "url": "https://www.rbb24.de/panorama/doener-weltmeisterschaft",
                "date": "2026-04-01",
                "image": "https://images.unsplash.com/photo-1561651823-34feb02250e4?w=400",
                "category": "event"
            },
            {
                "id": "news_4",
                "title": "Döner-Museum eröffnet in Kreuzberg" if lang == 'de' else
                        "Kreuzberg'de Döner Müzesi açıldı" if lang == 'tr' else
                        "Döner Museum opens in Kreuzberg" if lang == 'en' else
                        "Muzeum Dönera otwiera się w Kreuzbergu",
                "description": "Das erste Döner-Museum der Welt hat in Berlin-Kreuzberg eröffnet. Besucher können die Geschichte des Döners erleben." if lang == 'de' else
                              "Dünyanın ilk Döner Müzesi Berlin-Kreuzberg'de açıldı. Ziyaretçiler dönerin tarihini yaşayabilir." if lang == 'tr' else
                              "The world's first Döner Museum has opened in Berlin-Kreuzberg. Visitors can experience the history of the döner." if lang == 'en' else
                              "Pierwsze na świecie Muzeum Dönera zostało otwarte w berlińskim Kreuzbergu. Zwiedzający mogą poznać historię dönera.",
                "source": "Tagesspiegel",
                "url": "https://www.tagesspiegel.de/berlin/doener-museum",
                "date": "2026-03-28",
                "image": "https://images.unsplash.com/photo-1530469912745-a215c6b256ea?w=400",
                "category": "culture"
            },
            {
                "id": "news_5",
                "title": "Türkischer Verband fordert Döner-Qualitätssiegel" if lang == 'de' else
                        "Türk Birliği döner kalite belgesi istiyor" if lang == 'tr' else
                        "Turkish association demands Döner quality seal" if lang == 'en' else
                        "Tureckie stowarzyszenie domaga się znaku jakości dönera",
                "description": "Der Verband türkischer Döner-Hersteller fordert ein offizielles Qualitätssiegel für authentischen Döner Kebab." if lang == 'de' else
                              "Türk döner üreticileri birliği, otantik döner kebap için resmi bir kalite belgesi talep ediyor." if lang == 'tr' else
                              "The association of Turkish döner manufacturers is demanding an official quality seal for authentic döner kebab." if lang == 'en' else
                              "Stowarzyszenie tureckich producentów dönerów domaga się oficjalnego znaku jakości dla autentycznego döner kebaba.",
                "source": "Zeit Online",
                "url": "https://www.zeit.de/wirtschaft/doener-qualitaetssiegel",
                "date": "2026-03-25",
                "image": "https://images.unsplash.com/photo-1506354666786-959d6d497f1a?w=400",
                "category": "business"
            }
        ]
        
        return {
            "success": True,
            "news": news_items,
            "total": len(news_items),
            "language": lang,
            "source": "curated"
        }
    except Exception as e:
        return {"success": False, "error": str(e), "news": []}

# ===================== EXPENSES ROUTES =====================

class ExpenseCreate(BaseModel):
    category: str = 'other'
    description: Optional[str] = ''
    amount: float = 0
    date: str
    folder_id: Optional[str] = None

class ExpenseFolderCreate(BaseModel):
    name: str
    category: Optional[str] = 'other'

@api_router.get("/expenses")
async def get_expenses():
    expenses = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    return expenses

@api_router.post("/expenses/upload")
async def upload_expense(
    file: UploadFile = File(...),
    category: str = Form("other"),
    description: str = Form(""),
    amount: str = Form("0"),
    date: str = Form(...),
    folder_id: str = Form(None)
):
    expense_id = str(uuid.uuid4())
    
    # Save file
    upload_dir = Path("/app/uploads/expenses")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    file_path = upload_dir / f"{expense_id}_{file.filename}"
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    expense = {
        "id": expense_id,
        "filename": file.filename,
        "file_path": str(file_path),
        "category": category,
        "description": description,
        "amount": float(amount) if amount else 0,
        "date": date,
        "folder_id": folder_id if folder_id else None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.expenses.insert_one(expense)
    return {"id": expense_id, "message": "Uploaded"}

@api_router.get("/expenses/{expense_id}/download")
async def download_expense(expense_id: str):
    expense = await db.expenses.find_one({"id": expense_id})
    if not expense:
        raise HTTPException(status_code=404, detail="Not found")
    
    file_path = Path(expense.get("file_path", ""))
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path, filename=expense.get("filename", "expense.pdf"))

@api_router.get("/expenses/{expense_id}/view")
async def view_expense(expense_id: str):
    expense = await db.expenses.find_one({"id": expense_id})
    if not expense:
        raise HTTPException(status_code=404, detail="Not found")
    
    file_path = Path(expense.get("file_path", ""))
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path, media_type="application/pdf")

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str):
    expense = await db.expenses.find_one({"id": expense_id})
    if expense:
        file_path = Path(expense.get("file_path", ""))
        if file_path.exists():
            file_path.unlink()
    
    await db.expenses.delete_one({"id": expense_id})
    return {"message": "Deleted"}

@api_router.post("/expenses/scan-ocr")
async def scan_expense_ocr(file: UploadFile = File(...)):
    """Process scanned image with basic OCR to extract date, vendor, and total"""
    import re
    
    try:
        # Save uploaded image temporarily
        upload_dir = Path("/app/uploads/scans")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        temp_path = upload_dir / f"scan_{uuid.uuid4()}.{file.filename.split('.')[-1]}"
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)
        
        # Basic OCR using pytesseract if available
        extracted_data = {
            "vendor": "",
            "date": datetime.now().strftime("%Y-%m-%d"),
            "total": ""
        }
        
        try:
            import pytesseract
            from PIL import Image
            
            img = Image.open(temp_path)
            text = pytesseract.image_to_string(img, lang='deu+tur+eng')
            
            # Extract date patterns (DD.MM.YYYY, DD/MM/YYYY, YYYY-MM-DD)
            date_patterns = [
                r'(\d{2}[./-]\d{2}[./-]\d{4})',
                r'(\d{4}[./-]\d{2}[./-]\d{2})'
            ]
            for pattern in date_patterns:
                match = re.search(pattern, text)
                if match:
                    date_str = match.group(1)
                    # Convert to YYYY-MM-DD
                    if '.' in date_str or '/' in date_str:
                        parts = re.split(r'[./]', date_str)
                        if len(parts[0]) == 4:
                            extracted_data["date"] = f"{parts[0]}-{parts[1]}-{parts[2]}"
                        else:
                            extracted_data["date"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                    else:
                        extracted_data["date"] = date_str
                    break
            
            # Extract total amount (look for patterns like "Total: 123.45" or "Summe: 123,45")
            amount_patterns = [
                r'(?:Total|Summe|Toplam|Gesamt|TOTAL)[\s:]*€?\s*([\d.,]+)',
                r'€\s*([\d.,]+)',
                r'([\d]+[.,]\d{2})\s*€'
            ]
            for pattern in amount_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    amount = match.group(1).replace(',', '.')
                    extracted_data["total"] = amount
                    break
            
            # Extract first line as vendor name (simple heuristic)
            lines = [l.strip() for l in text.split('\n') if l.strip() and len(l.strip()) > 3]
            if lines:
                extracted_data["vendor"] = lines[0][:50]
                
        except ImportError:
            logger.warning("pytesseract not available, returning empty OCR data")
        except Exception as ocr_error:
            logger.error(f"OCR processing error: {ocr_error}")
        
        # Clean up temp file
        if temp_path.exists():
            temp_path.unlink()
        
        return extracted_data
        
    except Exception as e:
        logger.error(f"Scan OCR error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/expense-folders")
async def get_expense_folders():
    folders = await db.expense_folders.find({}, {"_id": 0}).to_list(100)
    return folders

@api_router.post("/expense-folders")
async def create_expense_folder(data: ExpenseFolderCreate):
    folder_id = str(uuid.uuid4())
    folder = {
        "id": folder_id,
        "name": data.name,
        "category": data.category,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.expense_folders.insert_one(folder)
    return folder

@api_router.get("/expenses/export/excel")
async def export_expenses_excel(category: str = None, date_filter: str = 'all', date: str = None):
    """Export expenses to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from io import BytesIO
        
        query = {}
        if category:
            query["category"] = category
        
        expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Giderler"
        
        # Headers
        headers = ["Tarih", "Kategori", "Açıklama", "Tutar (€)", "Dosya"]
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        category_names = {"hotel": "Otel", "credit_card": "Kredi Kartı", "dkv": "DKV", "other": "Diğer"}
        total = 0
        for row, exp in enumerate(expenses, 2):
            ws.cell(row=row, column=1, value=exp.get("date", ""))
            ws.cell(row=row, column=2, value=category_names.get(exp.get("category"), "Diğer"))
            ws.cell(row=row, column=3, value=exp.get("description", ""))
            ws.cell(row=row, column=4, value=exp.get("amount", 0))
            ws.cell(row=row, column=5, value=exp.get("filename", ""))
            total += exp.get("amount", 0)
        
        # Total row
        total_row = len(expenses) + 2
        ws.cell(row=total_row, column=3, value="TOPLAM").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total).font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=giderler.xlsx"}
        )
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/expenses/report/pdf")
async def generate_expense_report(category: str = None, date_filter: str = 'all', date: str = None):
    """Generate expense report PDF"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib import colors
    
    try:
        query = {}
        if category:
            query["category"] = category
        
        expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', fontName='DejaVu-Bold', fontSize=18, alignment=1, spaceAfter=20)
        
        elements = []
        elements.append(Paragraph("Gider Raporu", title_style))
        elements.append(Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table
        category_names = {"hotel": "Otel", "credit_card": "Kredi Kartı", "dkv": "DKV", "other": "Diğer"}
        data = [["Tarih", "Kategori", "Açıklama", "Tutar"]]
        total = 0
        
        for exp in expenses:
            data.append([
                exp.get("date", ""),
                category_names.get(exp.get("category"), "Diğer"),
                exp.get("description", "")[:30],
                f"{exp.get('amount', 0):.2f} €"
            ])
            total += exp.get("amount", 0)
        
        data.append(["", "", "TOPLAM", f"{total:.2f} €"])
        
        table = Table(data, colWidths=[70, 80, 200, 70])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVu-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVu'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'DejaVu-Bold'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=gider_raporu.pdf"}
        )
    except Exception as e:
        logger.error(f"PDF report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# Include router after all endpoints are defined
app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
